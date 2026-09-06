from django.conf import settings
from django.test import Client, RequestFactory, TestCase
from django.urls import reverse
from django.contrib.auth.hashers import check_password, make_password
from openpyxl import load_workbook
from django.core.files.uploadedfile import SimpleUploadedFile
from django.utils import timezone
from datetime import date, datetime, timedelta
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
import hashlib
import json
import os
import uuid
from urllib.parse import parse_qs, urlparse
from unittest.mock import patch

PRINCIPAL_DEFAULT_CUSTOM_ID = "PRN-DEFAULT"
PRINCIPAL_DEFAULT_PASSWORD = "Principal123"

from pypdf import PdfReader
from reportlab.pdfgen import canvas

from .forms import AdminPracticeMaterialForm
from .models import Material, User, Section, Assessment, Notification, Course, Note, LiveAssessmentSession, School, SchoolCalendar, CalendarEvent, StoryReadingProgress
from .reading_stt import (
    ReadingMatcher,
    align_story_transcript,
    story_word_states_from_results,
    analyze_reading,
    analyze_sentence_reading,
    language_code_for,
    synthesize_read_aloud_audio,
    target_phrase_hints,
    target_aware_syllable_stitching,
    syllable_context_metrics,
    strict_syllabic_word_match,
    transcribe_audio_bytes_v2_chirp3,
    v1_model_for_language,
    word_numbers_in_transcript,
)
from .hunt_scoring import classify_speech, normalize_speech, stars_for_points


def test_section_create(**kwargs):
    school = kwargs.pop("school", None)
    if school is None:
        suffix = uuid.uuid4().hex.upper()
        school = School.objects.create(name=f"Fixture School {suffix}", code=f"FIXTURE-{suffix}")
    return Section.objects.create(school=school, **kwargs)
from .management.commands.seed_official_crla_assessments import OFFICIAL_CRLA_CONTENT
from .views import _active_school_calendar, _apply_progression_unlock_override, _aral_eligible_classification, _create_notification, _notify_principals, _material_response_payload, _fallback_material_items_from_text, _build_material_items_from_ocr_layout, _build_image_upload_debug_info, _adapted_reading_level_from_attempts, _adapted_reading_level_label, _assessment_fluency_score, _assessment_score_payload, _build_reading_report_pdf, _derive_dashboard_greeting_name, _display_reading_level, _build_latest_reading_level_payload, _primary_school, _save_admin_practice_material, _selected_school_calendar, _sync_assessment_workflow_state, _official_crla_assessment_labels, _official_assessment_availability_for_student
from .weekly_digest import send_weekly_digest
from .scoring import build_assessment_score_payload


class WordDecodingLanguageTests(TestCase):
    def setUp(self):
        self.material = Material.objects.create(
            title='English decoding',
            item_type='word',
            language='English',
            content_json={
                'activity_slug': 'word_decoding',
                # Simulate a material whose template data predates a language edit.
                'language': 'Filipino',
                'items': [{'word': 'cat'}],
            },
        )

    def test_page_uses_the_saved_material_language(self):
        response = self.client.get(reverse('word_decoding_page'), {'material_id': self.material.id})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '"language":"English"')
        self.assertContains(response, 'word_decoding_google_tts.js')
        self.assertContains(response, 'data-word-decoding-language="English"')

    @patch('pabasa_app.views.transcribe_audio_bytes_with_model', return_value=('cat', 'chirp_3', ''))
    def test_google_speech_uses_the_saved_material_language(self, transcribe):
        session = self.client.session
        session['user_id'] = 1
        session.save()

        response = self.client.post(
            reverse('word_decoding_transcribe_api'),
            {
                'audio': SimpleUploadedFile('word.webm', b'audio', content_type='audio/webm'),
                'material_id': self.material.id,
                'target_word': 'cat',
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(transcribe.call_args.kwargs['language_code'], 'en-PH')


class LetterSoundCorrespondenceLanguageTests(TestCase):
    def test_page_uses_saved_material_language_and_male_tts_bridge(self):
        material = Material.objects.create(
            title='English correspondence',
            item_type='word',
            language='English',
            content_json={
                'activity_slug': 'letter-sound-correspondence',
                'language': 'Filipino',
                'items': [{'letter': 'A'}],
            },
        )

        response = self.client.get(reverse('letter_sound_correspondence_page'), {'material_id': material.id})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '"language":"English"')
        self.assertContains(response, 'letter_sound_correspondence_google_tts.js')
        self.assertContains(response, 'data-letter-correspondence-language="English"')


class ClassMaterialsApiTests(TestCase):
    def test_get_class_materials_groups_vowel_materials_under_vowel_bucket(self):
        teacher = User.objects.create(
            custom_id="TCHR-1001",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher1001@example.com",
            password_hash=make_password("teacher-password"),
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Class A",
            class_code="CLS-A1001",
            subject="Reading",
            is_active=True,
        )
        material = Material.objects.create(
            title="Vowel Drill",
            item_type="vowel",
            content_text="a\ne",
            content_json={"items": ["a", "e"]},
            type="assessment",
            status="published",
            section=section,
            teacher=teacher,
            is_active=True,
        )

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session.save()

        response = self.client.get(reverse("get_class_materials"), {"class_code": section.class_code})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn("vowel", data["materials"])
        self.assertEqual(len(data["materials"]["vowel"]), 1)
        self.assertEqual(data["materials"]["vowel"][0]["id"], f"material-{material.id}")
        self.assertEqual(data["materials"]["vowel"][0]["item_type"], "vowel")

    def test_get_class_materials_hides_bosy_official_card_for_eligible_completed_student(self):
        teacher = User.objects.create(
            custom_id="TCHR-1002",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher1002@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-1002",
            role="student",
            first_name="Sam",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="student1002@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-1002",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        Material.objects.create(
            title="Teacher Intervention Text",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            type="assessment",
            status="published",
            section=section,
            teacher=teacher,
            is_active=True,
        )
        self._login_student(student)

        with patch("pabasa_app.views._official_crla_assessment_phase", return_value="pretest"):
            response = self.client.get(reverse("get_class_materials"), {"class_code": section.class_code})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["official_assessments"], [])
        word_titles = [item["title"] for item in data["materials"]["word"]]
        self.assertIn("Teacher Intervention Text", word_titles)
        self.assertFalse(any("BoSY" in title for title in word_titles))


class SchoolCalendarSelectionTests(TestCase):
    def test_selected_school_calendar_respects_teacher_school(self):
        school_a = School.objects.create(name="School A", code="SCH-A")
        school_b = School.objects.create(name="School B", code="SCH-B")

        calendar_a = SchoolCalendar.objects.create(school_year="2026-2027", current_term=1, is_active=True)
        calendar_b = SchoolCalendar.objects.create(school_year="2027-2028", current_term=1, is_active=True)

        CalendarEvent.objects.create(
            school_calendar=calendar_a,
            term=1,
            title="School A Opening Block",
            event_type="school_opening",
            start_date=date(2027, 6, 1),
            end_date=date(2027, 6, 7),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar_a,
            term=1,
            title="School A Closing Block",
            event_type="school_closing",
            start_date=date(2027, 9, 30),
            end_date=date(2027, 10, 6),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar_b,
            term=1,
            title="School B Opening Block",
            event_type="school_opening",
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 7),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar_b,
            term=1,
            title="School B Closing Block",
            event_type="school_closing",
            start_date=date(2026, 9, 30),
            end_date=date(2026, 10, 6),
        )

        teacher = User.objects.create(
            custom_id="TCHR-SCHOOL-A",
            role="teacher",
            first_name="Ava",
            last_name="Teacher",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher-school-a@example.com",
            password_hash=make_password("password"),
            school_record=school_a,
        )
        other_teacher = User.objects.create(
            custom_id="TCHR-SCHOOL-B",
            role="teacher",
            first_name="Ben",
            last_name="Teacher",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=1991,
            email="teacher-school-b@example.com",
            password_hash=make_password("password"),
            school_record=school_b,
        )

        Section.objects.create(
            school=school_a,
            school_calendar=calendar_a,
            class_code="SCHA-001",
            class_name="School A Class",
            teacher=teacher,
            is_active=True,
            subject="Reading",
            grade_level="Grade 2",
            section="A",
        )
        Section.objects.create(
            school=school_b,
            school_calendar=calendar_b,
            class_code="SCHB-001",
            class_name="School B Class",
            teacher=other_teacher,
            is_active=True,
            subject="Reading",
            grade_level="Grade 2",
            section="B",
        )

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = "teacher"
        session.save()

        request = RequestFactory().get("/dashboard/teacher/")
        request.session = session

        with patch("pabasa_app.views.date", wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 7, 15)
            selected = _selected_school_calendar(request)

        self.assertEqual(selected, calendar_a)


class SchoolCalendarAdminTests(TestCase):
    def _login_admin(self, admin):
        session = self.client.session
        session["user_id"] = admin.id
        session["user_role"] = admin.role
        session["custom_id"] = admin.custom_id
        session["first_name"] = admin.first_name
        session["last_name"] = admin.last_name
        session["email"] = admin.email
        session.save()

    def test_admin_school_calendar_calculates_assessment_week_from_term_blocks(self):
        admin = User.objects.create(
            custom_id="ADM-1001",
            role="admin",
            first_name="Ava",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1985,
            email="admin1001@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(
            school_year="2026-2027",
            current_term=1,
            is_active=True,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Opening Block",
            event_type="school_opening",
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="End-of-Term Block",
            event_type="school_closing",
            start_date=date(2027, 5, 1),
            end_date=date(2027, 5, 31),
        )
        self._login_admin(admin)

        response = self.client.post(reverse("admin_school_calendar"), {
            "action": "save_term_blocks",
            "calendar_id": calendar.id,
            "term_1_opening": "2026-08-01",
            "term_1_closing": "2027-05-01",
        })

        self.assertEqual(response.status_code, 302)
        event = CalendarEvent.objects.get(school_calendar=calendar, event_type="pre_assessment")
        self.assertEqual(event.title, "Pre-Assessment Week")
        self.assertEqual(event.start_date, date(2026, 8, 3))
        self.assertEqual(event.end_date, date(2026, 8, 7))

        page = self.client.get(reverse("admin_school_calendar"), {"calendar_id": calendar.id})
        self.assertEqual(page.status_code, 200)
        self.assertContains(page, "Pre-Assessment Week")
        self.assertContains(page, '"event_type": "pre_assessment"')

    def test_admin_school_calendar_renders_year_view_and_soft_yellow_opening_block(self):
        admin = User.objects.create(
            custom_id="ADM-1002",
            role="admin",
            first_name="Ari",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=1985,
            email="admin1002@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(school_year="2026-2027", current_term=1, is_active=True)
        self._login_admin(admin)

        response = self.client.get(reverse("admin_school_calendar"), {"calendar_id": calendar.id})

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, "Year View")
        self.assertContains(response, 'id="calendarYearBtn">Year</button>')
        self.assertContains(response, 'id="calendarMonthBtn">Month</button>')
        self.assertContains(response, 'id="calendarWeekBtn">Week</button>')
        self.assertContains(response, "school_opening: '#facc15'")

    def test_admin_school_calendar_saves_progressive_term_blocks(self):
        admin = User.objects.create(
            custom_id="ADM-1003",
            role="admin",
            first_name="Ari",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=1985,
            email="admin1003@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(school_year="2026-2027", current_term=1, is_active=True)
        self._login_admin(admin)

        response = self.client.post(
            reverse("admin_school_calendar"),
            {
                "action": "save_term_blocks",
                "calendar_id": calendar.id,
                "term_1_opening": "2026-06-01",
                "term_1_closing": "2026-08-31",
                "term_2_opening": "2026-09-14",
                "term_2_closing": "2026-11-30",
            },
        )

        self.assertEqual(response.status_code, 302)
        self.assertEqual(
            CalendarEvent.objects.filter(school_calendar=calendar, event_type="school_opening").count(), 2
        )
        self.assertEqual(
            CalendarEvent.objects.filter(school_calendar=calendar, event_type="school_closing").count(), 2
        )
        self.assertEqual(
            CalendarEvent.objects.get(school_calendar=calendar, term=2, event_type="school_closing").start_date,
            date(2026, 11, 30),
        )
        self.assertEqual(
            CalendarEvent.objects.get(school_calendar=calendar, term=1, event_type="school_opening").end_date,
            date(2026, 6, 5),
        )
        self.assertEqual(
            CalendarEvent.objects.get(school_calendar=calendar, term=2, event_type="school_closing").end_date,
            date(2026, 12, 11),
        )
        self.assertEqual(
            CalendarEvent.objects.get(school_calendar=calendar, term=1, event_type="pre_assessment").start_date,
            date(2026, 6, 8),
        )
        self.assertEqual(
            CalendarEvent.objects.get(school_calendar=calendar, term=2, event_type="midline_assessment").start_date,
            date(2026, 9, 21),
        )
        holiday = CalendarEvent.objects.get(school_calendar=calendar, title="Independence Day", start_date=date(2026, 6, 12))
        self.assertEqual(holiday.event_type, "holiday")
        self.assertEqual(holiday.end_date, date(2026, 6, 12))

    def test_term_assessment_uses_the_next_monday_after_a_midweek_opening(self):
        admin = User.objects.create(
            custom_id="ADM-1003", role="admin", first_name="Mia", last_name="Admin", middle_initial="", suffix="",
            sex="female", birth_month=1, birth_day=2, birth_year=1985, email="admin1003@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(school_year="2027-2028", current_term=1, is_active=True)
        self._login_admin(admin)

        response = self.client.post(reverse("admin_school_calendar"), {
            "action": "save_term_blocks", "calendar_id": calendar.id,
            "term_1_opening": "2027-06-02", "term_1_closing": "2027-08-31",
        })

        self.assertEqual(response.status_code, 302)
        assessment = CalendarEvent.objects.get(school_calendar=calendar, term=1, event_type="pre_assessment")
        self.assertEqual(assessment.start_date, date(2027, 6, 7))
        self.assertEqual(assessment.end_date, date(2027, 6, 11))

    def test_admin_can_start_the_next_school_year_automatically(self):
        admin = User.objects.create(
            custom_id="ADM-1004", role="admin", first_name="Noel", last_name="Admin", middle_initial="", suffix="",
            sex="male", birth_month=1, birth_day=2, birth_year=1985, email="admin1004@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(school_year="2025-2026", current_term=1, is_active=True)
        CalendarEvent.objects.create(
            school_calendar=calendar, term=1, title="Opening Block", event_type="school_opening",
            start_date=date(2025, 6, 2), end_date=date(2025, 6, 6),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar, term=3, title="End-of-Term Block", event_type="school_closing",
            start_date=date(2026, 5, 18), end_date=date(2026, 5, 29),
        )
        self._login_admin(admin)

        response = self.client.post(reverse("admin_school_calendar"), {"action": "start_next_school_year"}, HTTP_ACCEPT="application/json")

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["school_year"], "2026-2027")
        self.assertTrue(SchoolCalendar.objects.get(school_year="2026-2027").is_active)

    def test_admin_can_save_multiple_suspension_dates(self):
        admin = User.objects.create(
            custom_id="ADM-1005", role="admin", first_name="Sam", last_name="Admin", middle_initial="", suffix="",
            sex="male", birth_month=1, birth_day=2, birth_year=1985, email="admin1005@example.com",
            password_hash=make_password("admin-password"),
        )
        calendar = SchoolCalendar.objects.create(school_year="2026-2027", current_term=1, is_active=True)
        self._login_admin(admin)

        response = self.client.post(reverse("admin_school_calendar"), {
            "action": "save_suspensions", "calendar_id": calendar.id,
            "suspension_dates": "2026-06-15,2026-06-16",
        })

        self.assertEqual(response.status_code, 302)
        self.assertEqual(CalendarEvent.objects.filter(school_calendar=calendar, title="Class Suspension").count(), 2)
        suspension = CalendarEvent.objects.filter(school_calendar=calendar, title="Class Suspension").first()
        response = self.client.post(reverse("admin_school_calendar"), {
            "action": "delete_suspension", "calendar_id": calendar.id, "event_id": suspension.id,
        })
        self.assertEqual(response.status_code, 302)
        self.assertEqual(CalendarEvent.objects.filter(school_calendar=calendar, title="Class Suspension").count(), 1)

    def _create_official_crla_calendar(self, *, pre_start, pre_end, post_start, post_end):
        calendar = SchoolCalendar.objects.create(
            school_year='2026-2027',
            current_term=1,
            is_active=True,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Start of Classes',
            event_type='start_of_classes',
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 1),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='End of Classes',
            event_type='end_of_classes',
            start_date=date(2027, 5, 31),
            end_date=date(2027, 5, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Opening Block',
            event_type='school_opening',
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Closing Block',
            event_type='school_closing',
            start_date=date(2026, 8, 31),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Pre-Assessment Week',
            event_type='pre_assessment',
            start_date=pre_start,
            end_date=pre_end,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Post-Assessment Week',
            event_type='post_assessment',
            start_date=post_start,
            end_date=post_end,
        )
        return calendar

    def _login_student(self, student):
        session = self.client.session
        session['user_id'] = student.id
        session['user_role'] = 'student'
        session['custom_id'] = student.custom_id
        session['first_name'] = student.first_name
        session['last_name'] = student.last_name
        session['email'] = student.email
        session.save()

    def test_get_class_materials_keeps_teacher_materials_and_excludes_official_crla(self):
        teacher = User.objects.create(
            custom_id="TCHR-1001",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher1001@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-1001",
            role="student",
            first_name="Sam",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="student1001@example.com",
            password_hash=make_password("student-password"),
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-1001",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 5),
            post_start=date(2026, 8, 8),
            post_end=date(2026, 8, 11),
        )

        bosy_payload = OFFICIAL_CRLA_CONTENT["bosy_crla_pretest"]
        bosy, _ = Material.objects.get_or_create(
            system_assessment_key="bosy_crla_pretest",
            defaults={
                "title": bosy_payload["title"],
                "item_type": "paragraph",
                "content_text": "\n".join([
                    *bosy_payload["words"],
                    *bosy_payload["sentences"],
                    *[p["content"] for p in bosy_payload["passages"]],
                ]),
                "content_json": {
                    "assessment_key": "bosy_crla_pretest",
                    "language": "Filipino",
                    "words": bosy_payload["words"],
                    "sentences": bosy_payload["sentences"],
                    "passages": bosy_payload["passages"],
                    "items": (
                        [{"type": "word", "text": word} for word in bosy_payload["words"]]
                        + [{"type": "sentence", "text": sentence} for sentence in bosy_payload["sentences"]]
                        + [{"type": "paragraph", "text": passage["content"], "title": passage["title"]} for passage in bosy_payload["passages"]]
                    ),
                },
                "assessment_kind": "crla",
                "assessment_set": "crla",
                "type": "assessment",
                "status": "published",
                "student_access": True,
                "section": None,
                "teacher": teacher,
                "is_active": True,
                "is_official_reading": True,
                "is_system_owned": True,
                "system_assessment_period": bosy_payload["period"],
                "system_assessment_phase": bosy_payload["phase"],
                "language": "Filipino",
                "source_type": "shared",
                "code": bosy_payload["code"],
            },
        )
        eosy_payload = OFFICIAL_CRLA_CONTENT["eosy_crla_posttest"]
        eosy, _ = Material.objects.get_or_create(
            system_assessment_key="eosy_crla_posttest",
            defaults={
                "title": eosy_payload["title"],
                "item_type": "paragraph",
                "content_text": "\n".join([
                    *eosy_payload["words"],
                    *eosy_payload["sentences"],
                    *[p["content"] for p in eosy_payload["passages"]],
                ]),
                "content_json": {
                    "assessment_key": "eosy_crla_posttest",
                    "language": "Filipino",
                    "words": eosy_payload["words"],
                    "sentences": eosy_payload["sentences"],
                    "passages": eosy_payload["passages"],
                    "items": (
                        [{"type": "word", "text": word} for word in eosy_payload["words"]]
                        + [{"type": "sentence", "text": sentence} for sentence in eosy_payload["sentences"]]
                        + [{"type": "paragraph", "text": passage["content"], "title": passage["title"]} for passage in eosy_payload["passages"]]
                    ),
                },
                "assessment_kind": "crla",
                "assessment_set": "crla",
                "type": "assessment",
                "status": "published",
                "student_access": True,
                "section": None,
                "teacher": teacher,
                "is_active": True,
                "is_official_reading": True,
                "is_system_owned": True,
                "system_assessment_period": eosy_payload["period"],
                "system_assessment_phase": eosy_payload["phase"],
                "language": "Filipino",
                "source_type": "shared",
                "code": eosy_payload["code"],
            },
        )
        teacher_material = Material.objects.create(
            title="Teacher Reading",
            item_type="word",
            content_text="alpha",
            content_json={"items": ["alpha"], "language": "English"},
            assessment_kind="regular",
            assessment_set="word",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
        )
        session = self.client.session
        session["user_id"] = student.id
        session["user_role"] = "student"
        session["custom_id"] = student.custom_id
        session["first_name"] = student.first_name
        session["last_name"] = student.last_name
        session["email"] = student.email
        session.save()

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 4)
            response = self.client.get(reverse('get_class_materials'), {'class_code': section.class_code})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertIn('materials', data)
        self.assertIn('official_assessments', data)
        self.assertEqual(sorted(data['materials'].keys()), ['paragraph', 'sentence', 'vowel', 'word'])

        word_items = data['materials']['word']
        self.assertTrue(any(item['raw_id'] == teacher_material.id for item in word_items))
        self.assertTrue(any(item['title'] == 'Teacher Reading' for item in word_items))
        teacher_item = next(item for item in word_items if item['raw_id'] == teacher_material.id)
        self.assertEqual(teacher_item['language'], 'English')
        self.assertEqual(teacher_item['source_type'], 'personal')
        self.assertEqual(teacher_item['assessment_kind'], 'regular')
        self.assertFalse(teacher_item['is_official_reading'])
        self.assertFalse(teacher_item['is_system_owned'])
        self.assertFalse(any(item['raw_id'] == eosy.id for item in word_items))
        self.assertFalse(any(item['raw_id'] == bosy.id for item in word_items))

        official_items = data['official_assessments']
        self.assertEqual(official_items, [])

        eligible_student = User.objects.create(
            custom_id="STU-1002",
            role="student",
            first_name="Ellie",
            last_name="Eligible",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="eligible1002@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"reader_classification": "Low Emerging Readers", "aral_eligible": True}},
        )
        ineligible_student = User.objects.create(
            custom_id="STU-1003",
            role="student",
            first_name="Ivy",
            last_name="Ineligible",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="ineligible1003@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"reader_classification": "Readers at Grade Level", "aral_eligible": False}},
        )
        section.add_student(eligible_student)
        section.add_student(ineligible_student)
        self._login_student(eligible_student)
        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 9)
        eligible_response = self.client.get(reverse('get_class_materials'), {'class_code': section.class_code})
        self.assertEqual(eligible_response.status_code, 200)
        self.assertIn('official_assessments', eligible_response.json())
        eligible_word_items = eligible_response.json()['materials']['word']
        self.assertTrue(any(item['title'] == 'Teacher Reading' for item in eligible_word_items))
        self.assertFalse(any(item['raw_id'] == bosy.id for item in eligible_word_items))
        self.assertFalse(any(item['raw_id'] == eosy.id for item in eligible_word_items))
        eligible_official_items = eligible_response.json()['official_assessments']
        self.assertTrue(any(item['title'] == eosy_payload['title'] for item in eligible_official_items))
        self.assertTrue(any(item['raw_id'] == eosy.id for item in eligible_official_items))
        self.assertFalse(any(item['raw_id'] == bosy.id for item in eligible_official_items))

        self._login_student(ineligible_student)
        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 9)
            ineligible_response = self.client.get(reverse('get_class_materials'), {'class_code': section.class_code})
        self.assertEqual(ineligible_response.status_code, 200)
        self.assertIn('official_assessments', ineligible_response.json())
        ineligible_word_items = ineligible_response.json()['materials']['word']
        self.assertFalse(any(item['title'] == eosy_payload['title'] for item in ineligible_word_items))
        self.assertFalse(any(item['raw_id'] == eosy.id for item in ineligible_word_items))
        self.assertFalse(any(item['raw_id'] == bosy.id for item in ineligible_word_items))
        self.assertTrue(any(item['title'] == 'Teacher Reading' for item in ineligible_word_items))
        self.assertEqual(ineligible_response.json()['official_assessments'], [])


class DashboardAchievementBadgeTests(TestCase):
    def test_dashboard_template_contains_practice_star_achievement(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "dashboard.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn("'practice-star'", content)
        self.assertIn("Complete every level in Free Mode, Color Mode, and Hunt Mode", content)
        self.assertIn("pabasa_practice_progress_v1", content)
        self.assertNotIn("pabasa_practice_sessions_completed", content)
        self.assertNotIn(">= 10", content)


class DashboardGreetingNameTests(TestCase):
    def test_uses_first_name_when_available(self):
        self.assertEqual(_derive_dashboard_greeting_name(first_name="Jamie", full_name="Jamie Reader"), "Jamie")

    def test_uses_first_word_of_full_name_for_legacy_accounts(self):
        self.assertEqual(_derive_dashboard_greeting_name(first_name="", full_name="Maria Clara Dela Cruz"), "Maria")

    def test_falls_back_to_student_when_no_name_data_exists(self):
        self.assertEqual(_derive_dashboard_greeting_name(first_name="", full_name=""), "Student")


class AssessmentPageTemplateTests(TestCase):
    def test_assessment_page_includes_vowel_material_support(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "assessment.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn('const materialTypes = ["word", "sentence", "paragraph", "vowel"];', content)
        self.assertIn('vowel: "{% url \'reading_vowel_page\' %}"', content)

    def test_assessment_page_start_link_passes_crla_payload_to_reader(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "assessment.html"
        content = template_path.read_text(encoding="utf-8")

        official_card_source = content.split('const buildOfficialAssessmentCard =', 1)[1].split('const officialAssessmentsHtml =', 1)[0]
        self.assertIn('official_assessment_id=', official_card_source)
        self.assertIn('official_assessment_data=', official_card_source)
        self.assertIn('crla_fresh=1', official_card_source)

    def test_custom_material_card_does_not_add_crla_query_parameters(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "assessment.html"
        content = template_path.read_text(encoding="utf-8")
        custom_card_source = content.split('const materialSourceType =', 1)[1].split('renderedRows.push', 1)[0]

        self.assertNotIn('official_assessment_id=', custom_card_source)
        self.assertNotIn('official_assessment_data=', custom_card_source)
        self.assertNotIn('crla_fresh', custom_card_source)

    def test_story_reading_card_uses_dedicated_route_from_template_identity(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "assessment.html"
        content = template_path.read_text(encoding="utf-8")
        custom_card_source = content.split('const materialSourceType =', 1)[1].split('renderedRows.push', 1)[0]

        self.assertIn("{% url 'story_reading_page' %}", content)
        self.assertIn('const isStoryReading = [m.template_title, m.template_type', custom_card_source)
        self.assertIn('`${storyReadingUrl}?id=${encodeURIComponent(m.id)}', custom_card_source)


class ReadingLaunchClassificationTests(TestCase):
    def _login_student(self):
        student = User.objects.create(
            custom_id=f'STU-LAUNCH-{User.objects.count()}', role='student',
            first_name='Sam', last_name='Reader', sex='Male',
            birth_month=1, birth_day=1, birth_year=2015,
            email=f'student-launch-{User.objects.count()}@example.com',
            password_hash=make_password('password'),
        )
        session = self.client.session
        session['user_id'] = student.id
        session['user_role'] = student.role
        session['custom_id'] = student.custom_id
        session.save()
        return student

    def _sentence_material(self, *, code, source_type='template', assessment_kind='regular', is_official=False):
        return Material.objects.create(
            title='Sentence Reading Practice', code=code, item_type='sentence',
            content_text='The moon is bright tonight.',
            content_json={
                'template_title': 'Sentence Reading Practice',
                'template_lesson': 'Sentence Reading',
                'template_type': 'Sentence Reading Practice',
                'language': 'English',
                'items': [{'sentence': 'The moon is bright tonight.'}],
            },
            language='English', type='assessment', source_type=source_type,
            assessment_kind=assessment_kind, is_official_reading=is_official,
            is_system_owned=is_official, student_access=True,
        )

    def test_reading_ui_sentence_route_never_renders_sentence_bot_for_official_crla(self):
        self._login_student()
        material = self._sentence_material(
            code='CRLA-SENTENCE-ROUTE', source_type='shared',
            assessment_kind='crla', is_official=True,
        )

        response = self.client.get(reverse('reading_sentence_page'), {
            'id': f'material-{material.id}',
            'official_assessment_id': f'material-{material.id}',
            'item_type': 'sentence',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'pabasa_app/reading_sentence_page.html')
        self.assertTemplateNotUsed(response, 'pabasa_app/sentence_bot_page.html')
        self.assertNotContains(response, 'pabasa_app/js/sentence_bot.js')

    def test_reading_ui_sentence_route_never_renders_sentence_bot_from_item_type(self):
        self._login_student()
        material = self._sentence_material(code='REGULAR-SENTENCE-ROUTE', source_type='template')

        response = self.client.get(reverse('reading_sentence_page'), {
            'id': f'material-{material.id}', 'item_type': 'sentence',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'pabasa_app/reading_sentence_page.html')
        self.assertTemplateNotUsed(response, 'pabasa_app/sentence_bot_page.html')

    def test_sentence_bot_activity_route_renders_valid_regular_template(self):
        self._login_student()
        material = self._sentence_material(code='VALID-SENTENCE-BOT')

        response = self.client.get(reverse('sentence_bot_page'), {
            'id': f'material-{material.id}',
        })

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'pabasa_app/sentence_bot_page.html')
        self.assertContains(response, 'pabasa_app/js/sentence_bot.js')
        self.assertContains(response, 'sentence-bot-root')

    def test_sentence_bot_activity_route_rejects_official_crla_material(self):
        self._login_student()
        material = self._sentence_material(
            code='CRLA-ON-BOT-ROUTE', source_type='template',
            assessment_kind='crla', is_official=True,
        )

        response = self.client.get(reverse('sentence_bot_page'), {
            'id': f'material-{material.id}',
        })

        self.assertRedirects(response, reverse('assessment'), fetch_redirect_response=False)

    def test_sentence_bot_activity_route_rejects_non_template_material(self):
        self._login_student()
        material = self._sentence_material(code='PERSONAL-ON-BOT-ROUTE', source_type='personal')

        response = self.client.get(reverse('sentence_bot_page'), {
            'id': f'material-{material.id}',
        })

        self.assertRedirects(response, reverse('assessment'), fetch_redirect_response=False)

    def test_sentence_bot_completion_persists_and_locks_reopening(self):
        student = self._login_student()
        material = self._sentence_material(code='PERSISTED-SENTENCE-BOT')

        active_response = self.client.get(reverse('sentence_bot_page'), {
            'id': f'material-{material.id}',
        })
        self.assertEqual(active_response.status_code, 200)
        self.assertContains(active_response, 'assessment_reader.js')

        completion_response = self.client.post(
            reverse('sentence_bot_complete'),
            data=json.dumps({
                'material_id': f'material-{material.id}',
                'activity_type': 'sentence_bot',
                'scores': {'duration_seconds': 12, 'speech_recognition_used': True},
            }),
            content_type='application/json',
        )
        self.assertEqual(completion_response.status_code, 200)
        self.assertTrue(completion_response.json()['success'])
        result = material.assessment_results.get(student=student, attempt_status='completed')
        self.assertEqual(result.correct_items, 1)
        self.assertEqual(result.items_completed, 1)
        self.assertEqual(result.total_score, 100)

        reopened_response = self.client.get(reverse('sentence_bot_page'), {
            'id': f'material-{material.id}', 'item_type': 'sentence', 'retake': '1',
        })
        self.assertEqual(reopened_response.status_code, 200)
        self.assertTrue(reopened_response.context['sentence_bot_completed'])
        self.assertNotContains(reopened_response, 'assessment_reader.js')
        self.assertContains(reopened_response, '"completed":true')
        self.assertContains(reopened_response, '"correct_sentences":1')
        self.assertContains(reopened_response, '"total_sentences":1')

        duplicate_response = self.client.post(
            reverse('sentence_bot_complete'),
            data=json.dumps({'material_id': material.id, 'activity_type': 'sentence_bot'}),
            content_type='application/json',
        )
        self.assertEqual(duplicate_response.status_code, 200)
        self.assertEqual(material.assessment_results.filter(student=student, attempt_status='completed').count(), 1)

    def test_sentence_bot_skipped_sentences_do_not_become_perfect_score(self):
        student = self._login_student()
        material = self._sentence_material(code='SKIPPED-SENTENCE-BOT')
        material.content_json['items'] = [
            {'sentence': 'One sentence.'},
            {'sentence': 'Two sentence.'},
            {'sentence': 'Three sentence.'},
            {'sentence': 'Four sentence.'},
            {'sentence': 'Five sentence.'},
            {'sentence': 'Six sentence.'},
            {'sentence': 'Seven sentence.'},
        ]
        material.save(update_fields=['content_json'])

        response = self.client.post(
            reverse('sentence_bot_complete'),
            data=json.dumps({
                'material_id': f'material-{material.id}',
                'activity_type': 'sentence_bot',
                'scores': {'correct_items': 0, 'duration_seconds': 5},
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['correct_items'], 0)
        result = material.assessment_results.get(student=student, attempt_status='completed')
        self.assertEqual(result.items_completed, 7)
        self.assertEqual(result.correct_items, 0)
        self.assertEqual(result.accuracy, 0)
        self.assertEqual(result.total_score, 0)

    def test_sentence_bot_completion_persists_perfect_and_partial_sentence_counts(self):
        student = self._login_student()
        sentences = [{'sentence': f'Sentence {index}.'} for index in range(1, 8)]

        for correct_items in (7, 5):
            material = self._sentence_material(code=f'RESULT-{correct_items}-SENTENCE-BOT')
            material.content_json['items'] = sentences
            material.save(update_fields=['content_json'])
            response = self.client.post(
                reverse('sentence_bot_complete'),
                data=json.dumps({
                    'material_id': f'material-{material.id}',
                    'activity_type': 'sentence_bot',
                    'scores': {'correct_items': correct_items, 'duration_seconds': 8},
                }),
                content_type='application/json',
            )

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.json()['correct_items'], correct_items)
            reopened = self.client.get(reverse('sentence_bot_page'), {'id': f'material-{material.id}'})
            completion = json.loads(reopened.context['sentence_bot_completion_json'])
            self.assertEqual(completion['correct_sentences'], correct_items)
            self.assertEqual(completion['total_sentences'], 7)
            self.assertNotEqual(completion['correct_sentences'], 7 if correct_items == 5 else 0)

    def test_sentence_bot_listening_control_is_not_a_navigation_control(self):
        sentence_bot_source = (Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'sentence_bot.js').read_text(encoding='utf-8')
        reader_source = (Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js').read_text(encoding='utf-8')

        self.assertIn('advanceButton.textContent = isFinalSentence ? "Finish" : "Skip";', sentence_bot_source)
        self.assertIn('advanceButton.addEventListener("click", () => nextButton?.click());', sentence_bot_source)
        self.assertNotIn('payload.correct_items = state.total;', sentence_bot_source)
        self.assertIn('if (isSentenceBot) return;', reader_source)
        self.assertIn(': (isSentenceBot ? false : (!isRecording || (onLastItem && isLastPage)))', reader_source)

    def test_sentence_bot_completion_message_has_perfect_and_partial_result_variants(self):
        sentence_bot_source = (Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'sentence_bot.js').read_text(encoding='utf-8')

        self.assertIn('function completionMessageFor(correct, total)', sentence_bot_source)
        self.assertIn('Woohoo! Pippo knows all the sentences now!', sentence_bot_source)
        self.assertIn('Great job! Pippo learned ${correct} out of ${total} sentences!', sentence_bot_source)
        self.assertIn('Nice work! Pippo learned ${correct} out of ${total} sentences!', sentence_bot_source)
        self.assertIn('Keep practicing! Pippo is still learning these sentences.', sentence_bot_source)
        self.assertIn('completionMessageFor(state.learned, state.total)', sentence_bot_source)

    def test_custom_material_launch_is_redirected_to_clean_non_crla_url(self):
        material = Material.objects.create(
            title='English Reading',
            code='VFPN-455',
            item_type='word',
            content_text='Cat\nDog\nDove\nZebra\nMouse',
            content_json={'items': ['Cat', 'Dog', 'Dove', 'Zebra', 'Mouse'], 'language': 'English'},
            language='English',
            type='assessment',
            source_type='personal',
            assessment_kind='regular',
            is_official_reading=False,
            is_system_owned=False,
            student_access=True,
        )
        response = self.client.get(reverse('reading_word_page'), {
            'test': material.title,
            'code': material.code,
            'id': f'material-{material.id}',
            'official_assessment_id': '999',
            'official_assessment_data': json.dumps({'official_title': 'Wrong CRLA title'}),
            'crla_fresh': '1',
            'content': material.content_text,
            'item_type': material.item_type,
            'language': material.language,
        })

        self.assertEqual(response.status_code, 302)
        query = parse_qs(urlparse(response.url).query)
        self.assertEqual(query['test'], ['English Reading'])
        self.assertEqual(query['code'], ['VFPN-455'])
        self.assertEqual(query['id'], [f'material-{material.id}'])
        self.assertEqual(query['item_type'], ['word'])
        self.assertEqual(query['language'], ['English'])
        self.assertNotIn('official_assessment_id', query)
        self.assertNotIn('official_assessment_data', query)
        self.assertNotIn('crla_fresh', query)

    def test_genuine_official_material_keeps_crla_query_parameters(self):
        self._login_student()
        material = Material.objects.create(
            title='Official CRLA', code='CRLA-TEST', item_type='word',
            content_text='Word', content_json={'items': ['Word']},
            type='assessment', assessment_kind='crla', is_official_reading=True,
            is_system_owned=True, student_access=True,
        )
        response = self.client.get(reverse('reading_word_page'), {
            'id': f'material-{material.id}',
            'official_assessment_id': str(material.id),
            'official_assessment_data': '{}',
            'crla_fresh': '1',
        })

        self.assertEqual(response.status_code, 200)

    def test_rendered_custom_reader_uses_persisted_material_and_has_no_official_payload(self):
        self._login_student()
        material = Material.objects.create(
            title='English Reading', code='CUSTOM-INTERNAL', item_type='word',
            content_text='Cat\nDog\nDove\nZebra\nMouse',
            content_json={'items': ['Cat', 'Dog', 'Dove', 'Zebra', 'Mouse'], 'language': 'English'},
            language='English', type='assessment', source_type='personal',
            assessment_kind='regular', is_official_reading=False,
            is_system_owned=False, student_access=True,
        )

        response = self.client.get(reverse('reading_word_page'), {
            'test': 'Wrong URL title',
            'code': 'VFPN-455',
            'id': f'material-{material.id}',
            'content': 'Wrong URL content',
            'item_type': 'word',
            'language': 'English',
        })

        self.assertEqual(response.status_code, 200)
        payload = response.context['custom_material_launch_data']
        self.assertEqual(payload['id'], f'material-{material.id}')
        self.assertEqual(payload['title'], 'English Reading')
        self.assertEqual(payload['code'], 'VFPN-455')
        self.assertEqual(payload['language'], 'English')
        self.assertEqual(payload['item_type'], 'word')
        self.assertEqual(payload['items'], ['Cat', 'Dog', 'Dove', 'Zebra', 'Mouse'])
        self.assertEqual(payload['content'], 'Cat\nDog\nDove\nZebra\nMouse')
        self.assertFalse(payload['is_official_reading'])
        self.assertNotIn('crla_official_assessment_data', response.context)
        self.assertContains(response, 'window.__PABASA_OFFICIAL_ASSESSMENT__ = null;')
        self.assertContains(response, 'window.__PABASA_CUSTOM_MATERIAL__ = {')
        self.assertContains(response, 'Cat')

    def test_reader_javascript_never_uses_cached_crla_items_for_custom_launch(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        custom_branch = content.split('if (!isOfficialAssessmentLaunch) {', 1)[1].split('// Prioritize the specific class code', 1)[0]

        self.assertIn('customMaterialData || liveContent', custom_branch)
        self.assertNotIn("sessionStorage.getItem('pabasa_crla_assessment_items')", custom_branch)

    def test_phrase_listening_button_resets_microphone_without_completing_activity(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        reset_branch = content.split('const resetPhraseListening = () => {', 1)[1].split('const startReading = () => {', 1)[0]
        start_branch = content.split('const startReading = () => {', 1)[1].split("if (mode === 'phrase')", 1)[0]

        self.assertIn('if (mode !== "phrase" || !isRecording) return false;', reset_branch)
        self.assertIn('itemResultVersion += 1;', reset_branch)
        self.assertIn('isRecording = false;', reset_branch)
        self.assertIn('stopSpeechRecognition();', reset_branch)
        self.assertIn('syncPhraseMicrophoneButton();', reset_branch)
        self.assertNotIn('showCompletion(', reset_branch)
        self.assertNotIn('phraseReadingCompleted', reset_branch)
        self.assertLess(start_branch.index('if (resetPhraseListening()) return;'), start_branch.index('if (isSpeechResponsePending()) return;'))

    def test_official_crla_start_button_cannot_manually_finish_an_active_reading_attempt(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        start_reading = content.split('const startReading = () => {', 1)[1].split('startAssessmentTimer();', 1)[0]
        update_ui = content.split('function updateUI() {', 1)[1].split('function animateCurrentItem()', 1)[0]
        speech_controls = content.split('function updateSpeechProcessingControls() {', 1)[1].split('function resetSyllableStitching()', 1)[0]

        self.assertIn('if (isCrla) return;', start_reading)
        self.assertLess(start_reading.index('if (isCrla) return;'), start_reading.index('stopReading();'))
        self.assertIn('spinner-border spinner-border-sm', update_ui)
        self.assertIn('<span>Reading...</span>', update_ui)
        self.assertIn('btnStartReading.disabled = true;', update_ui)
        self.assertIn('btnStartReading.disabled = false;', update_ui)
        self.assertNotIn('Finish Reading', update_ui)
        self.assertIn('button === btnStartReading', speech_controls)
        self.assertIn('&& isCrla', speech_controls)
        self.assertIn('button.disabled = true;', speech_controls)

    def test_official_crla_final_reading_next_persists_a_skip_before_existing_completion(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        final_skip = content.split('async function skipFinalCrlaReadingItem() {', 1)[1].split('function goToNextPageOrItem()', 1)[0]
        navigation = content.split('function updateAssessmentNavigationButtons() {', 1)[1].split('function updateSpeechProcessingControls()', 1)[0]
        next_handler = content.split('nextBtn?.addEventListener("click", async () => {', 1)[1].split('function isInteractiveElement', 1)[0]
        comprehension = content.split('crlaQuestionNextBtn?.addEventListener("click", async () => {', 1)[1].split('crlaQuestionReadAloudBtn?.addEventListener', 1)[0]

        self.assertIn('const isCrlaReading = isCrla', navigation)
        self.assertIn('!isCrlaReading && (!isRecording || (onLastItem && isLastPage))', navigation)
        self.assertIn('itemScores[currentIndex] = {', final_skip)
        self.assertIn('skipped: true,', final_skip)
        self.assertIn('await persistLockedItemResult(currentIndex);', final_skip)
        self.assertLess(final_skip.index('await persistLockedItemResult(currentIndex);'), final_skip.index('showCompletion(true);'))
        self.assertIn('await stopReading({ allowIdleStoryCompletion: true });', final_skip)
        self.assertIn('storyMiscueCount += readableWordCount(getCurrentDisplayText());', final_skip)
        self.assertEqual(next_handler.count('await skipFinalCrlaReadingItem()'), 2)
        self.assertIn('await completeCRLASpokenAttempt("", questionIndex);', comprehension)
        self.assertIn('await finishCRLAComprehension();', comprehension)

    def test_crla_miscue_branch_advances_local_paragraph_cursor(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        miscue_branch = content.split('if (!itemLocked[currentIndex] && (', 1)[1].split('if (Number(data.matched || 0) > 0)', 1)[0]

        self.assertIn('data.word_syllable_ranges[activeWordIndex][1]', miscue_branch)
        self.assertIn('currentSyllableIndex = resolvedWordEnd;', miscue_branch)
        self.assertNotIn('transitionToItem(', miscue_branch)
        context_guard = content.split('function isCurrentSpeechContext(context) {', 1)[1].split('async function sendAudioChunk', 1)[0]
        self.assertIn('context.syllableIndex === currentSyllableIndex', context_guard)
        update_ui = content.split('function updateUI() {', 1)[1].split('function animateCurrentItem()', 1)[0]
        self.assertNotIn('currentSyllableIndex = 0;', update_ui)
        self.assertIn('evaluatedParagraphWordIndex(data.word_syllable_ranges, context?.syllableIndex)', content.replace('\n', ' '))
        self.assertIn('paragraphWordResults[activeWordIndex] !== "miscue"', content)
        self.assertIn('paragraphWordResults[readableWordIndex] === "miscue"', content)
        self.assertNotIn('|| readableWordIndex === wrongWordIndex', miscue_branch)

    def test_crla_future_word_results_do_not_create_paragraph_visual_state(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordParagraphWordResult(', 1)[1].split('function consumeSequentialParagraphWordResults', 1)[0]

        self.assertIn('expectedIndex === activeWordIndex', recorder)
        self.assertNotIn('wordResults.forEach', recorder)
        self.assertNotIn('expectedIndex > throughIndex', recorder)

    def test_crla_paragraph_miscue_state_is_sticky_across_redraws(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordParagraphWordResult(', 1)[1].split('function evaluatedParagraphWordIndex', 1)[0]

        self.assertIn('paragraphWordResults[activeWordIndex] !== "miscue"', recorder)
        self.assertIn('paragraphWordResults[activeWordIndex] = status;', recorder)
        self.assertIn('paragraphWordResults[readableWordIndex] === "miscue"', content)

    def test_crla_later_correct_result_cannot_overwrite_paragraph_miscue(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordParagraphWordResult(', 1)[1].split('function evaluatedParagraphWordIndex', 1)[0]

        self.assertIn('if (paragraphWordResults[activeWordIndex] !== "miscue")', recorder)
        self.assertNotIn('paragraphWordResults[activeWordIndex] = "correct"', recorder)

    def test_story_finalized_metrics_use_total_words_minus_miscues(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        helper = content.split('function calculateFinalizedStoryMetrics(', 1)[1].split('function correctWordsRead()', 1)[0]

        self.assertIn('const wordsRead = Math.max(0, totalWords - miscues);', helper)
        self.assertIn('wordsRead / totalWords', helper)
        self.assertIn('wordsRead / Math.max(durationSeconds / 60, 1 / 60)', helper)
        self.assertNotIn('correctWordsRead()', helper)
        self.assertNotIn('120', helper)

        for scenario, total_words, miscues, expected_words_read, expected_accuracy, expected_wpm in (
            ('zero miscues', 4, 0, 4, 100, 4),
            ('one substitution', 4, 1, 3, 75, 3),
            ('one omission', 4, 1, 3, 75, 3),
            ('one insertion', 4, 1, 3, 75, 3),
            ('multiple miscues', 6, 3, 3, 50, 3),
        ):
            self.assertEqual(max(0, total_words - miscues), expected_words_read)
            self.assertEqual(round((expected_words_read / total_words) * 100), expected_accuracy, scenario)
            self.assertEqual(expected_words_read / 1, expected_wpm, scenario)

    def test_story_finalized_metrics_are_used_at_both_story_persistence_points(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        completion = content.split('async function showStoryCompletionScreen()', 1)[1].split('function hideStoryCompletionScreen()', 1)[0]
        stop_reading = content.split('const stopReading = async () => {', 1)[1].split('btnStartReading?.addEventListener', 1)[0]

        self.assertIn('calculateFinalizedStoryMetrics(', completion)
        self.assertIn('persistedStoryState?.duration_seconds', completion)
        self.assertIn('words_read: storyMetrics.wordsRead', completion)
        self.assertIn('wpm: storyMetrics.wpm', completion)
        self.assertNotIn('words_read: correctWordsRead()', completion)

        self.assertIn('calculateFinalizedStoryMetrics(', stop_reading)
        self.assertIn('words_read: storyMetrics.wordsRead', stop_reading)
        self.assertIn('story_read_percent: storyMetrics.accuracy', stop_reading)
        self.assertIn('wpm: storyMetrics.wpm', stop_reading)
        self.assertNotIn('const wordsRead = correctWordsRead();', stop_reading)

    def test_story_insertion_note_uses_deduplicated_alignment_without_target_highlight(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordStoryAlignmentMiscues(', 1)[1].split('function storyAlignmentHasInsertionMiscue(', 1)[0]
        insertion_detector = content.split('function storyAlignmentHasInsertionMiscue(', 1)[1].split('function resetStorySegmentState(', 1)[0]
        callback = content.split('const storyMiscueEvent = recordStoryAlignmentMiscues(data, context);', 1)[1].split('const itemCorrectWords =', 1)[0]

        self.assertIn('if (storyMiscueResponseKeys.has(responseKey)) return { accepted: false };', recorder)
        self.assertIn('return { accepted: true }', recorder)
        self.assertIn('type || "").toLowerCase() === "insertion"', insertion_detector)
        self.assertIn('recognizedWords.some((_, index) => !representedRecognizedIndexes.has(index))', insertion_detector)
        self.assertIn('storyMiscueEvent.accepted', callback)
        self.assertIn('Extra or repeated word detected.', content)
        self.assertNotIn('paragraphWordResults["insertion"]', content)

    def test_story_self_correction_is_strict_and_story_scoped(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordStoryAlignmentMiscues(', 1)[1].split('function storyAlignmentHasInsertionMiscue(', 1)[0]
        same_callback = content.split('function storySameCallbackSelfCorrection(', 1)[1].split('function storySelfCorrectionCandidate(', 1)[0]
        immediate = content.split('function isImmediateStorySelfCorrection(', 1)[1].split('function recordStoryAlignmentMiscues(', 1)[0]
        stop_reading = content.split('const stopReading = async () => {', 1)[1].split('btnStartReading?.addEventListener', 1)[0]

        self.assertIn('let pendingStorySelfCorrection = null;', content)
        self.assertIn('alignmentMiscues !== 1 || substitutions.length !== 1', content)
        self.assertIn('recognizedWords.length === 1', immediate)
        self.assertIn('Number(context?.syllableIndex) === candidate.expectedWordEnd', immediate)
        self.assertIn('wrongWord === expectedWord', same_callback)
        self.assertIn('extraIndexes.size !== 1', same_callback)
        self.assertIn('return { accepted: true, deferred: true }', recorder)
        self.assertIn('return { accepted: true, selfCorrection: candidate }', recorder)
        self.assertIn('paragraphWordResults[correctedIndex] = "correct";', content)
        self.assertIn('commitPendingStorySelfCorrection();', stop_reading)
        self.assertIn('!storyMiscueEvent.selfCorrection', content)

    def test_story_self_correction_commits_before_a_nonqualifying_zero_miscue_callback(self):
        """An intervening accepted callback must invalidate the one-callback correction window."""
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        recorder = content.split('function recordStoryAlignmentMiscues(', 1)[1].split('function storyAlignmentHasInsertionMiscue(', 1)[0]
        pending_branch = recorder.split('if (hadPendingCandidate) {', 1)[1].split('if (!hasAlignmentMiscues) return { accepted: false };', 1)[0]

        self.assertIn('const hasAlignmentMiscues = Number.isFinite(alignmentMiscues) && alignmentMiscues > 0;', recorder)
        self.assertIn('if (storyMiscueResponseKeys.has(responseKey)) return { accepted: false };', pending_branch)
        self.assertIn('if (isImmediateStorySelfCorrection(data, context, candidate))', pending_branch)
        self.assertIn('commitPendingStorySelfCorrection();', pending_branch)
        self.assertIn('if (!hasAlignmentMiscues) return { accepted: true };', pending_branch)
        self.assertLess(
            pending_branch.index('if (isImmediateStorySelfCorrection(data, context, candidate))'),
            pending_branch.index('commitPendingStorySelfCorrection();'),
        )
        self.assertLess(
            pending_branch.index('commitPendingStorySelfCorrection();'),
            pending_branch.index('if (!hasAlignmentMiscues) return { accepted: true };'),
        )

    def test_shared_scoring_and_non_story_paths_remain_unchanged(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        shared_scores = content.split('function calculateScores()', 1)[1].split('function calculateFinalizedStoryMetrics(', 1)[0]

        self.assertIn('const matchedWords = correctWordsRead();', shared_scores)
        self.assertIn('accuracy: targetWordCount && speechRecognitionUsed', shared_scores)
        self.assertIn('wpm: Math.round((matchedWords', shared_scores)

    def test_story_segment_transitions_reset_segment_local_reading_state(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        resetter = content.split('function resetStorySegmentState(', 1)[1].split('function currentSpeechContext()', 1)[0]

        self.assertIn('currentSyllableIndex = 0;', resetter)
        self.assertIn('paragraphWordResults = {};', resetter)
        self.assertIn('resetSyllableStitching();', resetter)

        automatic = content.split("traceEndSession('handleSpeechResult.storySegmentComplete'", 1)[1].split('if (currentIndex >= items.length - 1)', 1)[0]
        navigation = content.split('prevBtn?.addEventListener("click"', 1)[1].split("if (currentStoryState === \"story_comprehension\")", 1)[0]
        self.assertEqual(automatic.count('resetStorySegmentState('), 1)
        self.assertEqual(navigation.count('resetStorySegmentState('), 2)

    def test_story_segment_diagnostics_are_toggleable_and_read_only(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        debug_setup = content.split('const storyDebugStorageKey', 1)[1].split('const liveContent', 1)[0]

        self.assertIn('urlParams.get("story_debug") === "1"', debug_setup)
        self.assertIn('window.setCrlaStoryDebug = function (enabled)', debug_setup)
        self.assertIn('console.log("[CRLA_STORY_DEBUG]", detail);', debug_setup)
        for event in ('segment_initialization', 'segment_transition', 'stt_callback', 'highlight_state', 'segment_completion'):
            self.assertIn(f'event: "{event}"', content)

    def test_story_miscue_branch_consumes_only_contiguous_paragraph_results(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        consumer = content.split('function consumeSequentialParagraphWordResults(', 1)[1].split('function evaluatedParagraphWordIndex', 1)[0]
        miscue_branch = content.split('if (!itemLocked[currentIndex] && (', 1)[1].split('if (Number(data.matched || 0) > 0)', 1)[0]

        self.assertIn('for (let targetIndex = activeWordIndex; resultsByTargetIndex.has(targetIndex); targetIndex += 1)', consumer)
        self.assertIn('paragraphWordResults[targetIndex] = status;', consumer)
        self.assertIn('currentSyllableIndex = finalWordEnd;', consumer)
        self.assertIn('consumeSequentialParagraphWordResults(', miscue_branch)
        self.assertIn('Number(data.current_word_index ?? activeWordIndex)', miscue_branch)
        self.assertIn('mode === "paragraph" && paragraphChunkMiscueResult', miscue_branch)
        self.assertIn('renderSyllableDisplayWithError(data, highlightedMiscueIndex', miscue_branch)

    def test_story_miscue_consumption_stops_at_speculative_future_substitution(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        consumer = content.split('function consumeSequentialParagraphWordResults(', 1)[1].split('function evaluatedParagraphWordIndex', 1)[0]

        self.assertIn('targetIndex < confirmedWordIndex && status === "correct"', consumer)
        self.assertIn('targetIndex === confirmedWordIndex && status === "miscue"', consumer)
        self.assertIn('status === "correct"', consumer)
        self.assertIn('resultType === "multi_token_substitution"', consumer)
        self.assertIn('break;', consumer)
        self.assertNotIn('recognized_index', consumer)

    def test_story_miscue_completion_requires_authoritative_server_completion(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        miscue_branch = content.split('if (!itemLocked[currentIndex] && (', 1)[1].split('if (Number(data.matched || 0) > 0)', 1)[0]

        completion_condition = miscue_branch.split('const paragraphChunkCompleted', 1)[1].split(';', 1)[0]
        self.assertIn('data.complete === true', completion_condition)
        self.assertIn('finalResolvedTargetIndex === data.word_syllable_ranges.length - 1', completion_condition)
        self.assertIn('post_miscue_sequential_consumption', miscue_branch)

    def test_crla_correct_word_then_miscue_does_not_paint_next_word(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')
        result_handling = content.split('const activeWordIndex = mode === "paragraph"', 1)[1].split('const hasProgressRegression', 1)[0]

        self.assertIn('evaluatedParagraphWordIndex(data.word_syllable_ranges, context?.syllableIndex)', result_handling)
        self.assertIn('recordParagraphWordResult(data.word_results, activeWordIndex)', result_handling)
        self.assertNotIn('const activeWordIndex = Number(data.current_word_index || 0)', result_handling)
        self.assertIn('renderSyllableDisplayWithError(data, activeWordIndex', content)

    def test_regular_materials_render_my_materials_completion_shell_for_each_reading_type(self):
        self._login_student()
        reader_names = {
            'word': 'reading_word_page',
            'sentence': 'reading_sentence_page',
            'paragraph': 'reading_para_page',
        }

        for item_type, reader_name in reader_names.items():
            with self.subTest(item_type=item_type):
                material = Material.objects.create(
                    title=f'{item_type.title()} Material',
                    code=f'REG-{item_type.upper()}',
                    item_type=item_type,
                    content_text='One two three',
                    content_json={'items': ['One two three'], 'language': 'English'},
                    language='English', type='assessment', source_type='personal',
                    assessment_kind='regular', is_official_reading=False,
                    is_system_owned=False, student_access=True,
                )
                response = self.client.get(reverse(reader_name), {
                    'id': str(material.id),
                    'test': material.title,
                    'item_type': item_type,
                })

                self.assertEqual(response.status_code, 200)
                self.assertTrue(response.context['is_my_materials_completion'])
                self.assertContains(response, 'window.__PABASA_MY_MATERIALS__ = true;')
                self.assertContains(response, '<div class="completion-kicker">My Materials</div>', html=True)
                self.assertContains(response, '>Items Correctly Read<', count=1)
                self.assertContains(response, '>Try Again<', count=1)
                self.assertContains(response, '>Done / Back to Materials<', count=1)
                self.assertNotContains(response, '<div class="completion-kicker">Reading Assessment Results</div>', html=True)
                self.assertNotContains(response, '>Score Breakdown<')
                self.assertNotContains(response, 'Calculating your score breakdown')
                self.assertNotContains(response, '>Continue to Sentence Reading')

    def test_story_reading_has_independent_route_template_and_script(self):
        self._login_student()
        material = Material.objects.create(
            title='Sam and His Hat', code='STORY-BOOK', item_type='paragraph',
            content_text='Sam has a big hat. The hat is blue. Sam puts it on.',
            content_json={
                'template_title': 'Story Reading',
                'template_lesson': 'Story Reading',
                'storyTitle': 'Sam and His Hat',
                'storyText': 'Sam has a big hat. The hat is blue. Sam puts it on.',
                'language': 'English',
            },
            language='English', type='assessment', source_type='template',
            assessment_kind='regular', is_official_reading=False,
            is_system_owned=False, student_access=True,
        )

        response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'pabasa_app/story_reading_flipbook.html')
        self.assertContains(response, 'id="storybookApp"')
        self.assertContains(response, 'id="oralReadingButton"')
        self.assertNotContains(response, 'id="storyTitle"')
        self.assertContains(response, 'Start Reading')
        self.assertContains(response, 'Listen to Story')
        self.assertContains(response, 'pabasa_app/js/story_reading_flipbook.js')
        self.assertNotContains(response, 'assessment_reader.js')

        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'story_reading_flipbook.js'
        script = script_path.read_text(encoding='utf-8')
        self.assertIn('navigator.mediaDevices.getUserMedia', script)
        self.assertIn('new MediaRecorder', script)
        self.assertIn('const SPEECH_CHUNK_MS = 3200;', script)
        self.assertIn('const SPEECH_REQUEST_TIMEOUT_MS = 35000;', script)
        self.assertIn("fetch('/api/reading/transcribe/'", script)
        self.assertIn("fetch('/api/reading/read-aloud/'", script)
        self.assertIn('class="book-opening"', script)
        self.assertIn('Read at your own pace', script)
        self.assertIn("pages = [''];", script)
        self.assertIn('index += 2', script)
        self.assertIn('sentences.slice(index, index + 2)', script)
        self.assertIn('pageCount.textContent = `Page ${start + 1} of ${pages.length}`;', script)
        self.assertIn('oralReadingButton.disabled = onTitlePage;', script)
        self.assertIn('listenButton.disabled = onTitlePage;', script)
        self.assertNotIn('Begin on the next page', script)
        self.assertNotIn('Turn the page to start reading the story aloud.', script)
        self.assertIn('class="story-word"', script)
        self.assertIn('cursorFromTranscript', script)
        self.assertIn("word.classList.toggle('is-read'", script)

        assessment_script = (Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js').read_text(encoding='utf-8')
        self.assertIn('["sentence", "paragraph"].includes(mode) ? 10000 : 2400', assessment_script)
        self.assertNotIn('SpeechRecognition', script)
        self.assertNotIn('speechSynthesis', script)

    def test_story_reading_displays_each_material_assigned_week_after_refresh(self):
        self._login_student()
        for index, assigned_week in enumerate((1, 4, 8), start=1):
            with self.subTest(assigned_week=assigned_week):
                material = Material.objects.create(
                    title=f'Story Week {assigned_week}', code=f'STORY-WEEK-{index}', item_type='paragraph',
                    content_text='A story sentence.',
                    content_json={
                        'template_title': 'Story Reading',
                        'storyTitle': f'Story Week {assigned_week}',
                        'storyText': 'A story sentence.',
                        'language': 'English',
                    },
                    language='English', type='assessment', source_type='template',
                    assessment_kind='regular', student_access=True, assigned_week=assigned_week,
                    assigned_weeks=[assigned_week],
                )

                for _ in range(2):
                    response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
                    self.assertEqual(response.status_code, 200)
                    payload = response.context['story_reading_data']
                    self.assertEqual(payload['assigned_week'], assigned_week)
                    self.assertEqual(payload['assigned_week_display'], f'Week {assigned_week}')
                    self.assertContains(response, f'Week {assigned_week} <span aria-hidden="true">·</span> Story Reading', html=True)

    def test_story_reading_completion_updates_shared_assessment_result_and_restores_progress(self):
        student = self._login_student()
        material = Material.objects.create(
            title='The Blue Kite', code='STORY-PERSIST', item_type='paragraph',
            content_text='Mia has a blue kite.',
            content_json={
                'template_title': 'Story Reading',
                'storyTitle': 'The Blue Kite',
                'storyText': 'Mia has a blue kite.',
                'language': 'English',
            },
            language='English', type='assessment', source_type='template',
            assessment_kind='regular', student_access=True,
        )

        response = self.client.post(
            reverse('story_reading_complete'),
            data=json.dumps({
                'material_id': f'material-{material.id}',
                'story_title': 'The Blue Kite',
                'total_words': 5,
                'words_read': 5,
                'progress_percent': 100,
                'duration_seconds': 12,
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        self.assertTrue(response.json()['completed'])
        progress = StoryReadingProgress.objects.get(student=student, material=material)
        self.assertTrue(progress.completed)
        self.assertEqual(progress.words_read, 5)

        completion_response = self.client.post(
            reverse('record_assessment_completion'),
            data=json.dumps({
                'material_id': f'material-{material.id}',
                'activity_type': 'story_reading',
                'assessment_type': 'paragraph',
                'items_completed': 6,
                'correct_items': 5,
                'accuracy': 83.333333,
                'total_score': 83.333333,
                'duration_seconds': 12,
                'scores': {
                    'correct_items': 5,
                    'items_completed': 6,
                    'accuracy': 83.333333,
                    'total_score': 83.333333,
                    'duration_seconds': 12,
                },
            }),
            content_type='application/json',
        )
        self.assertEqual(completion_response.status_code, 200)
        self.assertTrue(completion_response.json()['success'])
        result = material.assessment_results.get(student=student, attempt_status='completed')
        self.assertEqual(result.correct_items, 5)
        self.assertEqual(result.items_completed, 6)
        self.assertAlmostEqual(result.total_score, 83.333333, places=2)

        reopened = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
        self.assertEqual(reopened.status_code, 200)
        self.assertTrue(reopened.context['story_reading_data']['completion']['completed'])
        self.assertEqual(reopened.context['story_reading_data']['return_url'], reverse('dashboard'))

    def test_story_reading_script_uses_shared_completion_flow_at_final_completion(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'story_reading_flipbook.js'
        script = script_path.read_text(encoding='utf-8')
        self.assertIn("fetch('/api/story-reading/complete/'", script)
        player_script = (Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'story_reading_player.js').read_text(encoding='utf-8')
        self.assertIn("fetch('/record-assessment-completion/'", player_script)
        self.assertIn('finishCompletion()', player_script)
        self.assertIn('state.scene >= scenes.length', player_script)

    def test_regular_paragraph_material_stays_on_existing_reader(self):
        self._login_student()
        material = Material.objects.create(
            title='Paragraph Material', code='REG-PARA-SEPARATE', item_type='paragraph',
            content_text='This remains in the existing paragraph reader.',
            content_json={'items': ['This remains in the existing paragraph reader.'], 'language': 'English'},
            language='English', type='assessment', source_type='personal',
            assessment_kind='regular', is_official_reading=False,
            is_system_owned=False, student_access=True,
        )

        response = self.client.get(reverse('reading_para_page'), {'id': f'material-{material.id}'})

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, 'pabasa_app/reading_para_page.html')
        self.assertContains(response, 'assessment_reader.js')
        self.assertNotContains(response, 'story_reading_flipbook.js')

    def test_official_crla_keeps_formal_completion_shell(self):
        self._login_student()
        material = Material.objects.create(
            title='Official CRLA Completion', code='CRLA-COMPLETION', item_type='word',
            content_text='One two three', content_json={'items': ['One', 'two', 'three']},
            type='assessment', assessment_kind='crla', is_official_reading=True,
            is_system_owned=True, student_access=True,
        )
        response = self.client.get(reverse('reading_word_page'), {
            'id': f'material-{material.id}',
            'official_assessment_id': str(material.id),
        })

        self.assertEqual(response.status_code, 200)
        self.assertFalse(response.context.get('is_my_materials_completion', False))
        self.assertContains(response, 'window.__PABASA_MY_MATERIALS__ = false;')
        self.assertContains(response, '<div class="completion-kicker">Reading Assessment Results</div>', html=True)
        self.assertContains(response, '>Your Reading Classification<', count=1)
        self.assertContains(response, '>Score Breakdown<', count=1)
        self.assertNotContains(response, '<div class="completion-kicker">My Materials</div>', html=True)

    def test_completion_javascript_selects_card_from_persisted_material_context(self):
        script_path = Path(__file__).resolve().parent / 'static' / 'pabasa_app' / 'js' / 'assessment_reader.js'
        content = script_path.read_text(encoding='utf-8')

        self.assertIn('const isMyMaterials = window.__PABASA_MY_MATERIALS__ === true;', content)
        self.assertIn('renderMyMaterialsCompletion(latestScores);', content)
        self.assertIn('completionClassificationValue.textContent = `${correct} / ${total}`;', content)
        self.assertIn('completionClassificationLabel.textContent = `${pluralLabel} Correctly Read`;', content)
        my_materials_renderer = content.split('function renderMyMaterialsCompletion(scores)', 1)[1].split('function setSpeechStatus', 1)[0]
        self.assertNotIn('resolveClassificationLabel', my_materials_renderer)
        self.assertNotIn('setCompletionClassification', my_materials_renderer)
        self.assertIn('if (!isMyMaterials && branchState.stage)', content)
        self.assertIn('const shouldShowClassification = !isMyMaterials && showClassification === true;', content)
        self.assertNotIn('urlParams.get("source") === "my_materials"', content)

    def test_filipino_story_sets_load_selected_scenes_and_images(self):
        self._login_student()
        expected = {
            'story-1-ang-umaga-ni-lito': ('filipino-set-1', 'Ang Umaga ni Lito', [
                'Maagang gumising si Lito.', 'Naghilamos siya ng mukha.', 'Nagsipilyo siya ng ngipin.',
                'Kumain siya ng almusal.', 'Isinuot niya ang uniporme.', 'Pumasok si Lito sa paaralan.',
            ]),
            'story-2-si-nena-at-ang-bulaklak': ('filipino-set-2', 'Si Nena at ang Bulaklak', [
                'May nakita si Nena na bulaklak.', 'Maliwanag ang kulay nito.', 'Diniligan niya ang bulaklak.',
                'Lumaki ito araw-araw.', 'Masaya si Nena sa kanyang halaman.', 'Inalagaan niya ito.',
            ]),
            'story-3-ang-masayang-araw': ('filipino-set-3', 'Ang Masayang Araw', [
                'Maaraw noong araw na iyon.', 'Lumabas sina Carlo at Ana.', 'Naglalaro sila sa parke.',
                'Tumakbo si Carlo sa damuhan.', 'Umupo si Ana sa ilalim ng puno.', 'Masaya silang umuwi.',
            ]),
            'story-4-ang-baon-ni-rosa': ('filipino-set-4', 'Ang Baon ni Rosa', [
                'May baon na tinapay si Rosa.', 'May dala rin siyang gatas.',
                'Umupo siya sa tabi ng kanyang kaibigan.', 'Ibinahagi niya ang kanyang tinapay.',
                'Nagpasalamat ang kanyang kaibigan.', 'Masaya silang kumain.',
            ]),
            'story-5-ang-nawalang-lapis': ('filipino-set-5', 'Ang Nawalang Lapis', [
                'Hinahanap ni Joel ang kanyang lapis.', 'Tumingin siya sa kanyang mesa.',
                'Tumingin din siya sa kanyang bag.', 'Nakita niya ito sa ilalim ng libro.',
                'Kinuha niya ang lapis.', 'Nagpatuloy siya sa pagsusulat.',
            ]),
        }
        for selector_key, (story_key, title, sentences) in expected.items():
            with self.subTest(selector_key=selector_key):
                material = Material.objects.create(
                    title='Filipino Story', code=f'FIL-{story_key}', item_type='paragraph',
                    content_json={
                        'template_title': 'Story Reading',
                        'language': 'Filipino',
                        'storyTitle': title,
                        'storyReading': {'language': 'Filipino', 'storyKey': selector_key},
                    },
                    language='Filipino', type='assessment', source_type='template',
                    assessment_kind='regular', student_access=True,
                )
                response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
                payload = response.context['story_reading_data']
                self.assertEqual(payload['story_key'], story_key)
                self.assertEqual(payload['title'], title)
                self.assertEqual(payload['text'].split('\n\n'), sentences)
                self.assertEqual(payload['text'].count('\n\n'), 5)
                self.assertEqual(payload['images'], [
                    f'/static/pabasa_app/images/story_reading/Filipino/Set_{story_key[-1]}/{index}.png'
                    for index in range(1, 7)
                ])

    def test_filipino_story_does_not_restore_progress_from_another_story(self):
        student = self._login_student()
        material = Material.objects.create(
            title='Filipino Story', code='FIL-PERSIST', item_type='paragraph',
            content_json={
                'template_title': 'Story Reading', 'language': 'Filipino',
                'storyTitle': 'Si Nena at ang Bulaklak',
                'storyReading': {'language': 'Filipino', 'storyKey': 'story-2-si-nena-at-ang-bulaklak'},
            },
            language='Filipino', type='assessment', source_type='template',
            assessment_kind='regular', student_access=True,
        )
        StoryReadingProgress.objects.create(
            student=student, material=material, story_title='Ang Umaga ni Lito',
            story_key='filipino-set-1', current_scene=6, current_time_seconds=60,
            completed=True,
        )
        response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
        self.assertIsNone(response.context['story_reading_data']['completion'])

    def test_english_story_sets_load_selected_scenes_and_images(self):
        self._login_student()
        expected = {
            'story-1-ben-and-the-little-pet': ('english-set-1', 'Ben and the Little Pet', [
                'Ben has a pet.', 'It is a little pet.', 'The pet is on a mat.',
                'Ben pats the pet.', 'The pet is happy.', 'Ben is happy too.',
            ]),
            'story-2-mia-and-the-red-ball': ('english-set-2', 'Mia and the Red Ball', [
                'Mia has a red ball.', 'She plays with the ball.', 'The ball rolls away.',
                'Mia runs after it.', 'She gets the ball back.', 'Mia is happy.',
            ]),
            'story-3-sams-big-hat': ('english-set-3', "Sam's Big Hat", [
                'Sam has a big hat.', 'The hat is blue.', 'Sam puts it on.',
                'He goes outside.', 'The sun is hot.', 'Sam likes his hat.',
            ]),
            'story-4-the-little-fish': ('english-set-4', 'The Little Fish', [
                'A little fish is in a pond.', 'The fish swims in the water.',
                'It sees a green leaf.', 'The fish swims under it.',
                'Then it swims away.', 'The little fish is safe.',
            ]),
            'story-5-anas-new-book': ('english-set-5', "Ana's New Book", [
                'Ana has a new book.', 'The book has many pictures.', 'Ana sits on a mat.',
                'She opens the book.', 'She reads each page.', 'Ana likes her new book.',
            ]),
        }
        for selector_key, (story_key, title, sentences) in expected.items():
            with self.subTest(selector_key=selector_key):
                material = Material.objects.create(
                    title='English Story', code=f'ENG-{story_key}', item_type='paragraph',
                    content_json={
                        'template_title': 'Story Reading',
                        'language': 'English',
                        'storyTitle': title,
                        'storyReading': {'language': 'English', 'storyKey': selector_key},
                    },
                    language='English', type='assessment', source_type='template',
                    assessment_kind='regular', student_access=True,
                )
                response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
                payload = response.context['story_reading_data']
                self.assertEqual(payload['language'], 'English')
                self.assertEqual(payload['story_key'], story_key)
                self.assertEqual(payload['title'], title)
                self.assertEqual(payload['text'].split('\n\n'), sentences)
                self.assertEqual(payload['images'], [
                    f'/static/pabasa_app/images/story_reading/English/Set_{story_key[-1]}/{index}.png'
                    for index in range(1, 7)
                ])

    def test_english_story_does_not_restore_progress_from_another_story(self):
        student = self._login_student()
        material = Material.objects.create(
            title='English Story', code='ENG-PERSIST', item_type='paragraph',
            content_json={
                'template_title': 'Story Reading', 'language': 'English',
                'storyTitle': 'Mia and the Red Ball',
                'storyReading': {'language': 'English', 'storyKey': 'story-2-mia-and-the-red-ball'},
            },
            language='English', type='assessment', source_type='template',
            assessment_kind='regular', student_access=True,
        )
        StoryReadingProgress.objects.create(
            student=student, material=material, story_title='Ben and the Little Pet',
            story_key='english-set-1', current_scene=6, current_time_seconds=60,
            completed=True,
        )
        response = self.client.get(reverse('story_reading_page'), {'id': f'material-{material.id}'})
        self.assertIsNone(response.context['story_reading_data']['completion'])


class AssessmentWorkflowStateTests(TestCase):
    def _student_session(self, student):
        session = self.client.session
        session['user_id'] = student.id
        session['user_role'] = 'student'
        session['custom_id'] = student.custom_id
        session['first_name'] = student.first_name
        session['last_name'] = student.last_name
        session['email'] = student.email
        session.save()

    def _student_and_class(self):
        teacher = User.objects.create(
            custom_id="TCHR-2000",
            role="teacher",
            first_name="Taylor",
            last_name="Teacher",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="teacher2000@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-2000",
            role="student",
            first_name="Sam",
            last_name="Student",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=2012,
            email="student2000@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-2000",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._student_session(student)
        return teacher, student, section

    def test_bosy_page_waits_for_published_enabled_class_assessment(self):
        self._student_and_class()

        response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'no_assessment')
        self.assertContains(response, 'No Reading Assessment Available')
        self.assertContains(response, 'Your teacher has not yet published the Beginning of School Year Reading Assessment.')

    def test_bosy_page_shows_existing_workflow_for_published_enabled_class_assessment(self):
        teacher, _student, section = self._student_and_class()
        material = Material.objects.create(
            title="Official BoSY CRLA",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
        )

        response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'pretest')
        self.assertEqual(response.context['crla_material_id'], material.id)
        self.assertContains(response, 'Start Assessment')

    def test_all_below_grade_classifications_are_aral_eligible(self):
        self.assertTrue(_aral_eligible_classification("Low Emerging Readers"))
        self.assertTrue(_aral_eligible_classification("High Emerging Readers"))
        self.assertTrue(_aral_eligible_classification("Developing Readers"))
        self.assertTrue(_aral_eligible_classification("Transitioning Readers"))
        self.assertFalse(_aral_eligible_classification("Readers at Grade Level"))

    def test_sync_assessment_workflow_state_persists_crla_completion_and_aral_eligibility(self):
        student = User.objects.create(
            custom_id="STU-2001",
            role="student",
            first_name="Jamie",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="jamie@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {}},
        )
        assessment = Assessment.objects.create(
            title="CRLA Pre-Test",
            assessment_kind='crla',
            assessment_type='word',
            teacher=student,
            status='published',
            is_active=True,
        )

        _sync_assessment_workflow_state(
            student,
            score_payload={
                'crla_classification': 'Low Emerging Readers',
                'adapted_reading_level': 'Low Emerging Readers',
            },
            assessment=assessment,
        )

        student.refresh_from_db()
        state = student.preference.get("reading_assessment_state", {})
        self.assertEqual(state["reader_classification"], "Low Emerging Readers")
        self.assertTrue(state["aral_eligible"])
        self.assertTrue(state["crla_pretest_completed"])
        self.assertEqual(state["current_phase"], "materials")

    def test_assessment_page_context_exposes_saved_student_profile_metrics(self):
        student = User.objects.create(
            custom_id="STU-2002",
            role="student",
            first_name="Alex",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="alex@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "reader_classification": "Low Emerging Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                },
                "student_profile": {
                    "last_assessment_at": "2026-08-05T10:00:00+00:00",
                    "total_score": 72,
                },
            },
        )
        session = self.client.session
        session['user_id'] = student.id
        session['user_role'] = 'student'
        session['custom_id'] = student.custom_id
        session['first_name'] = student.first_name
        session['last_name'] = student.last_name
        session['email'] = student.email
        session.save()

        response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['student_profile']['last_assessment_at'], "2026-08-05T10:00:00+00:00")
        self.assertEqual(response.context['student_profile']['total_score'], 72)
        self.assertEqual(response.context['stage'], "materials")

    def test_bosy_completion_persists_pretest_completed_and_eligibility(self):
        teacher = User.objects.create(
            custom_id="TCHR-2003",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara3@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-2003",
            role="student",
            first_name="Iris",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="iris3@example.com",
            password_hash=make_password("student-password"),
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-2003",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        material = Material.objects.filter(system_assessment_key="bosy_crla_pretest").first()
        self.assertIsNotNone(material)

        _sync_assessment_workflow_state(
            student,
            score_payload={
                "crla_classification": "Developing Readers",
                "classification": "Developing Readers",
            },
            material=material,
        )

        student.refresh_from_db()
        state = student.preference.get("reading_assessment_state", {})
        self.assertTrue(state.get("crla_pretest_completed"))
        self.assertTrue(state.get("aral_eligible"))
        self.assertEqual(state.get("reader_classification"), "Developing Readers")
        self.assertEqual(state.get("current_phase"), "materials")


class AssessmentPageFlowTests(TestCase):
    def _login_student(self, student):
        session = self.client.session
        session['user_id'] = student.id
        session['user_role'] = student.role
        session['custom_id'] = student.custom_id
        session['first_name'] = student.first_name
        session['last_name'] = student.last_name
        session['email'] = student.email
        session.save()
        return session

    def _create_official_crla_calendar(self, *, pre_start, pre_end, post_start, post_end):
        calendar = SchoolCalendar.objects.create(
            school_year='2026-2027',
            current_term=1,
            is_active=True,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Start of Classes',
            event_type='start_of_classes',
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 1),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='End of Classes',
            event_type='end_of_classes',
            start_date=date(2027, 5, 31),
            end_date=date(2027, 5, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Opening Block',
            event_type='school_opening',
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Closing Block',
            event_type='school_closing',
            start_date=date(2026, 8, 31),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Pre-Assessment Week',
            event_type='pre_assessment',
            start_date=pre_start,
            end_date=pre_end,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title='Post-Assessment Week',
            event_type='post_assessment',
            start_date=post_start,
            end_date=post_end,
        )
        return calendar

    def test_assessment_page_uses_calendar_pre_window_for_bosy_official_crla(self):
        teacher = User.objects.create(
            custom_id="TCHR-3101",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3101",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": False, "crla_posttest_completed": False}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3101",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 8),
            post_start=date(2026, 8, 9),
            post_end=date(2026, 8, 15),
        )
        pre_material = Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )
        post_material = Material.objects.create(
            title="End of School Year (EoSY) CRLA Post-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="eosy_crla_posttest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 5)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['crla_material_id'], pre_material.id)
        self.assertEqual(response.context['crla_official_assessment_data']['assessment_type'], 'pretest')
        self.assertEqual(response.context['crla_official_assessment_data']['official_title'], pre_material.title)

    def test_assessment_page_uses_calendar_midline_window_for_official_crla(self):
        teacher = User.objects.create(
            custom_id="TCHR-3102",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia3102@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3102",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia3102@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": False, "crla_posttest_completed": False}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class 2",
            class_code="READ-3102",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        calendar = SchoolCalendar.objects.create(
            school_year="2026-2027",
            current_term=1,
            is_active=True,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Start of Classes",
            event_type="start_of_classes",
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 1),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="End of Classes",
            event_type="end_of_classes",
            start_date=date(2027, 5, 31),
            end_date=date(2027, 5, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Opening Block",
            event_type="school_opening",
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Closing Block",
            event_type="school_closing",
            start_date=date(2027, 3, 1),
            end_date=date(2027, 3, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Midline Assessment Week",
            event_type="midline_assessment",
            start_date=date(2026, 11, 9),
            end_date=date(2026, 11, 13),
        )
        mid_material = Material.objects.create(
            title="Midline CRLA Mid-Test",
            item_type="word",
            content_text="write",
            content_json={"items": ["write"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="midline_crla_midtest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 11, 10)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['active_phase'], 'midtest')
        self.assertEqual(response.context['crla_material_id'], mid_material.id)
        self.assertEqual(response.context['crla_official_assessment_data']['assessment_type'], 'midtest')
        self.assertEqual(response.context['crla_official_assessment_data']['official_title'], mid_material.title)

    def test_assessment_page_does_not_activate_midline_outside_window(self):
        teacher = User.objects.create(
            custom_id="TCHR-3103",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia3103@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3103",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia3103@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": False, "crla_posttest_completed": False}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class 3",
            class_code="READ-3103",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        calendar = SchoolCalendar.objects.create(
            school_year="2026-2027",
            current_term=1,
            is_active=True,
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Start of Classes",
            event_type="start_of_classes",
            start_date=date(2026, 6, 1),
            end_date=date(2026, 6, 1),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="End of Classes",
            event_type="end_of_classes",
            start_date=date(2027, 5, 31),
            end_date=date(2027, 5, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Opening Block",
            event_type="school_opening",
            start_date=date(2026, 8, 1),
            end_date=date(2026, 8, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Closing Block",
            event_type="school_closing",
            start_date=date(2027, 3, 1),
            end_date=date(2027, 3, 31),
        )
        CalendarEvent.objects.create(
            school_calendar=calendar,
            term=1,
            title="Midline Assessment Week",
            event_type="midline_assessment",
            start_date=date(2026, 11, 9),
            end_date=date(2026, 11, 13),
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 11, 20)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertNotEqual(response.context['active_phase'], 'midtest')
        self.assertNotEqual(response.context['crla_official_assessment_data'].get('assessment_type'), 'midtest')

    def test_active_school_calendar_recognizes_school_opening_and_closing_bounds(self):
        calendar = self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 7),
            post_start=date(2026, 8, 8),
            post_end=date(2026, 8, 15),
        )

        active = _active_school_calendar(date(2026, 8, 8))

        self.assertIsNotNone(active)
        self.assertEqual(active.id, calendar.id)

    def test_get_class_materials_hides_official_crla_during_intervention_but_keeps_teacher_materials(self):
        teacher = User.objects.create(
            custom_id="TCHR-1004",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher1004@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-1004",
            role="student",
            first_name="Sam",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="student1004@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": True, "crla_posttest_completed": False, "reader_classification": "Low Emerging Readers", "aral_eligible": True}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-1004",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 5),
            post_start=date(2026, 8, 8),
            post_end=date(2026, 8, 11),
        )
        Material.objects.get_or_create(
            system_assessment_key="bosy_crla_pretest",
            defaults={
                "title": "Beginning of School Year (BoSY) CRLA Pre-Test",
                "item_type": "paragraph",
                "content_text": "read",
                "content_json": {"items": ["read"]},
                "assessment_kind": "crla",
                "assessment_set": "crla",
                "type": "assessment",
                "status": "published",
                "student_access": True,
                "section": None,
                "teacher": teacher,
                "is_active": True,
                "is_official_reading": True,
                "is_system_owned": True,
                "system_assessment_phase": "pretest",
                "language": "Filipino",
            },
        )
        Material.objects.get_or_create(
            system_assessment_key="eosy_crla_posttest",
            defaults={
                "title": "End of School Year (EoSY) CRLA Post-Test",
                "item_type": "paragraph",
                "content_text": "read",
                "content_json": {"items": ["read"]},
                "assessment_kind": "crla",
                "assessment_set": "crla",
                "type": "assessment",
                "status": "published",
                "student_access": True,
                "section": None,
                "teacher": teacher,
                "is_active": True,
                "is_official_reading": True,
                "is_system_owned": True,
                "system_assessment_phase": "posttest",
                "language": "Filipino",
            },
        )
        teacher_material = Material.objects.create(
            title="Teacher Reading",
            item_type="word",
            content_text="alpha",
            content_json={"items": ["alpha"]},
            assessment_kind="regular",
            assessment_set="word",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 6)
            response = self.client.get(reverse('get_class_materials'), {'class_code': section.class_code})

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['official_assessments'], [])
        word_items = data['materials']['word']
        self.assertTrue(any(item['title'] == 'Teacher Reading' for item in word_items))
        self.assertFalse(any(item['title'] == 'Beginning of School Year (BoSY) CRLA Pre-Test' for item in word_items))
        self.assertFalse(any(item['title'] == 'End of School Year (EoSY) CRLA Post-Test' for item in word_items))

    def test_get_class_materials_never_surfaces_official_crla_for_fresh_student_account(self):
        teacher = User.objects.create(
            custom_id="TCHR-1005",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="teacher1005@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-1005",
            role="student",
            first_name="Nina",
            last_name="New",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="student1005@example.com",
            password_hash=make_password("student-password"),
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-1005",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 7),
            post_start=date(2026, 8, 8),
            post_end=date(2026, 8, 15),
        )
        Material.objects.create(
            title="BoSY CRLA Pre-Test",
            item_type="word",
            content_text="official",
            content_json={"items": ["official"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=None,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
            system_assessment_period="bosy",
            system_assessment_phase="pretest",
        )
        Material.objects.create(
            title="Teacher Practice",
            item_type="word",
            content_text="alpha",
            content_json={"items": ["alpha"]},
            assessment_kind="regular",
            assessment_set="word",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
        )
        self._login_student(student)
        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 3)
            response = self.client.get(reverse('get_class_materials'), {'class_code': section.class_code})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['official_assessments'], [])
        self.assertTrue(any(item['title'] == 'Teacher Practice' for item in payload['materials']['word']))
        self.assertFalse(any(item['title'] == 'BoSY CRLA Pre-Test' for item in payload['materials']['word']))

    def test_assessment_page_routes_to_intervention_when_between_calendar_windows(self):
        teacher = User.objects.create(
            custom_id="TCHR-3102",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia2@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3102",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia2@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": True, "crla_posttest_completed": False, "reader_classification": "Low Emerging Readers", "aral_eligible": True, "current_phase": "materials"}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3102",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 7),
            post_start=date(2026, 8, 10),
            post_end=date(2026, 8, 15),
        )
        Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )
        Material.objects.create(
            title="End of School Year (EoSY) CRLA Post-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="eosy_crla_posttest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 8)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'original')
        self.assertFalse(response.context['official_assessment_available'])
        self.assertEqual(response.context['crla_material_id'], '')

    def test_assessment_page_uses_calendar_post_window_for_eosy_official_crla(self):
        teacher = User.objects.create(
            custom_id="TCHR-3103",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia3@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3103",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia3@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": False, "crla_posttest_completed": False}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3103",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 7),
            post_start=date(2026, 8, 8),
            post_end=date(2026, 8, 15),
        )
        pre_material = Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )
        post_material = Material.objects.create(
            title="End of School Year (EoSY) CRLA Post-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="eosy_crla_posttest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 8)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['crla_material_id'], post_material.id)
        self.assertEqual(response.context['crla_official_assessment_data']['assessment_type'], 'posttest')
        self.assertEqual(response.context['crla_official_assessment_data']['official_title'], post_material.title)

    def test_assessment_page_returns_unavailable_without_any_crla_window(self):
        teacher = User.objects.create(
            custom_id="TCHR-3104",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia4@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3104",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia4@example.com",
            password_hash=make_password("student-password"),
            preference={"reading_assessment_state": {"crla_pretest_completed": False, "crla_posttest_completed": False}},
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3104",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )
        Material.objects.create(
            title="End of School Year (EoSY) CRLA Post-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="eosy_crla_posttest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 8)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'unavailable')
        self.assertFalse(response.context['official_assessment_available'])
        self.assertEqual(response.context['crla_material_id'], '')
        self.assertEqual(response.context['workflow_title'], 'CRLA Assessment Currently Unavailable')

    def test_assessment_page_requires_saved_completion_record_before_showing_complete_card(self):
        teacher = User.objects.create(
            custom_id="TCHR-3105",
            role="teacher",
            first_name="Tia",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tia5@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3105",
            role="student",
            first_name="Pia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="pia5@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3105",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        self._create_official_crla_calendar(
            pre_start=date(2026, 8, 1),
            pre_end=date(2026, 8, 7),
            post_start=date(2026, 8, 10),
            post_end=date(2026, 8, 15),
        )
        material = Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )
        self._login_student(student)

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 5)
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'assessment')

        Assessment.objects.create(
            title=material.title,
            code=f"{material.code}-RESULT",
            teacher=teacher,
            material=material,
            student=student,
            attempt_status='completed',
            completed_at=timezone.now(),
            assessment_type='pretest',
            source_assessment=None,
            is_active=True,
        )

        with patch('pabasa_app.views.date', wraps=date) as mock_date:
            mock_date.today.return_value = date(2026, 8, 5)
            completed_response = self.client.get(reverse('assessment'))

        self.assertEqual(completed_response.status_code, 200)
        self.assertEqual(completed_response.context['stage'], 'original')

    def test_assessment_page_allows_aral_materials_after_pretest_eligibility(self):
        student = User.objects.create(
            custom_id="STU-3001",
            role="student",
            first_name="Ari",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="ari@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Low Emerging Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        self._login_student(student)

        response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'original')

    def test_assessment_page_pretest_landing_card_uses_bosy_labels(self):
        teacher = User.objects.create(
            custom_id="TCHR-3001",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3003",
            role="student",
            first_name="Noah",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="noah@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": False,
                    "crla_posttest_completed": False,
                    "reader_classification": "",
                    "aral_eligible": False,
                    "current_phase": "pretest",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3001",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        pre_material = Material.objects.filter(system_assessment_key="bosy_crla_pretest").first()
        self.assertIsNotNone(pre_material)

        self._login_student(student)
        with patch("pabasa_app.views._official_crla_assessment_phase", return_value="pretest"), patch(
            "pabasa_app.views._official_crla_material_for_student", return_value=pre_material
        ):
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'assessment')
        self.assertEqual(response.context['workflow_title'], 'Beginning of School Year Reading Assessment')
        self.assertEqual(response.context['workflow_message'], 'Complete the official BoSY CRLA Assessment prepared by your teacher.')
        self.assertEqual(response.context['workflow_subtitle'], 'Beginning of School Year reading assessment')

    def test_assessment_page_routes_eligible_bosy_completion_to_intervention(self):
        teacher = User.objects.create(
            custom_id="TCHR-3002",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara2@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3004",
            role="student",
            first_name="Iris",
            last_name="Learner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="iris@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3002",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        pre_material = Material.objects.filter(system_assessment_key="bosy_crla_pretest").first()
        self.assertIsNotNone(pre_material)

        self._login_student(student)
        with patch("pabasa_app.views._official_crla_assessment_phase", return_value="pretest"), patch(
            "pabasa_app.views._official_crla_material_for_student", return_value=pre_material
        ):
            response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'original')
        self.assertEqual(response.context['official_reading_assessments'], [])
        self.assertEqual(response.context['crla_assessment_title'], 'CRLA Assessment')

    def test_activate_aral_intervention_persists_active_state_and_routes_to_original_workflow(self):
        teacher = User.objects.create(
            custom_id="TCHR-3002A",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara2a@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3004A",
            role="student",
            first_name="Iris",
            last_name="Learner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="iris2@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                    "aral_status": "pending",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3002A",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)

        self._login_student(student)
        response = self.client.post(reverse("activate_aral_intervention"))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("assessment"), response.headers.get("Location", ""))
        student.refresh_from_db()
        state = student.preference.get("reading_assessment_state", {})
        self.assertEqual(state.get("aral_status"), "active")
        self.assertEqual(state.get("current_phase"), "materials")

        with patch("pabasa_app.views._official_assessment_availability_for_student", return_value={"available": False, "assessment_type": "intervention", "school_year": None, "current_term": None, "active_window": None}):
            followup = self.client.get(reverse("assessment"))

        self.assertEqual(followup.status_code, 200)
        self.assertEqual(followup.context["stage"], "original")

    def test_official_crla_assessment_labels_return_posttest_copy(self):
        labels = _official_crla_assessment_labels('posttest')
        self.assertEqual(labels['workflow_title'], 'End of School Year Reading Assessment')
        self.assertEqual(labels['workflow_message'], 'Complete the official EoSY CRLA Assessment prepared by your teacher.')
        self.assertEqual(labels['workflow_subtitle'], 'End of School Year reading assessment')

    def test_official_assessment_availability_hides_bosy_after_eligible_completion(self):
        student = User.objects.create(
            custom_id="STU-3005",
            role="student",
            first_name="Iris",
            last_name="Learner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="iris5@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        availability = _official_assessment_availability_for_student(student)
        self.assertFalse(availability["available"])
        self.assertEqual(availability["assessment_type"], "intervention")

    def test_official_assessment_availability_returns_eosy_when_posttest_window_is_active(self):
        teacher = User.objects.create(
            custom_id="TCHR-3005",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara5@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3006",
            role="student",
            first_name="Iris",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="iris6@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": False,
                    "reader_classification": "Developing Readers",
                    "aral_eligible": True,
                    "current_phase": "materials",
                }
            },
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3006",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        with patch("pabasa_app.views._official_crla_assessment_phase", return_value="posttest"):
            availability = _official_assessment_availability_for_student(student)
        self.assertTrue(availability["available"])
        self.assertEqual(availability["assessment_type"], "posttest")

    def test_completed_bosy_crla_routes_eligible_students_to_aral_flow(self):
        teacher = User.objects.create(
            custom_id="TCHR-3007",
            role="teacher",
            first_name="Tara",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=1990,
            email="tara7@example.com",
            password_hash=make_password("teacher-password"),
        )
        student = User.objects.create(
            custom_id="STU-3007",
            role="student",
            first_name="Mina",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="mina7@example.com",
            password_hash=make_password("student-password"),
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Reading Class",
            class_code="READ-3007",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        material = Material.objects.create(
            title="Beginning of School Year (BoSY) CRLA Pre-Test",
            item_type="word",
            content_text="read",
            content_json={"items": ["read"]},
            assessment_kind="crla",
            assessment_set="crla",
            type="assessment",
            status="published",
            student_access=True,
            section=section,
            teacher=teacher,
            is_active=True,
            is_official_reading=True,
            is_system_owned=True,
            system_assessment_key="bosy_crla_pretest",
        )

        _sync_assessment_workflow_state(
            student,
            score_payload={
                "crla_classification": "Developing Readers",
                "classification": "Developing Readers",
            },
            material=material,
        )

        student.refresh_from_db()
        state = student.preference.get("reading_assessment_state", {})
        self.assertTrue(state.get("aral_eligible"))
        self.assertEqual(state.get("reader_classification"), "Developing Readers")
        self.assertEqual(state.get("current_phase"), "materials")
        self.assertEqual(
            state.get("crla_windows", {}).get("pretest", {}).get("classification"),
            "Developing Readers",
        )

        self._login_student(student)
        with patch("pabasa_app.views._official_assessment_availability_for_student", return_value={"available": False, "assessment_type": "intervention", "school_year": None, "current_term": None, "active_window": None}):
            response = self.client.get(reverse("assessment"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["stage"], "original")
        self.assertEqual(response.context["eligibility"]["reader_classification"], "Developing Readers")
        self.assertTrue(response.context["eligibility"]["aral_eligible"])

    def test_assessment_page_shows_completion_card_for_non_eligible_students_after_posttest(self):
        student = User.objects.create(
            custom_id="STU-3002",
            role="student",
            first_name="Mina",
            last_name="Reader",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=2,
            birth_year=2012,
            email="mina@example.com",
            password_hash=make_password("student-password"),
            preference={
                "reading_assessment_state": {
                    "crla_pretest_completed": True,
                    "crla_posttest_completed": True,
                    "reader_classification": "Readers at Grade Level",
                    "aral_eligible": False,
                    "current_phase": "complete",
                }
            },
        )
        self._login_student(student)

        response = self.client.get(reverse('assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context['stage'], 'complete')


class TeacherSignupTemplateTests(TestCase):
    def test_teacher_signup_template_includes_privacy_step_and_consent_controls(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "teacher_signup.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn('data-signup-step="3"', content)
        self.assertIn("I agree to the Privacy Policy and Terms of Service", content)
        self.assertIn("${stepLabels[currentStep]} ${currentStep + 1}/${steps.length}", content)


class StudentSignupTemplateTests(TestCase):
    def test_student_signup_template_includes_privacy_step_and_consent_controls(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "student_signup.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn('data-signup-step="3"', content)
        self.assertIn("I agree to the Privacy Policy and Terms of Service", content)
        self.assertIn("Step ${currentStep + 1} of ${steps.length}", content)

    def test_student_signup_template_removes_grade_selector_and_enforces_grade_2(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "student_signup.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertNotIn('id="studentGradeLevel"', content)
        self.assertNotIn('name="grade_level"', content)
        self.assertIn('Grade 2', content)


class AssessmentResultsPageTests(TestCase):
    def test_completion_page_uses_child_friendly_summary_copy(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "reading_assessment_base.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn("Great job completing your reading assessment! Keep practicing to improve your reading skills.", content)
        self.assertNotIn("Your results show your current reading performance", content)
        self.assertNotIn("completionPerformanceInterpretation", content)

    def test_completion_page_has_no_score_loading_placeholder(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "reading_assessment_base.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertNotIn("completion-loading", content)
        self.assertNotIn("Calculating your score breakdown...", content)

    def test_completion_page_shows_only_final_classification(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "reading_assessment_base.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn("Your Reading Classification", content)
        self.assertNotIn("Score Breakdown", content)
        self.assertNotIn("RESULTS OVERVIEW", content)
        self.assertNotIn("READING DETAILS", content)
        self.assertNotIn("Calculating your score breakdown...", content)
        self.assertNotIn("Correct words read", content)
        self.assertNotIn("Reading type", content)

    def test_completion_flow_returns_without_classification_overlay(self):
        script_path = Path(__file__).resolve().parent / "static" / "pabasa_app" / "js" / "assessment_reader.js"
        content = script_path.read_text(encoding="utf-8")

        self.assertNotIn("show_classification", content)
        self.assertNotIn("goBackToAssessments(true)", content)
        self.assertIn("goBackToAssessments()", content)

    def test_build_reading_report_pdf_omits_performance_interpretation(self):
        report = {
            "student_name": "Jane Doe",
            "student_id": "1001",
            "grade_level": "Grade 2",
            "email": "jane@example.com",
            "joined_classes": ["Class A"],
            "course_name": "Reading",
            "course_code": "R1",
            "reading_level": "Transitioning Readers",
            "accuracy": 88,
            "wpm": 68,
            "fluency_score": 84,
            "duration_seconds": 120,
            "time_score": 90,
            "pronunciation_score": 82,
            "final_score": 85,
            "summary": "Strong performance",
            "recommendation": "Keep practicing",
            "completed_at": timezone.now().isoformat(),
        }

        pdf_bytes = _build_reading_report_pdf(report)
        reader = PdfReader(BytesIO(pdf_bytes))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)

        self.assertIn("Reading Level", text)
        self.assertNotIn("Performance Interpretation", text)

    def test_build_reading_report_pdf_includes_derived_classification_labels(self):
        report = {
            "student_name": "Jane Doe",
            "student_id": "1001",
            "grade_level": "Grade 2",
            "email": "jane@example.com",
            "joined_classes": ["Class A"],
            "course_name": "Reading",
            "course_code": "R1",
            "reading_level": "Readers at Grade Level",
            "accuracy": 88,
            "wpm": 68,
            "fluency_score": 84,
            "duration_seconds": 120,
            "time_score": 90,
            "pronunciation_score": 82,
            "final_score": 85,
            "summary": "Strong performance",
            "recommendation": "Keep practicing",
            "completed_at": timezone.now().isoformat(),
        }

        pdf_bytes = _build_reading_report_pdf(report)
        reader = PdfReader(BytesIO(pdf_bytes))
        text = "\n".join(page.extract_text() or "" for page in reader.pages)

        self.assertIn("Reading Classification", text)
        self.assertIn("CRLA Reading Classification", text)
        self.assertIn("Readers at Grade Level", text)
        self.assertIn("Phil-IRI Classification", text)
        self.assertIn("Independent", text)
        self.assertIn("PABASA Level", text)
        self.assertIn("Expert Reader", text)


class ReadingMatcherTests(TestCase):
    def test_material_languages_use_philippine_stt_locales(self):
        self.assertEqual(language_code_for("English"), "en-PH")
        self.assertEqual(language_code_for("Filipino"), "fil-PH")
        self.assertEqual(language_code_for("Tagalog"), "fil-PH")

    def test_philippine_locales_use_supported_v1_models(self):
        self.assertEqual(v1_model_for_language("latest_short", "en-PH"), "command_and_search")
        self.assertEqual(v1_model_for_language("latest_short", "fil-PH"), "")

    @patch("google.cloud.speech_v2.SpeechClient")
    @patch("pabasa_app.reading_stt.google_stt_credentials", return_value=object())
    def test_chirp3_uses_synchronous_recognition_for_short_clips(self, credentials, speech_client):
        client = speech_client.return_value
        client.recognize.return_value = SimpleNamespace(results=[
            SimpleNamespace(alternatives=[SimpleNamespace(transcript="Magandang umaga")]),
        ])

        transcript = transcribe_audio_bytes_v2_chirp3(
            b"short reading clip",
            "fil-PH",
            "test-project",
            "us",
            "unused-service-account.json",
        )

        self.assertEqual(transcript, "Magandang umaga")
        request = client.recognize.call_args.kwargs["request"]
        self.assertEqual(request.recognizer, "projects/test-project/locations/us/recognizers/_")
        self.assertEqual(request.config.model, "chirp_3")
        self.assertEqual(request.config.language_codes, ["fil-PH"])
        self.assertEqual(request.content, b"short reading clip")
        self.assertEqual(client.recognize.call_args.kwargs["timeout"], 12)

    @patch("pabasa_app.reading_stt._post_google_tts", return_value="encoded-audio")
    def test_read_aloud_uses_filipino_voice_for_tagalog_locale(self, post_google_tts):
        synthesize_read_aloud_audio("Magandang araw", "test-key", language_code="fil-PH")

        _, payload = post_google_tts.call_args.args
        self.assertEqual(payload["voice"]["languageCode"], "fil-PH")
        self.assertEqual(payload["voice"]["name"], "fil-ph-Neural2-A")

    def test_english_homophone_is_accepted(self):
        result = analyze_reading("two", 0, "too", language_code="en-US")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])

    def test_non_homophone_is_rejected(self):
        result = analyze_reading("cat", 0, "cut", language_code="en-US")

        self.assertEqual(result["correct_word_count"], 0)
        self.assertFalse(result["complete"])

    def test_cmu_homophones_are_not_applied_to_filipino(self):
        result = analyze_reading("two", 0, "too", language_code="fil-PH")

        self.assertEqual(result["correct_word_count"], 0)
        self.assertFalse(result["complete"])

    def test_filipino_syllable_tokens_match_one_target_word(self):
        result = analyze_reading("kabayo", 0, "ka ba yo", language_code="fil-PH")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])
        self.assertEqual(result["matched"], 3)

    def test_filipino_diphthong_word_ay_is_accepted(self):
        result = analyze_reading("ay", 0, "ay", language_code="fil-PH")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])
        self.assertEqual(result["matched"], 1)

    def test_filipino_spoken_vowel_tokens_preserve_exact_tokens(self):
        self.assertEqual(ReadingMatcher.normalize_spoken_word("ay"), "ay")
        self.assertEqual(ReadingMatcher.normalize_spoken_word("aye"), "aye")
        self.assertTrue(ReadingMatcher("ay", 0, "fil-PH").words_match("aye", "ay"))
        self.assertTrue(ReadingMatcher("ay", 0, "fil-PH").words_match("ay", "a"))

    def test_tass_stitches_filipino_syllables_across_chunks(self):
        first_analysis, first_context, first_applied = target_aware_syllable_stitching(
            "Tatay", 0, "", "ta", "fil-PH"
        )
        stitched, next_context, applied = target_aware_syllable_stitching(
            "Tatay", 0, first_context, "tay", "fil-PH"
        )

        self.assertEqual(first_analysis, "ta")
        self.assertEqual(first_context, "ta")
        self.assertFalse(first_applied)
        self.assertEqual(stitched, "ta tay")
        self.assertEqual(next_context, "")
        self.assertTrue(applied)
        self.assertTrue(analyze_reading("Tatay", 0, stitched, "fil-PH")["complete"])

    def test_tass_discards_non_prefix_syllables(self):
        analysis_text, context, applied = target_aware_syllable_stitching(
            "Tatay", 0, "ta", "bo", "fil-PH"
        )

        self.assertEqual(analysis_text, "bo")
        self.assertEqual(context, "")
        self.assertFalse(applied)

    def test_tass_counts_target_syllables_in_context(self):
        self.assertEqual(
            syllable_context_metrics("Kabayo", 0, "kaba", "fil-PH"),
            (2, 3, 66.67),
        )
        self.assertEqual(
            syllable_context_metrics("Tatay", 0, "ta tay", "fil-PH"),
            (2, 2, 100.0),
        )

    def test_tass_context_metrics_reject_non_prefix_context(self):
        self.assertEqual(
            syllable_context_metrics("Tatay", 0, "bo", "fil-PH"),
            (0, 2, 0),
        )

    def test_tass_is_disabled_for_english(self):
        analysis_text, context, applied = target_aware_syllable_stitching(
            "today", 0, "to", "day", "en-PH"
        )

        self.assertEqual(analysis_text, "day")
        self.assertEqual(context, "")
        self.assertFalse(applied)

    def test_strict_syllabic_word_match_uses_authoritative_parts(self):
        self.assertTrue(strict_syllabic_word_match("mansanas", "mansanas", ["man", "sa", "nas"], "fil-PH"))
        self.assertTrue(strict_syllabic_word_match("mansanas", "man sa nas", ["man", "sa", "nas"], "fil-PH"))
        self.assertTrue(strict_syllabic_word_match("mansanas", "man-sa-nas", ["man", "sa", "nas"], "fil-PH"))
        self.assertFalse(strict_syllabic_word_match("mansanas", "mansa nas", ["man", "sa", "nas"], "fil-PH"))
        self.assertFalse(strict_syllabic_word_match("mansanas", "mansa-nas", ["man", "sa", "nas"], "fil-PH"))
        self.assertFalse(strict_syllabic_word_match("mansanas", "", ["man", "sa", "nas"], "fil-PH"))

    def test_filipino_joined_syllables_allow_one_stt_vowel_error(self):
        result = analyze_reading("puno", 0, "po no", language_code="fil-PH")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])

    def test_filipino_target_adds_whole_word_and_syllable_hints(self):
        self.assertEqual(target_phrase_hints("Araw Puno", "fil-PH"), ["araw", "a", "raw", "puno", "pu", "no"])
        self.assertEqual(target_phrase_hints("Araw", "en-PH"), [])

    def test_english_tokens_are_not_joined_into_one_target_word(self):
        result = analyze_reading("somebody", 0, "some body", language_code="en-PH")

        self.assertEqual(result["correct_word_count"], 0)
        self.assertFalse(result["complete"])

    def test_word_numbers_in_english_transcript(self):
        self.assertEqual(
            word_numbers_in_transcript("I read 19 of 1,000 words."),
            "I read nineteen of one thousand words.",
        )

    def test_word_numbers_in_filipino_transcript(self):
        self.assertEqual(word_numbers_in_transcript("Bumasa ng 15", "fil-PH"), "Bumasa ng labinlima")

    def test_filipino_numeric_target_and_spoken_number_word_match(self):
        result = analyze_reading("15", 0, "labinlima", language_code="fil-PH")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])

    def test_word_numbers_does_not_rewrite_decimal_values(self):
        self.assertEqual(word_numbers_in_transcript("Score: 19.5"), "Score: 19.5")

    def test_wrong_word_does_not_complete_target(self):
        result = analyze_reading("water", 0, "apple")

        self.assertEqual(result["matched"], 0)
        self.assertEqual(result["correct_word_count"], 0)
        self.assertFalse(result["complete"])

    def test_numeric_text_is_not_treated_as_list_marker(self):
        result = analyze_reading("19", 0, "19")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])
        self.assertEqual(result["matched"], 1)

    def test_numeric_target_and_spoken_number_words_match(self):
        result = analyze_reading("19", 0, "nineteen")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])
        self.assertEqual(result["matched"], 1)

    def test_numeric_target_and_spoken_digit_match(self):
        result = analyze_reading("Nineteen", 0, "19")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertTrue(result["complete"])
        self.assertTrue(result["matched"] > 0)

    def test_normalize_words_preserves_numeric_tokens_with_punctuation(self):
        self.assertEqual(ReadingMatcher.normalize_words("19."), ["19"])
        self.assertEqual(ReadingMatcher.normalize_words("19,"), ["19"])

    def test_story_alignment_reports_single_substitution_without_rating_entire_sentence_wrong(self):
        result = align_story_transcript(
            "Si Ana ay pumunta sa bahay.",
            "Si Ana ay puminta sa bahay.",
            language_code="fil-PH",
        )
        self.assertEqual(result["total_words"], 6)
        self.assertEqual(result["correct_words"], 5)
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(result["word_results"][3]["result"], "miscue")
        self.assertEqual(result["word_results"][3]["type"], "substitution")

    def test_story_alignment_groups_two_token_tatay_miscue_without_resolving_may(self):
        target = "Iba't ibang tao ang sumasakay sa jeepney ni Tatay. May mga estudyante."
        result = align_story_transcript(
            target,
            "iba't ibang tao ang sumasakay sa jeepney ni na nay",
            language_code="fil-PH",
            start_word_index=0,
        )
        tatay = next(item for item in result["word_results"] if item.get("expected_index") == 8)
        self.assertEqual(tatay["recognized"], "na nay")
        self.assertEqual(tatay["result"], "miscue")
        self.assertEqual(tatay["type"], "multi_token_substitution")
        self.assertEqual(tatay["recognized_start_index"], 8)
        self.assertEqual(tatay["recognized_end_index"], 10)
        self.assertNotIn(9, [item.get("expected_index") for item in result["word_results"]])

    def test_story_alignment_does_not_merge_li_with_following_correct_word(self):
        result = align_story_transcript("bibilis pa sa akin", "li pa sa akin", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "miscue"), (1, "correct"), (2, "correct"), (3, "correct")],
        )

    def test_story_alignment_keeps_genuine_consecutive_miscues_separate(self):
        result = align_story_transcript("tatay may", "lolo bukas", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["recognized"], item["type"]) for item in result["word_results"]],
            [(0, "lolo", "substitution"), (1, "bukas", "substitution")],
        )

    def test_story_alignment_two_token_transition_preserves_one_token_correct_reading(self):
        result = align_story_transcript("tatay may bata", "tatay may bata", start_word_index=0)
        self.assertEqual([item["type"] for item in result["word_results"]], ["correct", "correct", "correct"])

    def test_story_alignment_two_token_transition_preserves_repeated_words(self):
        result = align_story_transcript("may may bata", "may may bata", start_word_index=0)
        self.assertEqual([item["expected_index"] for item in result["word_results"]], [0, 1, 2])
        self.assertTrue(all(item["result"] == "correct" for item in result["word_results"]))

    def test_story_alignment_multi_token_cursor_index_is_absolute(self):
        result = align_story_transcript("una pangalawa tatay may", "na nay", start_word_index=2)
        self.assertEqual(len(result["word_results"]), 1)
        self.assertEqual(result["word_results"][0]["expected_index"], 2)

    def test_story_alignment_one_token_recognized_index_remains_callback_relative(self):
        result = align_story_transcript("una pangalawa tatay may", "tatay", start_word_index=2)
        self.assertEqual(result["word_results"][0]["recognized_index"], 0)

    def test_story_alignment_multi_token_span_keeps_existing_consumer_fields(self):
        result = align_story_transcript("tatay may", "na nay", start_word_index=0)
        item = result["word_results"][0]
        self.assertEqual(item["expected_index"], 0)
        self.assertEqual(item["recognized_index"], 0)
        self.assertEqual(item["recognized_start_index"], 0)
        self.assertEqual(item["recognized_end_index"], 2)
        self.assertEqual(item["result"], "miscue")

    def test_story_cursor_relative_alignment_resolves_multi_word_chunk(self):
        result = align_story_transcript(
            "ako ang pinakamabilis tumakbo sa bahay",
            "pinakamabilis tumakbo sa",
            language_code="fil-PH",
            start_word_index=2,
        )
        resolved = [(item["expected_index"], item["result"]) for item in result["word_results"]]
        self.assertEqual(resolved, [(2, "correct"), (3, "correct"), (4, "correct")])

    def test_story_cursor_alignment_keeps_complete_correct_sequence(self):
        result = align_story_transcript(
            "Ang bata ay pumunta",
            "Ang bata ay pumunta",
            start_word_index=0,
        )
        self.assertEqual(result["miscues"], 0)
        self.assertEqual(
            [(item["expected"], item["result"]) for item in result["word_results"]],
            [("ang", "correct"), ("bata", "correct"), ("ay", "correct"), ("pumunta", "correct")],
        )

    def test_story_cursor_alignment_marks_middle_omission_and_keeps_following_word(self):
        result = align_story_transcript(
            "Ang bata ay pumunta",
            "Ang bata pumunta",
            start_word_index=0,
        )
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(
            [(item["expected"], item["recognized"], item["result"], item["type"]) for item in result["word_results"]],
            [
                ("ang", "ang", "correct", "correct"),
                ("bata", "bata", "correct", "correct"),
                ("ay", None, "miscue", "omission"),
                ("pumunta", "pumunta", "correct", "correct"),
            ],
        )

    def test_story_cursor_alignment_preserves_substitution_behavior(self):
        result = align_story_transcript(
            "Ang bata ay pumunta",
            "Ang bata ba pumunta",
            start_word_index=0,
        )
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(
            [(item["expected"], item["recognized"], item["type"]) for item in result["word_results"]],
            [
                ("ang", "ang", "correct"),
                ("bata", "bata", "correct"),
                ("ay", "ba", "substitution"),
                ("pumunta", "pumunta", "correct"),
            ],
        )

    def test_story_cursor_alignment_preserves_insertion_count(self):
        result = align_story_transcript(
            "Ang bata ay pumunta",
            "Ang bata ay mabilis pumunta",
            start_word_index=0,
        )
        self.assertEqual(result["miscues"], 1)
        self.assertTrue(all(
            item["result"] == "correct"
            for item in result["word_results"]
            if item.get("expected_index") is not None
        ))

    def test_story_alignment_same_callback_wrong_then_correct_is_one_raw_miscue(self):
        result = align_story_transcript("pumunta", "punta pumunta", start_word_index=0)
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(
            [(item["recognized"], item["result"], item["type"]) for item in result["word_results"]],
            [("punta", "miscue", "insertion"), ("pumunta", "correct", "correct")],
        )

    def test_story_alignment_pure_repetition_remains_one_raw_miscue(self):
        result = align_story_transcript("pumunta", "pumunta pumunta", start_word_index=0)
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(
            [(item["recognized"], item["result"], item["type"]) for item in result["word_results"]],
            [("pumunta", "miscue", "insertion"), ("pumunta", "correct", "correct")],
        )

    def test_story_cursor_alignment_preserves_trailing_unread_word_handling(self):
        result = align_story_transcript(
            "Ang bata ay pumunta sa bahay",
            "Ang bata ay pumunta",
            start_word_index=0,
        )
        self.assertEqual(result["miscues"], 0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "correct"), (1, "correct"), (2, "correct"), (3, "correct")],
        )

    def test_story_cursor_relative_alignment_is_independent_of_chunk_boundaries(self):
        expected = "ako ang pinakamabilis tumakbo sa bahay"
        one_chunk = align_story_transcript(expected, "ako ang pinakamabilis tumakbo sa bahay", start_word_index=0)
        split_chunks = [
            align_story_transcript(expected, "ako ang", start_word_index=0),
            align_story_transcript(expected, "pinakamabilis tumakbo", start_word_index=2),
            align_story_transcript(expected, "sa bahay", start_word_index=4),
        ]
        combined = [item for chunk in split_chunks for item in chunk["word_results"]]
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in combined],
            [(item["expected_index"], item["result"]) for item in one_chunk["word_results"]],
        )

    def test_story_cursor_relative_alignment_uses_next_repeated_word(self):
        result = align_story_transcript("ako ay ako rin", "ako rin", start_word_index=2)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(2, "correct"), (3, "correct")],
        )

    def test_story_cursor_relative_short_chunk_does_not_reconsider_word_zero(self):
        result = align_story_transcript("ako ang mabilis wala nang iba", "wala nang", start_word_index=3)
        self.assertEqual([item["expected_index"] for item in result["word_results"]], [3, 4])
        self.assertTrue(all(item["result"] == "correct" for item in result["word_results"]))

    def test_story_cursor_relative_miscue_does_not_mark_unattempted_tail(self):
        result = align_story_transcript(
            "ako ang mabilis tumakbo sa bahay bukas",
            "mabagal tumakbo sa",
            start_word_index=2,
        )
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(2, "miscue"), (3, "correct"), (4, "correct")],
        )
        self.assertNotIn(5, [item["expected_index"] for item in result["word_results"]])

    def test_story_post_miscue_chunk_exposes_trailing_correct_results(self):
        result = align_story_transcript("bibilis pa sa akin", "li pa sa akin", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "miscue"), (1, "correct"), (2, "correct"), (3, "correct")],
        )

    def test_story_post_miscue_chunk_can_contain_another_miscue(self):
        result = align_story_transcript("bibilis pa sa akin", "li bahay sa akin", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "miscue"), (1, "miscue"), (2, "correct"), (3, "correct")],
        )

    def test_story_post_miscue_chunk_preserves_correct_miscue_correct_order(self):
        result = align_story_transcript("una pangatlo huli", "una mali huli", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "correct"), (1, "miscue"), (2, "correct")],
        )

    def test_story_post_miscue_single_word_and_short_chunk_are_unchanged(self):
        single = align_story_transcript("bibilis", "li", start_word_index=0)
        short = align_story_transcript("bibilis pa sa akin", "li", start_word_index=0)
        self.assertEqual([(item["expected_index"], item["result"]) for item in single["word_results"]], [(0, "miscue")])
        self.assertEqual([(item["expected_index"], item["result"]) for item in short["word_results"]], [(0, "miscue")])

    def test_story_post_miscue_repeated_words_use_next_occurrence(self):
        result = align_story_transcript("ako ako rin", "mali ako rin", start_word_index=0)
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in result["word_results"]],
            [(0, "miscue"), (1, "correct"), (2, "correct")],
        )

    def test_story_post_miscue_results_are_equivalent_across_chunk_boundaries(self):
        expected = "bibilis pa sa akin"
        combined = align_story_transcript(expected, "li pa sa akin", start_word_index=0)["word_results"]
        split = (
            align_story_transcript(expected, "li", start_word_index=0)["word_results"]
            + align_story_transcript(expected, "pa sa", start_word_index=1)["word_results"]
            + align_story_transcript(expected, "akin", start_word_index=3)["word_results"]
        )
        self.assertEqual(
            [(item["expected_index"], item["result"]) for item in split],
            [(item["expected_index"], item["result"]) for item in combined],
        )

    def test_story_alignment_detects_omission_and_keeps_following_words_aligned(self):
        result = align_story_transcript(
            "Si Ana ay pumunta sa bahay.",
            "Si Ana pumunta sa bahay.",
            language_code="fil-PH",
        )
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(result["word_results"][2]["expected"], "ay")
        self.assertEqual(result["word_results"][2]["result"], "miscue")
        self.assertEqual(result["word_results"][2]["type"], "omission")
        self.assertEqual(result["word_results"][3]["expected"], "pumunta")
        self.assertEqual(result["word_results"][3]["result"], "correct")

    def test_story_alignment_handles_insertion_without_shifting_later_words(self):
        result = align_story_transcript(
            "Si Ana pumunta sa bahay.",
            "Si Ana ay pumunta sa bahay.",
            language_code="fil-PH",
        )
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(result["word_results"][2]["expected"], "pumunta")
        self.assertEqual(result["word_results"][2]["result"], "correct")
        self.assertEqual(result["word_results"][3]["expected"], "sa")
        self.assertEqual(result["word_results"][3]["result"], "correct")

    def test_story_alignment_normalizes_formatting_noise_and_keeps_genuine_differences_as_miscues(self):
        formatting_result = align_story_transcript(
            "Si Ana ay pumunta.",
            "si   ana   ay   pumunta",
            language_code="fil-PH",
        )
        self.assertEqual(formatting_result["miscues"], 0)

        genuine_result = align_story_transcript(
            "binti",
            "bente",
            language_code="fil-PH",
        )
        self.assertEqual(genuine_result["miscues"], 1)
        self.assertEqual(genuine_result["word_results"][0]["result"], "miscue")
        self.assertEqual(genuine_result["word_results"][0]["type"], "substitution")

    def test_story_alignment_critical_fragmentation_pagong_bagong(self):
        """Test that STT fragmentation artifacts are distinguished from genuine miscues.
        
        Expected: "Pagong"
        Genuine student miscue: "bagong" (different lexical item)
        STT output for miscue: "ba go ng" (fragmented)
        
        MUST mark as miscue, NOT as correct.
        """
        # Case 1: Harmless STT segmentation of correct word
        result = align_story_transcript(
            "Pagong",
            "pa gong",
            language_code="fil-PH",
        )
        self.assertEqual(result["correct_words"], 1, 
            "STT splitting 'Pagong' into 'pa gong' is a harmless segmentation artifact")
        self.assertEqual(result["miscues"], 0)
        self.assertEqual(result["word_results"][0]["result"], "correct")
        
        # Case 2: Genuine student miscue (bagong vs Pagong)
        result = align_story_transcript(
            "Pagong",
            "ba go ng",
            language_code="fil-PH",
        )
        self.assertEqual(result["correct_words"], 0, 
            "STT output 'ba go ng' represents student saying 'bagong', not 'Pagong'")
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(result["word_results"][0]["result"], "miscue")
        self.assertEqual(result["word_results"][0]["type"], "substitution")

    def test_story_alignment_critical_fragmentation_niya_ni(self):
        """Test that student omissions are not masked by STT fragmentation.
        
        Expected: "niya"
        Student says: "ni" (genuine omission)
        STT may fragment this in various ways: "ni", "ni a", etc.
        
        MUST remain a miscue, NOT automatically joined into "niya".
        """
        # Case 1: Simple omission - student says "ni" not "niya"
        result = align_story_transcript(
            "niya",
            "ni",
            language_code="fil-PH",
        )
        self.assertEqual(result["correct_words"], 0,
            "Student said 'ni', not 'niya'")
        self.assertEqual(result["miscues"], 1)
        self.assertEqual(result["word_results"][0]["result"], "miscue")
        
        # Case 2: Fragmented output "ni a" - must NOT be auto-joined into "niya"
        result = align_story_transcript(
            "niya",
            "ni a",
            language_code="fil-PH",
        )
        # The system should recognize these as separate tokens, not auto-join them
        # Result should be 1-2 tokens in recognized, but they should NOT match "niya" exactly
        self.assertEqual(result["correct_words"], 0,
            "Fragmented 'ni a' must not be auto-joined to become 'niya'")
        self.assertGreaterEqual(result["miscues"], 1,
            "At least one word should be marked as miscue")

    def test_story_alignment_mixed_passage_with_segmentation_and_miscue(self):
        """Test passage with both harmless segmentation and genuine miscues.
        
        Expected: "sabi ni Pagong"
        Student reads: "sabi" (correct) + "ni" (correct) + "bagong" (miscue)
        STT returns: "sabi ni ba go ng"
        
        Should align as: sabi=CORRECT, ni=CORRECT, Pagong=MISCUE
        """
        result = align_story_transcript(
            "sabi ni Pagong",
            "sabi ni ba go ng",
            language_code="fil-PH",
        )
        self.assertEqual(result["total_words"], 3)
        self.assertEqual(result["correct_words"], 2,
            "First two words (sabi, ni) are correct")
        self.assertEqual(result["miscues"], 1,
            "Third word (Pagong) is a miscue")
        
        # Verify individual word results
        self.assertEqual(result["word_results"][0]["expected"], "sabi")
        self.assertEqual(result["word_results"][0]["result"], "correct")
        
        self.assertEqual(result["word_results"][1]["expected"], "ni")
        self.assertEqual(result["word_results"][1]["result"], "correct")
        
        self.assertEqual(result["word_results"][2]["expected"], "pagong")
        self.assertEqual(result["word_results"][2]["result"], "miscue")
        self.assertEqual(result["word_results"][2]["type"], "substitution")

    def test_story_alignment_preserves_raw_recognized_text_for_audit(self):
        """Ensure raw STT output is preserved for later audit and debugging."""
        expected = "Pagong"
        recognized = "ba go ng"
        result = align_story_transcript(expected, recognized, language_code="fil-PH")
        
        # The raw input should be preserved in the result
        self.assertEqual(result["recognized_text"], recognized.lower().replace("  ", " "))
        # But the word_results should show the miscue properly
        self.assertEqual(result["word_results"][0]["result"], "miscue")

    def test_story_word_states_mark_only_attempted_words_as_read(self):
        states = story_word_states_from_results(
            "Naku, Kuneho, wala ka nang ibang sinabi",
            "Naku, Pagong",
            total_words=7,
        )
        self.assertEqual(states, [
            "correct",
            "miscue",
            "pending",
            "pending",
            "pending",
            "pending",
            "pending",
        ])

    def test_story_word_states_keep_unattempted_words_pending_after_misread(self):
        states = story_word_states_from_results(
            "ang bata ay naglalaro sa parke",
            "ang guro",
            total_words=6,
        )
        self.assertEqual(states, [
            "correct",
            "miscue",
            "pending",
            "pending",
            "pending",
            "pending",
        ])

    def test_story_word_states_mark_attempted_tail_words_correct_and_keep_rest_pending(self):
        states = story_word_states_from_results(
            "ang bata ay naglalaro sa parke",
            "ang bata ay",
            total_words=6,
        )
        self.assertEqual(states, [
            "correct",
            "correct",
            "correct",
            "pending",
            "pending",
            "pending",
        ])

    def test_story_word_states_generic_for_punctuated_and_non_punctuated_paragraphs(self):
        cases = [
            (
                "ang bata ay naglalaro sa parke",
                "ang guro",
                6,
                ["correct", "miscue", "pending", "pending", "pending", "pending"],
            ),
            (
                "Naku, Kuneho, wala ka nang ibang sinabi.",
                "Naku, Pagong.",
                7,
                ["correct", "miscue", "pending", "pending", "pending", "pending", "pending"],
            ),
            (
                "Maganda ang umaga at masaya ang araw.",
                "Maganda ang umaga at masaya ang buwan.",
                7,
                ["correct", "correct", "correct", "correct", "correct", "correct", "miscue"],
            ),
        ]

        for expected_text, recognized_text, total_words, expected_states in cases:
            with self.subTest(expected_text=expected_text, recognized_text=recognized_text):
                states = story_word_states_from_results(
                    expected_text,
                    recognized_text,
                    total_words=total_words,
                )
                self.assertEqual(states, expected_states)

    def test_story_word_states_use_dynamic_word_results_current_index(self):
        expected_text = "Bata, huminga at maglakad sa parke."
        states = story_word_states_from_results(
            expected_text,
            total_words=7,
            word_results=[
                {"expected_index": 0, "result": "correct"},
                {"expected_index": 1, "result": "miscue"},
                {"expected_index": 2, "result": "correct"},
            ],
        )
        self.assertEqual(states, [
            "correct",
            "miscue",
            "correct",
            "pending",
            "pending",
            "pending",
            "pending",
        ])

class AdaptedReadingLevelTests(TestCase):

    def test_adapted_reading_level_label_uses_expected_thresholds(self):
        self.assertEqual(_adapted_reading_level_label(0.90), "Readers at Grade Level")
        self.assertEqual(_adapted_reading_level_label(0.76), "Transitioning Readers")
        self.assertEqual(_adapted_reading_level_label(0.60), "Developing Readers")
        self.assertEqual(_adapted_reading_level_label(0.45), "High Emerging Readers")
        self.assertEqual(_adapted_reading_level_label(0.25), "Low Emerging Readers")

    def test_adapted_reading_level_averages_across_assessment_types(self):
        result = _adapted_reading_level_from_attempts([
            {"total_score": 80, "assessment_type": "word"},
            {"total_score": 80, "assessment_type": "sentence"},
            {"total_score": 80, "assessment_type": "paragraph"},
        ])

        self.assertEqual(result["adapted_level_score"], 0.76)
        self.assertEqual(result["adapted_reading_level"], "Transitioning Readers")
        self.assertEqual(result["adapted_reading_level_disclaimer"], "Great job completing your reading assessment! Your results show your current reading performance. Keep practicing to improve your reading skills.")

    def test_assessment_fluency_score_is_more_forgiving_for_accurate_slow_readers(self):
        self.assertEqual(_assessment_fluency_score(0.10, 95), 52)
        self.assertEqual(_assessment_fluency_score(0.20, 95), 60)

    def test_display_reading_level_uses_consistent_classification_labels(self):
        self.assertEqual(_display_reading_level("Transitioning", None), "Transitioning Readers")
        self.assertEqual(_display_reading_level(None, {"final_score": 60}), "High Emerging Readers")

    def test_display_reading_level_prefers_score_classification_over_stale_adapted_labels(self):
        self.assertEqual(
            _display_reading_level(None, {"final_score": 81, "adapted_reading_level": "Low Emerging Readers"}),
            "Transitioning Readers",
        )

    def test_build_latest_reading_level_payload_uses_latest_score_classification(self):
        payload = _build_latest_reading_level_payload({"total_score": 81, "assessment_type": "word"}, fallback="Low Emerging Readers")

        self.assertEqual(payload["reading_level"], "Transitioning Readers")
        self.assertEqual(payload["adapted_reading_level"], "Transitioning Readers")

    def test_assessment_score_payload_uses_word_multiplier_for_high_emerging_levels(self):
        result = _assessment_score_payload({
            "scores": {
                "fluency_score": 0,
                "accuracy": 0,
                "pronunciation_score": 0,
                "time_score": 0,
                "total_score": 56,
            },
            "assessment_type": "word",
        })

        self.assertEqual(result["adapted_level_score"], 0.5)
        self.assertEqual(result["adapted_reading_level"], "High Emerging Readers")

    def test_assessment_score_payload_uses_word_multiplier_for_lower_high_emerging_levels(self):
        result = _assessment_score_payload({
            "scores": {
                "fluency_score": 0,
                "accuracy": 0,
                "pronunciation_score": 0,
                "time_score": 0,
                "total_score": 47,
            },
            "assessment_type": "word",
        })

        self.assertEqual(result["adapted_level_score"], 0.42)
        self.assertEqual(result["adapted_reading_level"], "High Emerging Readers")

    def test_assessment_score_payload_uses_weighted_total_and_interpretation(self):
        result = _assessment_score_payload({
            "scores": {
                "fluency_score": 6,
                "accuracy": 60,
                "pronunciation_score": 60,
                "time_score": 0,
                "total_score": 31,
            },
            "assessment_type": "word",
        })

        self.assertEqual(result["overall_raw_score"], 49)
        self.assertEqual(result["final_score"], 44)
        self.assertEqual(result["total_score"], 44)
        self.assertFalse(result["passed"])
        self.assertEqual(result["performance_interpretation"], "Needs Support")
        self.assertEqual(result["adapted_reading_level"], "High Emerging Readers")

    def test_assessment_score_payload_uses_vowel_osps_multiplier_for_classification(self):
        result = _assessment_score_payload({
            "scores": {
                "fluency_score": 84,
                "accuracy": 84,
                "pronunciation_score": 84,
                "time_score": 0,
                "total_score": 84,
            },
            "assessment_type": "vowel",
        })

        self.assertEqual(result["overall_raw_score"], 80)
        self.assertEqual(result["final_score"], 68)
        self.assertEqual(result["crla_classification"], "High Emerging Readers")
        self.assertEqual(result["adapted_level_score"], 0.68)

    def test_assessment_score_payload_uses_vowel_multiplier_for_vc_materials(self):
        result = _assessment_score_payload({
            "scores": {
                "fluency_score": 84,
                "accuracy": 84,
                "pronunciation_score": 84,
                "time_score": 0,
                "total_score": 84,
            },
            "assessment_type": "vc",
        })

        self.assertEqual(result["osps_multiplier"], 0.85)
        self.assertEqual(result["final_score"], 68)


class CentralizedAssessmentScoringTests(TestCase):
    def test_build_assessment_score_payload_computes_time_score_from_pace(self):
        payload = build_assessment_score_payload({
            "assessment_type": "word",
            "correct_words": 45,
            "incorrect_words": 5,
            "skipped_words": 0,
            "duration_seconds": 60,
            "target_word_count": 50,
            "pronunciation_metrics": {"score": 80},
            "fluency_metrics": {"score": 70},
        })

        self.assertEqual(payload["time_score"], 100.0)
        self.assertEqual(payload["overall_raw_score"], 87)
        self.assertEqual(payload["final_score"], 78)

    def test_build_assessment_score_payload_uses_weighted_formula_for_all_metrics(self):
        payload = build_assessment_score_payload({
            "assessment_type": "word",
            "correct_words": 80,
            "incorrect_words": 20,
            "skipped_words": 0,
            "duration_seconds": 60,
            "target_word_count": 100,
            "pronunciation_metrics": {"score": 40},
            "fluency_metrics": {"score": 60},
            "time_score": 20,
        })

        self.assertEqual(payload["accuracy"], 80.0)
        self.assertEqual(payload["fluency_score"], 60.0)
        self.assertEqual(payload["pronunciation_score"], 40.0)
        self.assertEqual(payload["time_score"], 20.0)
        self.assertEqual(payload["overall_raw_score"], 74)
        self.assertEqual(payload["final_score"], 67)

    def test_build_assessment_score_payload_uses_raw_metrics_for_authoritative_score(self):
        payload = build_assessment_score_payload({
            "assessment_type": "word",
            "correct_words": 100,
            "incorrect_words": 0,
            "skipped_words": 0,
            "duration_seconds": 60,
            "target_word_count": 100,
            "pronunciation_metrics": {"score": 62.5},
            "fluency_metrics": {"score": 60},
        })

        self.assertEqual(payload["accuracy"], 100.0)
        self.assertEqual(payload["fluency_score"], 60.0)
        self.assertEqual(payload["pronunciation_score"], 62.5)
        self.assertEqual(payload["final_score"], 81)
        self.assertEqual(payload["crla_classification"], "Transitioning Readers")

    def test_build_assessment_score_payload_handles_missing_pronunciation_data(self):
        payload = build_assessment_score_payload({
            "assessment_type": "paragraph",
            "correct_words": 12,
            "incorrect_words": 8,
            "skipped_words": 0,
            "duration_seconds": 90,
            "target_word_count": 20,
            "fluency_metrics": {"score": 80},
        })

        self.assertEqual(payload["accuracy"], 60.0)
        self.assertEqual(payload["pronunciation_score"], 0.0)
        self.assertEqual(payload["final_score"], 55)
        self.assertEqual(payload["crla_classification"], "Low Emerging Readers")

    def test_build_assessment_score_payload_uses_zero_fluency_for_skipped_assessment(self):
        payload = build_assessment_score_payload({
            "assessment_type": "word",
            "correct_words": 0,
            "incorrect_words": 0,
            "skipped_words": 0,
            "duration_seconds": 0,
            "target_word_count": 20,
            "transcript": "",
            "speech_recognition_used": False,
            "needs_manual_review": False,
        })

        self.assertEqual(payload["fluency_score"], 0.0)
        self.assertEqual(payload["overall_raw_score"], 0)
        self.assertEqual(payload["final_score"], 0)
        self.assertEqual(payload["crla_classification"], "Low Emerging Readers")


class HuntScoringRuleTests(TestCase):
    def test_normalization_and_missing_confidence_fallback(self):
        self.assertEqual(normalize_speech(" C\u00c1t! "), "cat")
        self.assertEqual(classify_speech("Cat!", "cat"), ("Excellent", 2))
        self.assertEqual(classify_speech("dog", "cat"), ("Weak", 0))

    def test_confidence_thresholds(self):
        self.assertEqual(classify_speech("cat", "cat", .80), ("Excellent", 2))
        self.assertEqual(classify_speech("cat", "cat", .79), ("Mixed", 1))
        self.assertEqual(classify_speech("cat", "cat", .50), ("Mixed", 1))
        self.assertEqual(classify_speech("cat", "cat", .49), ("Weak", 0))
        self.assertEqual(classify_speech("dog", "cat", .99), ("Weak", 0))

    def test_star_thresholds_never_award_zero(self):
        self.assertEqual([stars_for_points(p) for p in (0, 4, 5, 7, 8, 10)], [1, 1, 2, 2, 3, 3])

    def test_frontend_has_duplicate_guard_and_non_scoring_checkpoint(self):
        source = (Path(__file__).resolve().parent / "static" / "pabasa_app" / "js" / "practice_reader.js").read_text(encoding="utf-8")
        template = (Path(__file__).resolve().parent / "templates" / "pabasa_app" / "practice_reader_base.html").read_text(encoding="utf-8")
        self.assertIn("if (!isHuntMode || huntResults[index]) return null", source)
        self.assertIn("if (currentIndex === 3 && huntCheckpointToast)", source)
        self.assertIn("if (result.points > 0)", source)
        self.assertIn("speechItemIndex !== currentIndex || huntAdvanceInProgress", source)
        self.assertIn("Try again — keep reading the same word.", source)
        self.assertIn("if (!huntResults[currentIndex]) finalizeHuntSpeechResult", source)
        self.assertIn("updateReadingToggleButton()", source)
        self.assertIn("practiceRecordingWindowMs(targetText)", source)
        self.assertIn("huntListeningDesired", source)
        self.assertIn("scheduleContinuousRecognition(currentIndex, 120)", source)
        self.assertIn("stopContinuousRecognitionByUser()", source)
        self.assertIn("Google Speech results will appear here while you read.", template)
        self.assertIn("Raw mic input", template)
        self.assertIn("Waiting for speech...", template)
        self.assertIn('id="huntReadAloudBtn"', template)
        self.assertIn('/api/reading/read-aloud/', source)
        self.assertIn('formData.append("tts_profile", "hunt")', source)
        self.assertIn("activeDot.appendChild(huntFlightBird)", source)
        self.assertIn("grid-template-columns: .72fr 1.12fr 1fr .72fr", template)
        self.assertIn("transform: translate(-50%,-50%)", template)
        self.assertIn("/api/practice/hunt/award-stars/", source)
        self.assertIn("if (huntAwardSubmitted)", source)
        self.assertIn('id="huntPointsDisplay">Points: 0/10', template)
        self.assertIn('Available Stars: {{ student_available_stars|default:0 }}', template)
        self.assertIn('starCount.textContent = `Available Stars: ${data.available_stars}`', source)

    def test_frontend_normalizes_prefixed_material_ids_and_accepts_legacy_game_types(self):
        source = (Path(__file__).resolve().parent / "static" / "pabasa_app" / "js" / "practice_reader.js").read_text(encoding="utf-8")
        self.assertIn("function normalizeMaterialId(value)", source)
        self.assertIn("mId === normalizeMaterialId(materialId)", source)
        self.assertIn("materialGameMode === selectedGameMode", source)
        self.assertIn("if (!itemTypeMatches && !legacyGameModeMatches) return false", source)

    def test_hunt_reader_preserves_the_treasure_map_in_dark_mode(self):
        template = (Path(__file__).resolve().parent / "templates" / "pabasa_app" / "practice_reader_base.html").read_text(encoding="utf-8")
        self.assertIn("hunt-treasure-map.svg", template)
        self.assertIn("body.dark-theme .practice-hunt-shell .hunt-card", template)
        self.assertIn("background-image:linear-gradient", template)


class StudentSignupCustomIdTests(TestCase):
    def setUp(self):
        self.client = Client()

    def _set_pending_student_signup(self, grade_level, email):
        session = self.client.session
        session["pending_student_signup"] = {
            "first_name": "Jamie",
            "last_name": "Reader",
            "email": email,
            "middle_initial": "",
            "suffix": "",
            "sex": "female",
            "birth_month": 1,
            "birth_day": 15,
            "birth_year": 2014,
            "password_hash": make_password("student-password"),
            "contact_no": "",
            "grade_level": grade_level,
            "section": "",
            "reading_level": "",
        }
        session["pending_student_signup_otp"] = "123456"
        session["pending_student_signup_otp_created"] = timezone.now().timestamp()
        session.save()

    @patch("pabasa_app.views._notify_admins")
    @patch("pabasa_app.views.send_student_confirmation_email")
    def test_verify_student_otp_uses_selected_grade_for_custom_id_prefix(self, mock_email, mock_notify):
        self._set_pending_student_signup("Grade 6", "grade6@example.com")

        response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["custom_id"], "G6-0001")
        self.assertEqual(User.objects.get(email="grade6@example.com").custom_id, "G6-0001")

    @patch("pabasa_app.views._notify_admins")
    @patch("pabasa_app.views.send_student_confirmation_email")
    def test_verify_student_otp_increments_custom_id_per_grade_prefix(self, mock_email, mock_notify):
        User.objects.create(
            custom_id="G3-0001",
            role="student",
            first_name="Gia",
            last_name="Three",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=2,
            birth_day=2,
            birth_year=2013,
            email="existing-g3@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 3",
        )
        User.objects.create(
            custom_id="G6-0001",
            role="student",
            first_name="Gino",
            last_name="Six",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=3,
            birth_day=3,
            birth_year=2012,
            email="existing-g6@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 6",
        )

        self._set_pending_student_signup("Grade 3", "next-g3@example.com")
        response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["custom_id"], "G3-0002")
        self.assertEqual(User.objects.get(email="next-g3@example.com").custom_id, "G3-0002")

    def test_similar_wrong_word_does_not_match_when_first_sound_differs(self):
        result = analyze_reading("house", 0, "mouse")

        self.assertEqual(result["matched"], 0)
        self.assertEqual(result["correct_word_count"], 0)
        self.assertFalse(result["complete"])

    def test_correct_words_advance_in_order_until_first_missing_target(self):
        result = analyze_reading("the water is cold", 0, "the apple is cold")

        self.assertEqual(result["correct_word_count"], 1)
        self.assertFalse(result["complete"])


class SentenceReadingWordResultTests(TestCase):
    target = "the water is cold"

    def test_all_words_correct_receive_one_point_each(self):
        result = analyze_sentence_reading(self.target, "the water is cold")
        self.assertTrue(result["complete"])
        self.assertEqual(result["correct_word_count"], 4)
        self.assertEqual([item["result"] for item in result["word_results"]], ["correct"] * 4)

    def test_wrong_word_is_pending_until_speech_moves_forward(self):
        pending = analyze_sentence_reading(self.target, "the apple")
        self.assertEqual(pending["word_results"][1]["result"], "pending")
        result = analyze_sentence_reading(self.target, "is cold", pending["word_results"])
        self.assertTrue(result["complete"])
        self.assertEqual(result["word_results"][1]["result"], "miscue")
        self.assertEqual(result["word_results"][1]["points"], 0)

    def test_immediate_correction_receives_point_without_miscue(self):
        result = analyze_sentence_reading(self.target, "the apple water is cold")
        corrected = result["word_results"][1]
        self.assertTrue(result["complete"])
        self.assertTrue(corrected["correct"])
        self.assertTrue(corrected["self_corrected"])
        self.assertFalse(corrected["miscue"])
        self.assertEqual(corrected["points"], 1)

    def test_multiple_miscues_are_recorded_individually(self):
        result = analyze_sentence_reading(self.target, "the is warm")
        result = analyze_sentence_reading(self.target, "again", result["word_results"])
        self.assertTrue(result["complete"])
        self.assertEqual([item["result"] for item in result["word_results"]], [
            "correct", "miscue", "correct", "miscue",
        ])

    def test_mixed_results_keep_self_correction_and_miscue_separate(self):
        result = analyze_sentence_reading(self.target, "the apple water cold")
        self.assertTrue(result["complete"])
        self.assertEqual(result["word_results"][1]["self_corrected"], True)
        self.assertEqual(result["word_results"][2]["result"], "miscue")
        self.assertEqual(result["correct_word_count"], 3)

    def test_silence_after_pending_final_word_confirms_miscue(self):
        pending = analyze_sentence_reading(self.target, "the water is warm")
        self.assertFalse(pending["complete"])
        result = analyze_sentence_reading(self.target, "", pending["word_results"])
        self.assertTrue(result["complete"])
        self.assertEqual(result["word_results"][-1]["result"], "miscue")

    def test_crla_clean_sentence_regression(self):
        result = analyze_sentence_reading(
            "naglalaba si tatay sa palanggana",
            "naglalaba si tatay sa palanggana",
            language_code="fil-PH",
        )
        self.assertEqual((result["correct_word_count"], result["miscues"]), (5, 0))

    def test_crla_intentional_miscue_regression(self):
        result = analyze_sentence_reading(
            "magpapalit ako ng kamiseta mamaya",
            "magpapalit ako ng damit mamaya",
            language_code="fil-PH",
        )
        self.assertEqual((result["correct_word_count"], result["miscues"]), (4, 1))
        self.assertEqual(
            [item["result"] for item in result["word_results"]],
            ["correct", "correct", "correct", "miscue", "correct"],
        )

    def test_crla_split_word_stt_reconstructs_only_current_target(self):
        result = analyze_sentence_reading(
            "nilinis nila ang agiw rito",
            "ni li ni si nila ang agiw rito",
            language_code="fil-PH",
        )
        self.assertEqual((result["correct_word_count"], result["miscues"]), (5, 0))
        self.assertEqual(
            [item["result"] for item in result["word_results"]],
            ["correct", "correct", "correct", "correct", "correct"],
        )
        self.assertEqual(result["word_results"][0]["type"], "split_token_reconstruction")


class StudentLrnTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.signup_data = {
            "first_name": "Lia",
            "last_name": "Santos",
            "email": "lia@example.com",
            "password": "Student123",
            "confirm_password": "Student123",
            "sex": "female",
            "birth_month": "1",
            "birth_day": "15",
            "birth_year": "2014",
            "grade_level": "Grade 2",
        }

    def test_student_signup_requires_exactly_twelve_digit_lrn(self):
        for lrn in ("", "12345678901", "1234567890123", "12345678901A"):
            with self.subTest(lrn=lrn):
                response = self.client.post(
                    reverse("register_student"),
                    {**self.signup_data, "lrn": lrn},
                )
                self.assertEqual(response.status_code, 400)

    def test_lrn_is_saved_after_otp_verification(self):
        session = self.client.session
        session["pending_student_signup"] = {
            **self.signup_data,
            "password_hash": make_password(self.signup_data["password"]),
            "lrn": "123456789012",
            "contact_no": "",
            "section": "",
            "reading_level": "",
        }
        session["pending_student_signup_otp"] = "123456"
        session["pending_student_signup_otp_created"] = timezone.now().timestamp()
        session.save()

        with patch("pabasa_app.views.send_student_confirmation_email"), patch("pabasa_app.views._notify_admins"):
            response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(User.objects.get(email="lia@example.com").lrn, "123456789012")

    def test_student_signup_page_exposes_required_lrn_field(self):
        response = self.client.get(reverse("student_signup"))
        self.assertContains(response, 'name="lrn"')
        self.assertContains(response, 'pattern="[0-9]{12}"')


class StudentSignupFlowTests(TestCase):
    def setUp(self):
        self.client = Client()
        self.base_payload = {
            "first_name": "Mia",
            "last_name": "Rivera",
            "email": "mia.new@example.com",
            "password": "Student123",
            "confirm_password": "Student123",
            "lrn": "123456789012",
            "sex": "female",
            "birth_month": "1",
            "birth_day": "5",
            "birth_year": "2014",
            "grade_level": "Grade 3",
            "section": "BONIFACIO",
        }

    @patch("pabasa_app.views.send_student_signup_otp_email")
    def test_new_student_signup_succeeds_and_creates_pending_otp(self, mock_email):
        response = self.client.post(reverse("register_student"), self.base_payload)

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["email"], self.base_payload["email"])
        session = self.client.session
        self.assertIn("pending_student_signup", session)
        self.assertIn("pending_student_signup_otp", session)
        self.assertIn("pending_student_signup_otp_created", session)

    @patch("pabasa_app.views.send_student_signup_otp_email")
    def test_existing_email_blocks_registration(self, mock_email):
        User.objects.create(
            custom_id="STU-EXIST-1",
            role="student",
            first_name="Existing",
            last_name="Email",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2014,
            email="mia.new@example.com",
            password_hash=make_password("student-password"),
        )
        payload = dict(self.base_payload)
        payload["email"] = "mia.new@example.com"
        payload["lrn"] = "123456789013"

        response = self.client.post(reverse("register_student"), payload)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "Email already registered")

    @patch("pabasa_app.views.send_student_signup_otp_email")
    def test_existing_lrn_blocks_registration(self, mock_email):
        User.objects.create(
            custom_id="STU-EXIST-2",
            role="student",
            first_name="Existing",
            last_name="LRN",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2014,
            email="existing-lrn@example.com",
            password_hash=make_password("student-password"),
            lrn="123456789012",
        )
        payload = dict(self.base_payload)
        payload["email"] = "different@example.com"

        response = self.client.post(reverse("register_student"), payload)

        self.assertEqual(response.status_code, 400)
        self.assertEqual(response.json()["error"], "LRN is already registered")

    @patch("pabasa_app.views.send_student_signup_otp_email")
    def test_existing_custom_id_blocks_registration_via_otp_creation(self, mock_email):
        User.objects.create(
            custom_id="G3-0001",
            role="student",
            first_name="Existing",
            last_name="Custom",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2014,
            email="existing-custom@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 3",
        )
        payload = dict(self.base_payload)
        payload["email"] = "custom-check@example.com"

        with patch("pabasa_app.views.generate_custom_id", return_value="G3-0001"), patch(
            "pabasa_app.views.send_student_confirmation_email"
        ), patch("pabasa_app.views._notify_admins"):
            session = self.client.session
            session["pending_student_signup"] = {
                "first_name": payload["first_name"],
                "last_name": payload["last_name"],
                "email": payload["email"],
                "middle_initial": "",
                "suffix": "",
                "sex": payload["sex"],
                "birth_month": int(payload["birth_month"]),
                "birth_day": int(payload["birth_day"]),
                "birth_year": int(payload["birth_year"]),
                "password_hash": make_password(payload["password"]),
                "contact_no": "",
                "lrn": payload["lrn"],
                "grade_level": payload["grade_level"],
                "section": "",
                "reading_level": "",
            }
            session["pending_student_signup_otp"] = "123456"
            session["pending_student_signup_otp_created"] = timezone.now().timestamp()
            session.save()
            response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 400)
        self.assertIn("already exists", response.json()["error"])

    @patch("pabasa_app.views.send_student_confirmation_email")
    @patch("pabasa_app.views._notify_admins")
    def test_verify_student_otp_succeeds_with_valid_pending_signup(self, mock_notify, mock_email):
        session = self.client.session
        session["pending_student_signup"] = {
            "first_name": "Mia",
            "last_name": "Rivera",
            "email": "mia.verify@example.com",
            "middle_initial": "",
            "suffix": "",
            "sex": "female",
            "birth_month": 1,
            "birth_day": 5,
            "birth_year": 2014,
            "password_hash": make_password("Student123"),
            "contact_no": "",
            "lrn": "123456789014",
            "grade_level": "Grade 3",
            "section": "",
            "reading_level": "",
        }
        session["pending_student_signup_otp"] = "123456"
        session["pending_student_signup_otp_created"] = timezone.now().timestamp()
        session.save()

        response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertTrue(User.objects.filter(email="mia.verify@example.com").exists())

    def test_verify_student_otp_reports_clear_error_when_pending_signup_is_missing(self):
        response = self.client.post(reverse("verify_student_otp"), {"otp": "123456"})

        self.assertEqual(response.status_code, 400)
        self.assertIn("registration session", response.json()["error"])


class PracticeProgressionTests(TestCase):
    def test_apply_progression_unlock_override_only_marks_ui_hint_without_changing_state(self):
        progression = {
            "sections": [
                {
                    "difficulty": "easy",
                    "levels": [
                        {"difficulty": "easy", "level": "level_1", "state": "locked", "unlocked": False, "button_label": "Locked"},
                        {"difficulty": "easy", "level": "level_2", "state": "locked", "unlocked": False, "button_label": "Locked"},
                    ],
                }
            ]
        }

        updated = _apply_progression_unlock_override(progression, "easy_level_2")
        levels = updated["sections"][0]["levels"]

        self.assertEqual(levels[0]["state"], "locked")
        self.assertFalse(levels[0]["unlocked"])
        self.assertEqual(levels[1]["state"], "locked")
        self.assertFalse(levels[1]["unlocked"])
        self.assertEqual(updated["ui_unlock_target"], "easy_level_2")


class SharedMaterialImportTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=5,
            birth_day=10,
            birth_year=1988,
            email="shared-import-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.other_teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}",
            role="teacher",
            first_name="Mina",
            last_name="Shared",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=7,
            birth_day=4,
            birth_year=1991,
            email="shared-source-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.shared_material = Material.objects.create(
            teacher=self.other_teacher,
            title="Shared Reading",
            item_type="word",
            prompt_text="One",
            content_text="One\nTwo",
            content_json={"items": ["One", "Two"], "language": "English"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )
        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session["first_name"] = self.teacher.first_name
        session["last_name"] = self.teacher.last_name
        session["email"] = self.teacher.email
        session["custom_id"] = self.teacher.custom_id
        session.save()

    def test_importing_shared_material_without_changes_reuses_original(self):
        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared Reading",
                "reading_type": "word",
                "content": "One\nTwo",
                "status": "published",
                "usage_type": "assessment",
                "source_type": "shared",
                "source_material_id": self.shared_material.id,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertTrue(data.get("reused", False))
        self.assertEqual(Material.objects.filter(teacher=self.teacher).count(), 0)
        self.assertEqual(Material.objects.filter(source_type="shared", teacher=self.other_teacher).count(), 1)

    def test_importing_shared_material_with_changes_creates_updated_duplicate(self):
        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared Reading",
                "reading_type": "word",
                "content": "One\nTwo\nThree",
                "status": "draft",
                "usage_type": "assessment",
                "source_type": "shared",
                "source_material_id": self.shared_material.id,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertFalse(data.get("reused", False))
        duplicate_materials = Material.objects.filter(teacher=self.teacher, source_type="shared")
        self.assertEqual(duplicate_materials.count(), 1)
        duplicate_material = duplicate_materials.get()
        self.assertTrue(duplicate_material.title.startswith("[UPDATED]"))
        self.assertEqual(duplicate_material.status, "draft")


class OcrLayoutGroupingTests(TestCase):
    def test_build_material_items_from_ocr_layout_returns_words_in_reading_order(self):
        layout = [
            {"text": "Hello", "left": 10, "top": 20, "width": 40, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 0},
            {"text": "world", "left": 60, "top": 20, "width": 40, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 1},
        ]

        items = _build_material_items_from_ocr_layout(layout, "word")

        self.assertEqual(items, ["Hello", "world"])

    def test_build_material_items_from_ocr_layout_groups_lines_into_sentences(self):
        layout = [
            {"text": "The", "left": 10, "top": 20, "width": 20, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 0},
            {"text": "quick", "left": 40, "top": 20, "width": 30, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 1},
            {"text": "brown", "left": 80, "top": 20, "width": 30, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 2},
            {"text": "fox", "left": 120, "top": 20, "width": 20, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 3},
            {"text": "jumps", "left": 10, "top": 45, "width": 35, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 1, "word_num": 0},
            {"text": "over", "left": 50, "top": 45, "width": 25, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 1, "word_num": 1},
            {"text": "the", "left": 80, "top": 45, "width": 20, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 1, "word_num": 2},
            {"text": "lazy", "left": 105, "top": 45, "width": 25, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 1, "word_num": 3},
            {"text": "dog", "left": 135, "top": 45, "width": 20, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 1, "word_num": 4},
        ]

        items = _build_material_items_from_ocr_layout(layout, "sentence")

        self.assertEqual(items, ["The quick brown fox", "jumps over the lazy dog"])

    def test_build_material_items_from_ocr_layout_groups_paragraphs_by_vertical_gap(self):
        layout = [
            {"text": "First", "left": 10, "top": 20, "width": 30, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 0},
            {"text": "paragraph", "left": 45, "top": 20, "width": 50, "height": 12, "conf": 95, "block_num": 0, "par_num": 0, "line_num": 0, "word_num": 1},
            {"text": "Second", "left": 10, "top": 70, "width": 34, "height": 12, "conf": 95, "block_num": 0, "par_num": 1, "line_num": 0, "word_num": 0},
            {"text": "paragraph", "left": 50, "top": 70, "width": 50, "height": 12, "conf": 95, "block_num": 0, "par_num": 1, "line_num": 0, "word_num": 1},
        ]

        items = _build_material_items_from_ocr_layout(layout, "paragraph")

        self.assertEqual(items, ["First paragraph", "Second paragraph"])


class MaterialUploadExtractionTests(TestCase):
    def test_build_image_upload_debug_info_reports_upload_size_and_hash(self):
        upload = SimpleUploadedFile("scan.png", b"abc123", content_type="image/png")

        info = _build_image_upload_debug_info(upload, source="received")

        self.assertEqual(info["source"], "received")
        self.assertEqual(info["size"], 6)
        self.assertEqual(info["sha256"], hashlib.sha256(b"abc123").hexdigest())
        self.assertEqual(info["content_type"], "image/png")

    def setUp(self):
        self.teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=5,
            birth_day=10,
            birth_year=1988,
            email="upload-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session["first_name"] = self.teacher.first_name
        session["last_name"] = self.teacher.last_name
        session["email"] = self.teacher.email
        session["custom_id"] = self.teacher.custom_id
        session.save()

    def test_extract_endpoint_honors_selected_pdf_pages(self):
        buffer = BytesIO()
        pdf_canvas = canvas.Canvas(buffer)
        pdf_canvas.drawString(72, 720, "Intro page")
        pdf_canvas.showPage()
        pdf_canvas.drawString(72, 720, "Page 2")
        pdf_canvas.showPage()
        pdf_canvas.drawString(72, 720, "Last page")
        pdf_canvas.save()
        buffer.seek(0)

        pdf_file = SimpleUploadedFile(
            "sample.pdf",
            buffer.read(),
            content_type="application/pdf",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": pdf_file, "selection_mode": "selected", "selected_pages": "2"},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["page_count"], 3)
        self.assertEqual(data["selected_pages"], [2])
        self.assertEqual(data["items"], ["Page 2"])

    @patch("pabasa_app.views._extract_text_from_image", return_value="")
    def test_extract_endpoint_returns_empty_items_without_warning_when_image_ocr_detects_no_text(self, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], [])
        self.assertEqual(data.get("warnings", []), [])
        self.assertEqual(data.get("warning_message", ""), "")
        mock_extract_text_from_image.assert_called_once()

    @patch("pabasa_app.views._extract_text_from_image", return_value="Alpha beta gamma")
    def test_extract_endpoint_exposes_alias_payload_fields_for_upload_ui(self, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], ["Alpha", "beta", "gamma"])
        self.assertEqual(data["extracted_items"], ["Alpha", "beta", "gamma"])
        self.assertEqual(data["extractedItems"], ["Alpha", "beta", "gamma"])
        mock_extract_text_from_image.assert_called_once()

    @patch("pabasa_app.views._extract_text_from_image", return_value="Alpha beta gamma")
    @patch("pabasa_app.views._build_extracted_material_items", return_value=("word", []))
    def test_extract_endpoint_returns_warning_response_when_extracted_text_cannot_be_split(self, mock_build_extracted_material_items, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], ["Alpha", "beta", "gamma"])
        self.assertTrue(any("could not be converted" in warning.lower() for warning in data.get("warnings", [])))
        mock_extract_text_from_image.assert_called_once()
        mock_build_extracted_material_items.assert_called_once()

    @patch("pabasa_app.views._build_extracted_material_items", side_effect=RuntimeError("boom"))
    def test_extract_endpoint_returns_warning_response_when_item_building_fails(self, mock_build_extracted_material_items):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], [])
        self.assertTrue(any("could not be processed" in warning.lower() for warning in data.get("warnings", [])))
        mock_build_extracted_material_items.assert_called_once()

    @patch("pabasa_app.views._extract_text_from_image", return_value="Alpha beta gamma")
    @patch("pabasa_app.views._build_extracted_material_items", return_value=("word", []))
    def test_extract_endpoint_falls_back_to_text_items_when_server_returns_no_items(self, mock_build_extracted_material_items, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], ["Alpha", "beta", "gamma"])
        mock_extract_text_from_image.assert_called_once()
        mock_build_extracted_material_items.assert_called_once()

    @patch("pabasa_app.views._extract_text_from_image", return_value={"text": "Line one\nLine two\n\nLine three", "layout": []})
    def test_extract_endpoint_preserves_newlines_for_ocr_text(self, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertIn("Line one\nLine two", data["text"])
        self.assertIn("\n\nLine three", data["text"])
        mock_extract_text_from_image.assert_called_once()

    @patch("pabasa_app.views._extract_text_from_image", return_value={"text": "", "layout": []})
    def test_extract_endpoint_returns_warning_when_image_ocr_yields_no_text(self, mock_extract_text_from_image):
        image_file = SimpleUploadedFile(
            "scan.png",
            b"not-a-real-image",
            content_type="image/png",
        )

        response = self.client.post(
            reverse("extract_reading_material_file"),
            {"file": image_file},
            format="multipart",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["items"], [])
        self.assertTrue(data.get("warnings") or data.get("warning_message"))
        mock_extract_text_from_image.assert_called_once()

    def test_fallback_material_items_preserve_paragraph_blocks(self):
        text = "First line\nSecond line\n\nThird line"
        self.assertEqual(_fallback_material_items_from_text(text), ["First line Second line", "Third line"])


class PrincipalReportsExportTests(TestCase):
    def test_default_timezone_is_asia_manila(self):
        self.assertEqual(settings.TIME_ZONE, "Asia/Manila")

    def setUp(self):
        self.user = User.objects.create(
            custom_id=f"ADM-{uuid.uuid4().hex[:8].upper()}",
            role="admin",
            first_name="Principal",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="principal@example.com",
            password_hash="hashed-password",
        )
        session = self.client.session
        session["user_id"] = self.user.id
        session["user_role"] = self.user.role
        session["first_name"] = self.user.first_name
        session["last_name"] = self.user.last_name
        session["email"] = self.user.email
        session["custom_id"] = self.user.custom_id
        session.save()

    def test_principal_reports_pdf_export_returns_pdf_response(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "school", "export": "pdf"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "application/pdf")
        self.assertTrue(response.content.startswith(b"%PDF"))

    def test_principal_reports_pdf_export_includes_summary_overview(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "school", "export": "pdf"})

        self.assertEqual(response.status_code, 200)
        reader = PdfReader(BytesIO(response.content))
        extracted_text = "\n".join(page.extract_text() or "" for page in reader.pages)

        self.assertTrue(extracted_text or response.content.startswith(b"%PDF"))


class PrincipalReportsPreviewTests(TestCase):
    def setUp(self):
        unique_suffix = uuid.uuid4().hex[:8].upper()
        self.principal = User.objects.create(
            custom_id=f"PRN-{unique_suffix}",
            role="principal",
            first_name="Jobelyn",
            last_name="Valdez",
            middle_initial="A",
            suffix="",
            sex="female",
            birth_month=6,
            birth_day=3,
            birth_year=1980,
            email="principal-preview@example.com",
            password_hash=make_password("Principal@123"),
        )
        self.teacher = User.objects.create(
            custom_id=f"TCH-{unique_suffix}",
            role="teacher",
            first_name="Rowan",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="teacher-preview@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.student = User.objects.create(
            custom_id=f"STD-{unique_suffix}",
            role="student",
            first_name="Ava",
            last_name="Learner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=2,
            birth_day=2,
            birth_year=2012,
            email="student-preview@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
        )
        self.section = test_section_create(
            class_code=f"G2-{unique_suffix}",
            class_name="Grade 2 Preview",
            header="Reading Class",
            description="Preview section",
            teacher=self.teacher,
            is_active=True,
            subject="Reading",
            students=[{
                "student_id": self.student.id,
                "custom_id": self.student.custom_id,
                "first_name": self.student.first_name,
                "last_name": self.student.last_name,
                "email": self.student.email,
                "joined_at": timezone.now().isoformat(),
                "is_active": True,
            }],
        )
        self.principal.school_record = self.section.school
        self.principal.save(update_fields=["school_record"])
        User.objects.filter(id__in=[self.teacher.id, self.student.id]).update(school_record=self.section.school)
        self.assessment = Assessment.objects.create(
            title="Preview Assessment",
            code="ASM-PRV1",
            assessment_type="word",
            status="published",
            teacher=self.teacher,
            section=self.section,
            is_active=True,
            attempt_no=1,
        )
        self.assessment.record_attempt(
            self.student,
            status="completed",
            total_score=87,
            accuracy=90,
            pronunciation_score=84,
            completed_at=timezone.now(),
        )
        session = self.client.session
        session["user_id"] = self.principal.id
        session["user_role"] = self.principal.role
        session["first_name"] = self.principal.first_name
        session["last_name"] = self.principal.last_name
        session["email"] = self.principal.email
        session["custom_id"] = self.principal.custom_id
        session.save()

    def test_principal_reports_preview_shows_live_assessment_data(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "assessment"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Assessment Report")
        self.assertContains(response, "Preview Assessment")
        self.assertContains(response, "87.0%")
        self.assertContains(response, "100%")

    def test_principal_reports_excel_export_still_returns_csv_response(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "assessment", "export": "excel"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response["Content-Type"], "text/csv")
        self.assertIn("attachment; filename=", response["Content-Disposition"])

    def test_principal_reports_page_uses_a_single_report_workflow(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "assessment"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Choose a report")
        self.assertNotContains(response, "Recently Generated Reports")

    def test_principal_reports_disables_grade_filter_for_non_grade_reports(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "school"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="gradeLevel"')
        self.assertContains(response, 'disabled')

    def test_principal_reports_preview_uses_distinct_headers_for_each_report_type(self):
        school_response = self.client.get(reverse("principal_reports"), {"report_type": "school"})
        grade_response = self.client.get(reverse("principal_reports"), {"report_type": "grade"})
        assessment_response = self.client.get(reverse("principal_reports"), {"report_type": "assessment"})

        self.assertEqual(school_response.status_code, 200)
        self.assertEqual(grade_response.status_code, 200)
        self.assertEqual(assessment_response.status_code, 200)
        self.assertContains(school_response, "School Name")
        self.assertContains(grade_response, "Grade")
        self.assertContains(assessment_response, "Assessment")

    def test_principal_reports_export_buttons_submit_current_report_selection(self):
        response = self.client.get(reverse("principal_reports"), {"report_type": "assessment"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'name="export"')
        self.assertContains(response, 'value="pdf"')
        self.assertContains(response, 'value="excel"')


class LiveAssessmentStartTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}",
            role="teacher",
            first_name="Tina",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=5,
            birth_day=10,
            birth_year=1988,
            email="live-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.student = User.objects.create(
            custom_id=f"STD-{uuid.uuid4().hex[:8].upper()}",
            role="student",
            first_name="Lia",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=6,
            birth_day=2,
            birth_year=2012,
            email="live-student@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
        )
        self.section = test_section_create(
            class_code=f"LIV-{uuid.uuid4().hex[:6].upper()}",
            class_name="Live Assessment Class",
            header="Reading",
            description="Live assessment test class",
            teacher=self.teacher,
            is_active=True,
            subject="Reading",
            students=[{
                "student_id": self.student.id,
                "custom_id": self.student.custom_id,
                "first_name": self.student.first_name,
                "last_name": self.student.last_name,
                "email": self.student.email,
                "joined_at": timezone.now().isoformat(),
                "is_active": True,
            }],
        )
        self.material = Material.objects.create(
            title="Live Assessment Material",
            code="MAT-LIVE-1",
            item_type="word",
            type="assessment",
            status="published",
            teacher=self.teacher,
            section=self.section,
            is_active=True,
        )
        self.course = Course.objects.create(
            code=f"CRS-{uuid.uuid4().hex[:6].upper()}",
            title="Live Course",
            description="Course for live assessment tests",
            teacher=self.teacher,
            is_active=True,
        )
        self.course.sections.add(self.section)
        self.course.materials.add(self.material)

        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session["first_name"] = self.teacher.first_name
        session["last_name"] = self.teacher.last_name
        session["email"] = self.teacher.email
        session["custom_id"] = self.teacher.custom_id
        session.save()

    def test_teacher_can_start_live_assessment_and_notify_students(self):
        response = self.client.post(
            reverse("start_live_assessment"),
            json.dumps({
                "course_id": self.course.id,
                "material_id": self.material.id,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertIn("session", body)
        self.assertTrue(body["session"]["url"])

        session = LiveAssessmentSession.objects.filter(id=body["session"]["id"]).first()
        self.assertIsNotNone(session)
        self.assertEqual(session.status, 'waiting')
        self.assertEqual(session.student_count, 1)

        notif = Notification.objects.filter(recipient=self.student).order_by("-created_at").first()
        self.assertIsNotNone(notif)
        self.assertIn("live", notif.title.lower())
        self.assertIn("/dashboard/live-assessment/", notif.action_url)
        self.assertIn("live_session_id=", notif.action_url)

    def test_teacher_start_live_assessment_closes_existing_active_session(self):
        existing = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='waiting',
            countdown_seconds=10,
        )

        response = self.client.post(
            reverse("start_live_assessment"),
            json.dumps({
                "course_id": self.course.id,
                "material_id": self.material.id,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])

        existing.refresh_from_db()
        self.assertEqual(existing.status, 'ended')
        self.assertIsNotNone(existing.ends_at)
        self.assertEqual(existing.student_ids, [self.student.id])
        self.assertTrue(any(
            'Existing live assessment session closed automatically before starting a new session.' in entry.get('message', '')
            for entry in existing.activity_log or []
        ))

        new_session = LiveAssessmentSession.objects.filter(id=body["session"]["id"]).first()
        self.assertIsNotNone(new_session)
        self.assertNotEqual(existing.id, new_session.id)
        self.assertEqual(new_session.status, 'waiting')

    def test_live_assessment_end_action_sets_ends_at_and_student_states(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={
                str(self.student.id): {
                    'status': 'reading',
                    'progress': 0.5,
                    'connection_status': 'connected',
                }
            },
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["session"]["status"], 'ended')

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        self.assertIsNotNone(session.ends_at)
        self.assertEqual(session.student_states[str(self.student.id)]["status"], 'completed')
        self.assertEqual(session.student_states[str(self.student.id)]["connection_status"], 'disconnected')

    def test_teacher_end_waiting_live_assessment_session(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='waiting',
            countdown_seconds=10,
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["session"]["status"], 'ended')

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        self.assertIsNotNone(session.ends_at)

    def test_teacher_end_session_action_is_idempotent_on_repeat_click(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={
                str(self.student.id): {
                    'status': 'reading',
                    'progress': 0.5,
                    'items_completed': 3,
                    'items_total': 6,
                    'elapsed_seconds': 12,
                    'connection_status': 'connected',
                }
            },
        )

        first_response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )
        second_response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )

        self.assertEqual(first_response.status_code, 200)
        self.assertEqual(second_response.status_code, 200)
        self.assertTrue(first_response.json()["success"])
        self.assertTrue(second_response.json()["success"])

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        completed_attempts = Assessment.objects.filter(
            student=self.student,
            attempt_status='completed',
            is_active=True,
        ).count()
        self.assertEqual(completed_attempts, 1)

    def test_teacher_end_session_finalizes_incomplete_participants(self):
        other_student = User.objects.create(
            custom_id=f"STD-{uuid.uuid4().hex[:8].upper()}",
            role="student",
            first_name="Kai",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=4,
            birth_day=8,
            birth_year=2013,
            email="live-student-2@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
        )
        self.section.students.append({
            "student_id": other_student.id,
            "custom_id": other_student.custom_id,
            "first_name": other_student.first_name,
            "last_name": other_student.last_name,
            "email": other_student.email,
            "joined_at": timezone.now().isoformat(),
            "is_active": True,
        })
        self.section.save(update_fields=['students', 'updated_at'])

        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id, other_student.id],
            student_count=2,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={
                str(self.student.id): {
                    'status': 'reading',
                    'progress': 0.5,
                    'items_completed': 3,
                    'items_total': 6,
                    'elapsed_seconds': 12,
                    'connection_status': 'connected',
                },
                str(other_student.id): {
                    'status': 'reading',
                    'progress': 0.25,
                    'items_completed': 1,
                    'items_total': 4,
                    'elapsed_seconds': 7,
                    'connection_status': 'connected',
                },
            },
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["session"]["status"], 'ended')

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        self.assertEqual(session.student_states[str(self.student.id)]['status'], 'completed')
        self.assertEqual(session.student_states[str(other_student.id)]['status'], 'completed')
        self.assertEqual(session.student_states[str(other_student.id)]['progress'], 1)
        self.assertEqual(session.student_states[str(other_student.id)]['items_completed'], 4)
        self.assertEqual(session.student_states[str(other_student.id)]['items_total'], 4)

        completed_attempts = Assessment.objects.filter(
            student__in=[self.student, other_student],
            attempt_status='completed',
            is_active=True,
        ).count()
        self.assertEqual(completed_attempts, 2)

    def test_teacher_end_session_reconciles_already_ended_incomplete_participants(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='ended',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            ends_at=timezone.now() - timedelta(seconds=1),
            student_states={
                str(self.student.id): {
                    'status': 'reading',
                    'progress': 0.5,
                    'items_completed': 3,
                    'items_total': 6,
                    'elapsed_seconds': 12,
                    'connection_status': 'connected',
                }
            },
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "end"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        self.assertEqual(session.student_states[str(self.student.id)]['status'], 'completed')
        completed_attempts = Assessment.objects.filter(
            student=self.student,
            attempt_status='completed',
            is_active=True,
        ).count()
        self.assertEqual(completed_attempts, 1)

    def test_stale_live_assessment_session_auto_ends_on_poll(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='waiting',
            countdown_seconds=10,
        )
        LiveAssessmentSession.objects.filter(id=session.id).update(created_at=timezone.now() - timedelta(hours=25))

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.get(reverse("live_assessment_session_state", kwargs={"session_id": session.id}))
        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])
        self.assertEqual(body["session"]["status"], 'ended')

        session.refresh_from_db()
        self.assertEqual(session.status, 'ended')
        self.assertIsNotNone(session.ends_at)

    def test_live_assessment_session_state_api_returns_200_after_session_started(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=3,
            start_at=timezone.now() - timedelta(seconds=1),
        )

        response = self.client.get(reverse("live_assessment_session_state", kwargs={"session_id": session.id}))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["session"]["status"], 'started')
        self.assertIn('reader_url', data["session"])

    def test_student_can_publish_live_assessment_state_updates(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={str(self.student.id): {'status': 'waiting', 'progress': 0, 'connection_status': 'waiting'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.post(
            reverse('live_assessment_student_state_update', kwargs={'session_id': session.id}),
            json.dumps({
                'status': 'reading',
                'items_completed': 2,
                'items_total': 6,
                'progress': 0.33,
                'elapsed_seconds': 12,
                'current_item': 'cat',
                'final_score': 88,
                'connection_status': 'connected',
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        session.refresh_from_db()
        student_state = session.student_states[str(self.student.id)]
        self.assertEqual(student_state['status'], 'reading')
        self.assertEqual(student_state['items_completed'], 2)
        self.assertEqual(student_state['items_total'], 6)
        self.assertEqual(student_state['progress'], 0.33)
        self.assertEqual(student_state['elapsed_seconds'], 12)
        self.assertEqual(student_state['final_score'], 88)
        self.assertEqual(student_state['connection_status'], 'connected')

    def test_record_assessment_completion_updates_live_session_score(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={str(self.student.id): {'status': 'reading', 'progress': 0.5, 'connection_status': 'connected'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.post(
            reverse('record_assessment_completion'),
            json.dumps({
                'material_id': f'material-{self.material.id}',
                'activity_type': 'assessment',
                'class_code': self.section.class_code,
                'live_session_id': session.id,
                'scores': {
                    'fluency_score': 90,
                    'accuracy': 88,
                    'pronunciation_score': 86,
                    'time_score': 94,
                    'duration_seconds': 15,
                    'word_count': 18,
                    'transcript': 'cat dog',
                    'speech_recognition_used': True,
                },
            }),
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()['success'])
        session.refresh_from_db()
        self.assertEqual(session.student_states[str(self.student.id)]['final_score'], 89)

    def test_teacher_can_pause_and_resume_live_assessment_session(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={str(self.student.id): {'status': 'reading', 'progress': 0}},
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "pause"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        pause_body = response.json()
        self.assertTrue(pause_body["success"])
        self.assertEqual(pause_body["session"]["status"], 'paused')
        session.refresh_from_db()
        self.assertEqual(session.status, 'paused')
        self.assertEqual(session.student_states[str(self.student.id)]["status"], 'paused')
        self.assertEqual(session.student_states[str(self.student.id)]["previous_status"], 'reading')

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({"action": "resume"}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        resume_body = response.json()
        self.assertTrue(resume_body["success"])
        self.assertEqual(resume_body["session"]["status"], 'started')
        session.refresh_from_db()
        self.assertEqual(session.status, 'started')
        self.assertEqual(session.student_states[str(self.student.id)]["status"], 'reading')
        self.assertNotIn('previous_status', session.student_states[str(self.student.id)])

    def test_save_settings_persists_selection_and_notifies_students_for_waiting_room(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[],
            student_count=0,
            status='waiting',
            countdown_seconds=10,
        )

        response = self.client.post(
            reverse("live_assessment_session_action", kwargs={"session_id": session.id}),
            json.dumps({
                "action": "save_settings",
                "selected_student_ids": [self.student.id],
                "countdown_seconds": 5,
                "timing_mode": "none",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body["success"])

        session.refresh_from_db()
        self.assertEqual(session.student_ids, [self.student.id])
        self.assertEqual(session.student_count, 1)
        self.assertEqual(session.countdown_seconds, 5)
        self.assertEqual(session.status, 'waiting')
        self.assertIn(str(self.student.id), session.student_states)
        self.assertEqual(session.student_states[str(self.student.id)]['status'], 'waiting')

        notif = Notification.objects.filter(recipient=self.student).order_by('-created_at').first()
        self.assertIsNotNone(notif)
        self.assertIn('/waiting/', notif.action_url)
        self.assertIn('live-assessment', notif.action_url)

    def test_dashboard_template_uses_live_assessment_available_heading(self):
        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.get(reverse('dashboard'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'Live Assessment Available')
        self.assertNotContains(response, 'Live Assessment in Progress')

    def test_live_session_assigns_roster_into_sequential_batches(self):
        from .views import _ensure_live_session_batches

        roster = [self.student.id] + list(range(1001, 1027))
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex, teacher=self.teacher, course=self.course, material=self.material,
            student_ids=roster, student_count=len(roster),
        )
        assignments = _ensure_live_session_batches(session)

        self.assertEqual(session.batch_size, 10)
        self.assertEqual(session.total_batches, 3)
        self.assertEqual([sid for sid in roster if assignments[str(sid)] == 1], roster[:10])
        self.assertEqual([sid for sid in roster if assignments[str(sid)] == 2], roster[10:20])
        self.assertEqual([sid for sid in roster if assignments[str(sid)] == 3], roster[20:])

    def test_live_session_next_batch_is_sequential_and_idempotently_guarded(self):
        from .views import _ensure_live_session_batches

        roster = [self.student.id]
        for index in range(10):
            student = User.objects.create(
                custom_id=f"BATCH-{uuid.uuid4().hex[:8].upper()}", role='student',
                first_name=f'Student{index}', last_name='Batch', email=f'batch-{uuid.uuid4().hex}@example.com',
                password_hash=make_password('student-password'), grade_level='Grade 2',
                middle_initial='', suffix='', sex='female', birth_month=6, birth_day=2, birth_year=2012,
            )
            roster.append(student.id)
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex, teacher=self.teacher, course=self.course, material=self.material,
            student_ids=roster, student_count=len(roster), status='started', start_at=timezone.now(),
            student_states={str(student_id): {'status': 'completed'} for student_id in roster[:10]},
        )
        _ensure_live_session_batches(session)
        session.save(update_fields=['batch_assignments', 'batch_size', 'current_batch', 'total_batches'])

        response = self.client.post(
            reverse('live_assessment_session_action', kwargs={'session_id': session.id}),
            json.dumps({'action': 'start_next_batch', 'target_batch': 2}), content_type='application/json',
        )
        self.assertEqual(response.status_code, 200)
        session.refresh_from_db()
        self.assertEqual(session.current_batch, 2)

        duplicate = self.client.post(
            reverse('live_assessment_session_action', kwargs={'session_id': session.id}),
            json.dumps({'action': 'start_next_batch', 'target_batch': 2}), content_type='application/json',
        )
        self.assertEqual(duplicate.status_code, 409)
        session.refresh_from_db()
        self.assertEqual(session.current_batch, 2)

    def test_student_active_invitation_endpoint_only_returns_late_joiner_modal(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            timing_mode='duration',
            duration_seconds=60,
            start_at=timezone.now() - timedelta(seconds=20),
            student_states={str(self.student.id): {'status': 'waiting', 'progress': 0, 'connection_status': 'waiting'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session['login_at'] = (timezone.now() + timedelta(seconds=5)).isoformat()
        student_session.save()

        response = student_client.get(reverse('live_assessment_active_invitation'))

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertIsNotNone(body['session'])
        self.assertTrue(body['session']['show_modal'])
        self.assertEqual(body['session']['id'], session.id)
        self.assertEqual(body['session']['timing_mode'], 'duration')
        self.assertIsNotNone(body['session']['remaining_seconds'])
        self.assertTrue(body['session']['join_url'].endswith(f'/dashboard/live-assessment/{session.id}/waiting/'))

    def test_student_active_invitation_endpoint_redirects_already_logged_in_students_to_waiting_room(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            timing_mode='none',
            duration_seconds=None,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={str(self.student.id): {'status': 'waiting', 'progress': 0, 'connection_status': 'waiting'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session['login_at'] = (timezone.now() - timedelta(hours=1)).isoformat()
        student_session.save()

        response = student_client.get(reverse('live_assessment_active_invitation'))

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertIsNotNone(body['session'])
        self.assertTrue(body['session']['redirect_to_waiting_room'])
        self.assertFalse(body['session']['show_modal'])

    def test_student_active_invitation_endpoint_returns_duration_session_details(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            timing_mode='duration',
            duration_seconds=60,
            start_at=timezone.now() - timedelta(seconds=20),
            student_states={str(self.student.id): {'status': 'reading', 'progress': 0.5, 'connection_status': 'connected'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.get(reverse('live_assessment_active_invitation'))

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['session']['id'], session.id)
        self.assertEqual(body['session']['timing_mode'], 'duration')
        self.assertIsNotNone(body['session']['remaining_seconds'])
        self.assertGreaterEqual(body['session']['remaining_seconds'], 0)
        self.assertLessEqual(body['session']['remaining_seconds'], 60)
        self.assertTrue(body['session']['join_url'].endswith(f'/dashboard/live-assessment/{session.id}/waiting/'))

    def test_student_active_invitation_endpoint_returns_no_limit_session_details(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            timing_mode='none',
            duration_seconds=None,
            start_at=timezone.now() - timedelta(seconds=10),
            student_states={str(self.student.id): {'status': 'reading', 'progress': 0.5, 'connection_status': 'connected'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session.save()

        response = student_client.get(reverse('live_assessment_active_invitation'))

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertEqual(body['session']['id'], session.id)
        self.assertEqual(body['session']['timing_mode'], 'none')
        self.assertIsNone(body['session']['remaining_seconds'])
        self.assertTrue(body['session']['join_url'].endswith(f'/dashboard/live-assessment/{session.id}/waiting/'))

    def test_student_active_invitation_uses_login_timestamp_for_late_join_modal(self):
        session = LiveAssessmentSession.objects.create(
            id=uuid.uuid4().hex,
            teacher=self.teacher,
            course=self.course,
            material=self.material,
            student_ids=[self.student.id],
            student_count=1,
            status='started',
            countdown_seconds=0,
            timing_mode='duration',
            duration_seconds=120,
            start_at=timezone.now() - timedelta(minutes=2),
            student_states={str(self.student.id): {'status': 'waiting', 'progress': 0, 'connection_status': 'waiting'}},
        )

        student_client = Client()
        student_session = student_client.session
        student_session['user_id'] = self.student.id
        student_session['user_role'] = 'student'
        student_session['first_name'] = self.student.first_name
        student_session['last_name'] = self.student.last_name
        student_session['email'] = self.student.email
        student_session['custom_id'] = self.student.custom_id
        student_session['login_at'] = (timezone.now() - timedelta(seconds=10)).isoformat()
        student_session.save()

        response = student_client.get(reverse('live_assessment_active_invitation'))

        self.assertEqual(response.status_code, 200)
        body = response.json()
        self.assertTrue(body['success'])
        self.assertIsNotNone(body['session'])
        self.assertTrue(body['session']['show_modal'])
        self.assertFalse(body['session']['redirect_to_waiting_room'])
        self.assertEqual(body['session']['id'], session.id)


class PrincipalAccountBootstrapTests(TestCase):
    def test_login_recreates_missing_principal_account_once(self):
        self.assertFalse(User.objects.filter(custom_id=PRINCIPAL_DEFAULT_CUSTOM_ID).exists())

        response = self.client.post(
            reverse("login_user"),
            {
                "custom_id": PRINCIPAL_DEFAULT_CUSTOM_ID,
                "password": PRINCIPAL_DEFAULT_PASSWORD,
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        principal = User.objects.get(custom_id=PRINCIPAL_DEFAULT_CUSTOM_ID)
        self.assertEqual(principal.role, "principal")
        self.assertTrue(check_password(PRINCIPAL_DEFAULT_PASSWORD, principal.password_hash))
        self.assertEqual(User.objects.filter(custom_id=PRINCIPAL_DEFAULT_CUSTOM_ID).count(), 1)

        second_response = self.client.post(
            reverse("login_user"),
            {
                "custom_id": PRINCIPAL_DEFAULT_CUSTOM_ID,
                "password": PRINCIPAL_DEFAULT_PASSWORD,
            },
        )

        self.assertEqual(second_response.status_code, 200)
        self.assertTrue(second_response.json()["success"])
        self.assertEqual(User.objects.filter(custom_id=PRINCIPAL_DEFAULT_CUSTOM_ID).count(), 1)


class ProfileUpdateTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(
            custom_id="TCH-0001",
            role="teacher",
            first_name="Old",
            last_name="Name",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="old@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = self.user.id
        session["user_role"] = self.user.role
        session["first_name"] = self.user.first_name
        session["last_name"] = self.user.last_name
        session["email"] = self.user.email
        session["custom_id"] = self.user.custom_id
        session.save()

    def test_profile_page_includes_hidden_save_flag(self):
        response = self.client.get(reverse("profile"))

        self.assertContains(response, 'name="save_account_details" value="true"', html=False)

    def test_profile_post_updates_user_record(self):
        response = self.client.post(
            reverse("profile"),
            {
                "save_account_details": "true",
                "first_name": "New",
                "last_name": "Name",
                "middle_initial": "Q",
                "suffix": "Jr.",
                "email": "new@example.com",
                "bio": "Updated bio",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()["success"], True)

        self.user.refresh_from_db()
        self.assertEqual(self.user.first_name, "New")
        self.assertEqual(self.user.last_name, "Name")
        self.assertEqual(self.user.middle_initial, "Q")
        self.assertEqual(self.user.suffix, "Jr.")
        self.assertEqual(self.user.email, "new@example.com")
        self.assertIn({"profile_info": {"bio": "Updated bio"}}, self.user.tags)


class MaterialCreationTests(TestCase):
    def test_add_reading_material_saves_selected_filipino_language(self):
        user = User.objects.create(
            custom_id="TCH-0002",
            role="teacher",
            first_name="Language",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="language@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Filipino reading",
                "content": "Araw\nBuwan",
                "reading_type": "word",
                "status": "published",
                "usage_type": "practice",
                "class_code": "",
                "language": "Filipino",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        material = Material.objects.latest("id")
        self.assertEqual(material.language, "Filipino")
        self.assertEqual(material.content_json.get("language"), "Filipino")
        self.assertEqual(material.type, "assessment")
        self.assertEqual(material.source_type, "shared")

    def test_material_response_payload_preserves_saved_language(self):
        material = Material.objects.create(
            teacher=self.teacher,
            title="Language reading",
            item_type="word",
            prompt_text="Araw",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Filipino"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )

        payload = _material_response_payload(material)

        self.assertEqual(payload["language"], "Filipino")

    def test_material_saved_language_display_uses_saved_value_or_not_set(self):
        material = Material.objects.create(
            teacher=self.teacher,
            title="Language reading",
            item_type="word",
            prompt_text="Araw",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Filipino"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )

        self.assertEqual(material.get_saved_language_display(), "Filipino")

        material.content_json = {"items": ["Araw"], "language": "English"}
        material.save(update_fields=["content_json", "updated_at"])
        self.assertEqual(material.get_saved_language_display(), "English")

        legacy_material = Material.objects.create(
            teacher=self.teacher,
            title="Legacy reading",
            item_type="word",
            prompt_text="Araw",
            content_text="Araw",
            content_json={},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )

        self.assertEqual(legacy_material.get_saved_language_display(), "Not Set")

    def test_add_reading_material_response_includes_shared_source_type(self):
        user = User.objects.create(
            custom_id="TCH-0003",
            role="teacher",
            first_name="Source",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="source@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared reading",
                "content": "Araw\nBuwan",
                "reading_type": "word",
                "status": "published",
                "usage_type": "assessment",
                "source_type": "shared",
                "class_code": "",
                "language": "Tagalog",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        material = Material.objects.latest("id")
        self.assertEqual(material.type, "assessment")
        self.assertEqual(material.source_type, "shared")

    def test_add_reading_material_reuses_existing_shared_material(self):
        user = User.objects.create(
            custom_id="TCH-0006",
            role="teacher",
            first_name="Reuse",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="reuse@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        existing = Material.objects.create(
            teacher=user,
            title="Shared reading",
            item_type="word",
            prompt_text="Araw",
            content_text="Araw\nBuwan",
            content_json={"items": ["Araw", "Buwan"], "language": "Tagalog"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared reading",
                "content": "Araw\nBuwan",
                "reading_type": "word",
                "status": "published",
                "usage_type": "assessment",
                "source_type": "shared",
                "source_material_id": existing.id,
                "class_code": "",
                "language": "Tagalog",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertTrue(payload["reused"])
        self.assertEqual(payload["created_count"], 0)
        self.assertEqual(payload["material_ids"], [existing.id])
        self.assertEqual(Material.objects.filter(source_type="shared", title="Shared reading").count(), 1)

    @patch("pabasa_app.views._compute_teacher_overview")
    def test_add_reading_material_reuse_skips_overview(self, mock_overview):
        user = User.objects.create(
            custom_id="TCH-0008",
            role="teacher",
            first_name="ReuseOverview",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="reuse-overview@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        existing = Material.objects.create(
            teacher=user,
            title="Shared reading",
            item_type="word",
            prompt_text="Araw",
            content_text="Araw\nBuwan",
            content_json={"items": ["Araw", "Buwan"], "language": "Tagalog"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared reading",
                "content": "Araw\nBuwan",
                "reading_type": "word",
                "status": "published",
                "usage_type": "assessment",
                "source_type": "shared",
                "source_material_id": existing.id,
                "class_code": "",
                "language": "Tagalog",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        mock_overview.assert_not_called()

    def test_add_reading_material_saves_vowel_and_vc_items_as_vowel(self):
        user = User.objects.create(
            custom_id="TCH-0007",
            role="teacher",
            first_name="Vowel",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="vowel@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Vowel reading",
                "content": "a\nbe\nmi",
                "reading_type": "word",
                "status": "published",
                "usage_type": "assessment",
                "class_code": "",
                "language": "English",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        material = Material.objects.latest("id")
        self.assertEqual(material.item_type, "vowel")
        self.assertEqual(material.content_json.get("items"), ["a", "be", "mi"])

    def test_add_reading_material_saves_multiple_paragraph_items(self):
        user = User.objects.create(
            custom_id="TCH-0003",
            role="teacher",
            first_name="Paragraph",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="paragraph@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Paragraph reading",
                "content": "First paragraph text.\n\nSecond paragraph text.",
                "reading_type": "paragraph",
                "status": "published",
                "usage_type": "practice",
                "class_code": "",
                "language": "English",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        material = Material.objects.latest("id")
        self.assertEqual(material.item_type, "paragraph")
        self.assertEqual(material.content_json.get("items"), ["First paragraph text.", "Second paragraph text."])

    def test_add_reading_material_saves_separate_sentence_items_from_multiline_content(self):
        user = User.objects.create(
            custom_id="TCH-0004",
            role="teacher",
            first_name="Sentence",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="sentence@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Sentence reading",
                "content": "First sentence.\nSecond sentence.\nThird sentence.",
                "reading_type": "sentence",
                "status": "published",
                "usage_type": "assessment",
                "class_code": "",
                "language": "English",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        material = Material.objects.latest("id")
        self.assertEqual(material.item_type, "sentence")
        self.assertEqual(material.content_json.get("items"), ["First sentence.", "Second sentence.", "Third sentence."])

    @patch("pabasa_app.views._compute_teacher_overview")
    def test_add_reading_material_skips_overview_by_default(self, mock_overview):
        user = User.objects.create(
            custom_id="TCH-0004",
            role="teacher",
            first_name="Fast",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="fast@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Fast create",
                "content": "alpha",
                "reading_type": "word",
                "status": "published",
                "usage_type": "practice",
                "class_code": "",
                "language": "English",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        mock_overview.assert_not_called()

    def test_add_reading_material_saves_shared_source_type(self):
        user = User.objects.create(
            custom_id="TCH-0005",
            role="teacher",
            first_name="Shared",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="shared@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = user.id
        session["user_role"] = user.role
        session["first_name"] = user.first_name
        session["last_name"] = user.last_name
        session["email"] = user.email
        session["custom_id"] = user.custom_id
        session.save()

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Shared reading",
                "content": "Araw\nBuwan",
                "reading_type": "word",
                "status": "published",
                "usage_type": "practice",
                "class_code": "",
                "language": "Tagalog",
                "source_type": "shared",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["material"]["source_type"], "shared")

        material = Material.objects.latest("id")
        self.assertEqual(material.source_type, "shared")

    def test_teacher_courses_api_includes_material_source_type(self):
        teacher = User.objects.create(
            custom_id="TCH-0004",
            role="teacher",
            first_name="Course",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="course@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="CRS-1001",
            class_name="Course 1",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        student = User.objects.create(
            custom_id="STD-1001",
            role="student",
            first_name="Metric",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2015,
            email="metric.student@example.com",
            password_hash="hashed-password",
        )
        section.students = [{
            "student_id": student.id,
            "custom_id": student.custom_id,
            "first_name": student.first_name,
            "last_name": student.last_name,
            "email": student.email,
            "is_active": True,
        }]
        section.save(update_fields=["students"])
        course = Course.objects.create(
            code="C-1001",
            title="Shared Course",
            description="",
            teacher=teacher,
        )
        course.sections.add(section)

        material = Material.objects.create(
            title="Imported reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "English"},
            type="practice",
            source_type="shared",
            status="published",
            difficulty_level="",
            is_active=True,
            section=section,
        )
        course.materials.add(material)
        assessment = Assessment.objects.create(
            title="Course assessment",
            code="ASM-1001",
            assessment_type="word",
            status="published",
            teacher=teacher,
            section=section,
            is_active=True,
            attempt_no=1,
        )
        course.assessments.add(assessment)
        Assessment.objects.create(
            title="Material progress",
            code="ASM-1002",
            assessment_type="word",
            status="published",
            teacher=teacher,
            section=section,
            material=material,
            student=student,
            attempt_status="completed",
            is_active=True,
            attempt_no=1,
            items_completed=2,
            total_score=100,
        )

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.get(reverse("get_teacher_courses_api"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        course_payload = next(item for item in payload["courses"] if item["id"] == course.id)
        material_payload = next(item for item in course_payload["materials"] if item["id"] == material.id)
        self.assertEqual(material_payload["source_type"], "shared")
        self.assertTrue(material_payload["is_shared_material"])
        self.assertEqual(course_payload["metrics"], {
            "sections": 1,
            "assessments": 1,
            "materials": 1,
            "students": 1,
            "average_progress": 100.0,
        })

    def test_template_material_uses_shared_teacher_and_student_pipeline(self):
        teacher = User.objects.create(
            custom_id="TCH-0013",
            role="teacher",
            first_name="Template",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="template.teacher@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        student = User.objects.create(
            custom_id="STU-0013",
            role="student",
            first_name="Template",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2015,
            email="template.student@example.com",
            password_hash="hashed-password",
        )
        section = test_section_create(
            class_code="TPL-1001",
            class_name="Template Class",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        section.add_student(student)
        section_count = Section.objects.count()
        course_count = Course.objects.count()

        teacher_session = self.client.session
        teacher_session["user_id"] = teacher.id
        teacher_session["user_role"] = teacher.role
        teacher_session["first_name"] = teacher.first_name
        teacher_session["last_name"] = teacher.last_name
        teacher_session["email"] = teacher.email
        teacher_session["custom_id"] = teacher.custom_id
        teacher_session.save()

        create_response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Template Reading",
                "content": json.dumps({
                    "template_title": "Letter & Sound Matching",
                    "template_lesson": "Lesson 1",
                    "template_type": "template",
                    "template_source": "template",
                    "items": ["A", "B"],
                }),
                "reading_type": "word",
                "status": "published",
                "usage_type": "assessment",
                "class_code": section.class_code,
                "assigned_week": 3,
                "assigned_weeks": [3],
                "language": "Tagalog",
                "randomize_order": False,
                "assessment_kind": "regular",
                "source_type": "template",
                "template_title": "Letter & Sound Matching",
                "template_lesson": "Lesson 1",
            }),
            content_type="application/json",
        )
        self.assertEqual(create_response.status_code, 200)
        create_payload = create_response.json()
        self.assertTrue(create_payload["success"])
        material_id = create_payload["material"]["raw_id"]
        material = Material.objects.get(id=material_id)
        self.assertEqual(material.source_type, "template")
        self.assertEqual(material.teacher_id, teacher.id)
        self.assertEqual(material.section_id, section.id)
        self.assertEqual(material.status, "published")
        self.assertEqual(material.assigned_weeks, [3])
        self.assertEqual(material.assigned_week, 3)
        self.assertFalse(material.student_access)

        attach_response = self.client.post(
            reverse("add_material_to_course"),
            json.dumps({"course_id": f"section-{section.id}", "material_id": material.id}),
            content_type="application/json",
        )
        self.assertEqual(attach_response.status_code, 200)
        self.assertTrue(attach_response.json()["success"])
        self.assertEqual(Section.objects.count(), section_count)
        self.assertEqual(Course.objects.count(), course_count)

        course_response = self.client.get(reverse("get_teacher_courses_api"))
        self.assertEqual(course_response.status_code, 200)
        course_payload = next(
            item for item in course_response.json()["courses"]
            if item["id"] == f"section-{section.id}"
        )
        course_material = next(item for item in course_payload["materials"] if item["id"] == material.id)
        self.assertEqual(course_material["source_type"], "template")
        self.assertEqual(course_material["assigned_weeks"], [3])
        self.assertEqual(course_material["status"], "published")

        toggle_response = self.client.post(
            reverse("toggle_material_student_access"),
            json.dumps({"material_id": material.id, "student_access": True}),
            content_type="application/json",
        )
        self.assertEqual(toggle_response.status_code, 200)
        self.assertTrue(toggle_response.json()["success"])

        material.refresh_from_db()
        self.assertTrue(material.student_access)

        student_session = self.client.session
        student_session["user_id"] = student.id
        student_session["user_role"] = student.role
        student_session["current_week"] = 3
        student_session.save()

        assessment_response = self.client.get(reverse("assessment"))
        self.assertEqual(assessment_response.status_code, 200)
        student_materials = assessment_response.context["student_assessment_materials"]
        self.assertTrue(any(item["id"] == material.id for item in student_materials))
        matched = next(item for item in student_materials if item["id"] == material.id)
        self.assertEqual(matched["title"], "Template Reading")

    def test_class_materials_api_includes_template_material(self):
        teacher = User.objects.create(
            custom_id="TCH-0014",
            role="teacher",
            first_name="API",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="api.teacher@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="FCNG-648",
            class_name="API Class",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        course = Course.objects.create(
            code="API-C1",
            title="API Course",
            description="",
            teacher=teacher,
        )
        course.sections.add(section)
        material = Material.objects.create(
            teacher=teacher,
            section=section,
            title="heheheh",
            item_type="word",
            content_text="heheheh",
            content_json={"items": ["heheheh"], "language": "English"},
            type="assessment",
            source_type="template",
            status="published",
            assigned_week=3,
            assigned_weeks=[3],
            student_access=False,
            is_active=True,
        )
        course.materials.add(material)

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.get(reverse("get_class_materials"), {"class_code": section.class_code})
        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertIn("all_materials", payload)
        self.assertTrue(any(item["title"] == "heheheh" and item["source_type"] == "template" for item in payload["all_materials"]))

    def test_teacher_courses_api_preserves_saved_material_language(self):
        teacher = User.objects.create(
            custom_id="TCH-0012",
            role="teacher",
            first_name="Language",
            last_name="Owner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="language-owner@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="LANG-1001",
            class_name="Language Class",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        course = Course.objects.create(
            code="C-LANG-1",
            title="Language Course",
            description="",
            teacher=teacher,
        )
        course.sections.add(section)
        material = Material.objects.create(
            teacher=teacher,
            title="Language reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="personal",
            status="published",
            is_active=True,
            section=section,
        )
        course.materials.add(material)

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.get(reverse("get_teacher_courses_api"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        course_payload = next(item for item in payload["courses"] if item["id"] == course.id)
        material_payload = next(item for item in course_payload["materials"] if item["id"] == material.id)
        self.assertEqual(material_payload["language"], "Tagalog")
        self.assertEqual(material_payload["content_json"]["language"], "Tagalog")

    def test_teacher_courses_api_keeps_personal_materials_unmarked_as_shared(self):
        current_teacher = User.objects.create(
            custom_id="TCH-0012",
            role="teacher",
            first_name="Current",
            last_name="Owner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="current-owner@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        other_teacher = User.objects.create(
            custom_id="TCH-0013",
            role="teacher",
            first_name="Other",
            last_name="Owner",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="other-owner@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        course = Course.objects.create(
            code="C-PERSONAL-1",
            title="Personal Course",
            description="",
            teacher=current_teacher,
        )
        material = Material.objects.create(
            teacher=other_teacher,
            title="Private reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="personal",
            status="published",
            is_active=True,
        )
        course.materials.add(material)

        session = self.client.session
        session["user_id"] = current_teacher.id
        session["user_role"] = current_teacher.role
        session["first_name"] = current_teacher.first_name
        session["last_name"] = current_teacher.last_name
        session["email"] = current_teacher.email
        session["custom_id"] = current_teacher.custom_id
        session.save()

        response = self.client.get(reverse("get_teacher_courses_api"))

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        course_payload = next(item for item in payload["courses"] if item["id"] == course.id)
        material_payload = next(item for item in course_payload["materials"] if item["id"] == material.id)
        self.assertEqual(material_payload["source_type"], "personal")
        self.assertEqual(material_payload["material_source"], "personal")
        self.assertFalse(material_payload["is_shared_material"])

    def test_delete_course_removes_course_and_related_records(self):
        teacher = User.objects.create(
            custom_id="TCH-0014",
            role="teacher",
            first_name="Delete",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="delete-course@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="DEL-1001",
            class_name="Delete Course Section",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        course = Course.objects.create(
            code="C-DELETE-1",
            title="Delete Course",
            description="",
            teacher=teacher,
        )
        course.sections.add(section)
        material = Material.objects.create(
            teacher=teacher,
            title="Course material",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="personal",
            status="published",
            is_active=True,
        )
        assessment = Assessment.objects.create(
            title="Course assessment",
            code="ASM-DELETE-1",
            assessment_type="word",
            status="published",
            teacher=teacher,
            section=section,
            is_active=True,
            attempt_no=1,
        )
        course.materials.add(material)
        course.assessments.add(assessment)

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.post(
            reverse("delete_course"),
            json.dumps({"course_id": course.id, "confirmation": "DELETE"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertFalse(Course.objects.filter(id=course.id).exists())
        self.assertFalse(material.refresh_from_db().is_active)
        self.assertFalse(assessment.refresh_from_db().is_active)

    def test_create_course_allows_empty_sections(self):
        teacher = User.objects.create(
            custom_id="TCH-0013",
            role="teacher",
            first_name="Empty",
            last_name="Course",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="empty-course@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.post(
            reverse("create_course"),
            json.dumps({"title": "Empty Course", "description": "No classes yet", "sections": []}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        course = Course.objects.get(id=payload["course"]["id"])
        self.assertEqual(course.title, "Empty Course")
        self.assertEqual(course.description, "No classes yet")
        self.assertEqual(course.sections.count(), 0)

    def test_delete_course_accepts_prefixed_course_id(self):
        teacher = User.objects.create(
            custom_id="TCH-0015",
            role="teacher",
            first_name="Delete",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="delete-course-prefixed@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        course = Course.objects.create(
            code="C-DELETE-2",
            title="Delete Course Prefixed",
            description="",
            teacher=teacher,
        )

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.post(
            reverse("delete_course"),
            json.dumps({"course_id": f"course-{course.id}", "confirmation": "DELETE"}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertFalse(Course.objects.filter(id=course.id).exists())

    def test_delete_course_rejects_missing_confirmation(self):
        teacher = User.objects.create(
            custom_id="TCH-0016",
            role="teacher",
            first_name="Confirm",
            last_name="Teacher",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="delete-course-confirmation@example.com",
            password_hash="hashed-password",
        )
        course = Course.objects.create(code="C-CONFIRM", title="Protected Course", teacher=teacher)
        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session.save()

        response = self.client.post(
            reverse("delete_course"),
            json.dumps({"course_id": course.id}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 400)
        self.assertFalse(response.json()["success"])
        self.assertTrue(Course.objects.filter(id=course.id).exists())

    def test_shared_courses_api_includes_own_shared_materials_without_personal_rows(self):
        current_teacher = User.objects.create(
            custom_id="TCH-0010",
            role="teacher",
            first_name="Current",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="current-shared@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        other_teacher = User.objects.create(
            custom_id="TCH-0011",
            role="teacher",
            first_name="Other",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="other-shared@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        current_course = Course.objects.create(
            code="C-CURRENT-1",
            title="Current Course",
            description="",
            teacher=current_teacher,
        )
        other_course = Course.objects.create(
            code="C-OTHER-1",
            title="Other Course",
            description="",
            teacher=other_teacher,
        )
        personal_material = Material.objects.create(
            teacher=current_teacher,
            title="Legacy private reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="personal",
            status="published",
            is_active=True,
        )
        current_shared_material = Material.objects.create(
            teacher=current_teacher,
            title="Current shared reading",
            item_type="word",
            content_text="Buwan",
            content_json={"items": ["Buwan"], "language": "Tagalog"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )
        shared_material = Material.objects.create(
            teacher=other_teacher,
            title="Original shared reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="shared",
            status="published",
            is_active=True,
        )
        current_course.materials.add(personal_material, current_shared_material)
        other_course.materials.add(shared_material)

        session = self.client.session
        session["user_id"] = current_teacher.id
        session["user_role"] = current_teacher.role
        session["first_name"] = current_teacher.first_name
        session["last_name"] = current_teacher.last_name
        session["email"] = current_teacher.email
        session["custom_id"] = current_teacher.custom_id
        session.save()

        response = self.client.get(reverse("get_teacher_courses_api"), {"shared": "true"})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        course_ids = {course["id"] for course in payload["courses"]}
        self.assertIn(current_course.id, course_ids)
        self.assertIn(other_course.id, course_ids)
        material_ids = {
            material["id"]
            for course in payload["courses"]
            for material in course["materials"]
        }
        self.assertNotIn(personal_material.id, material_ids)
        self.assertIn(current_shared_material.id, material_ids)
        self.assertIn(shared_material.id, material_ids)

    def test_courses_template_uses_selected_source_type_for_new_materials(self):
        template_path = Path(settings.BASE_DIR) / "pabasa_app" / "templates" / "pabasa_app" / "courses.html"
        template_content = template_path.read_text(encoding="utf-8")

        self.assertIn("const sourceType = ", template_content)
        self.assertIn("source_type: sourceType", template_content)
        self.assertNotIn("const sourceType = 'shared';", template_content)

    def test_add_material_to_course_response_includes_material_source_type(self):
        teacher = User.objects.create(
            custom_id="TCH-0006",
            role="teacher",
            first_name="Attach",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="attach@example.com",
            password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="ATT-1001",
            class_name="Attach 1",
            header="Reading Class",
            description="",
            teacher=teacher,
            subject="Reading",
        )
        course = Course.objects.create(
            code="ATT-C1",
            title="Attach Course",
            description="",
            teacher=teacher,
        )
        course.sections.add(section)
        material = Material.objects.create(
            title="Persistent shared reading",
            item_type="word",
            content_text="Araw",
            content_json={"items": ["Araw"], "language": "Tagalog"},
            type="assessment",
            source_type="shared",
            status="published",
            difficulty_level="",
            is_active=True,
            section=section,
            teacher=teacher,
        )

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.post(
            reverse("add_material_to_course"),
            json.dumps({"course_id": course.id, "material_id": material.id}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["material"]["source_type"], "shared")
        self.assertEqual(payload["material"]["material_source"], "shared")
        self.assertTrue(payload["material"]["is_shared_material"])

    def test_add_template_material_accepts_canonical_section_id(self):
        teacher = User.objects.create(
            custom_id="TCH-SECTION-ID", role="teacher", first_name="Class", last_name="Owner",
            middle_initial="", suffix="", sex="female", birth_month=1, birth_day=1,
            birth_year=1990, email="section-owner@example.com", password_hash="hashed-password",
            teacher_role="Teacher",
        )
        section = test_section_create(
            class_code="CLS-SECTION-ID", class_name="Grade 2 - Mabini", header="Reading Class",
            description="", teacher=teacher, subject="Filipino",
        )
        session = self.client.session
        session.update({"user_id": teacher.id, "user_role": "teacher", "email": teacher.email})
        session.save()
        template_content = {
            "template_title": "Picture-Word Matching",
            "template_lesson": "Word Recognition",
            "template_type": "Picture-Word Matching",
            "template_source": "template",
            "language": "Filipino",
            "items": [{"word": "Aso"}],
        }

        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Picture words", "content": json.dumps(template_content),
                "reading_type": "word", "status": "published", "source_type": "template",
                "template_title": "Picture-Word Matching", "template_lesson": "Word Recognition",
                "class_id": f"section-{section.id}", "class_code": section.class_code,
                "assigned_weeks": ["Week 1"], "language": "Filipino",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()["material"]
        material = Material.objects.get(id=payload["raw_id"])
        self.assertEqual(material.section_id, section.id)
        self.assertTrue(material.assigned_sections.filter(id=section.id).exists())
        self.assertEqual(payload["class_id"], section.id)
        self.assertEqual(payload["class_code"], section.class_code)
        self.assertEqual(payload["content_json"]["template_title"], "Picture-Word Matching")


class CanonicalSectionMaterialWorkflowTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id="TCH-CANONICAL", role="teacher", first_name="Canonical", last_name="Teacher",
            sex="female", birth_month=1, birth_day=1, birth_year=1990,
            email="canonical.teacher@example.com", password_hash=make_password("password"),
        )
        self.other_teacher = User.objects.create(
            custom_id="TCH-SHARED", role="teacher", first_name="Shared", last_name="Teacher",
            sex="female", birth_month=1, birth_day=1, birth_year=1990,
            email="shared.teacher@example.com", password_hash=make_password("password"),
        )
        self.section = test_section_create(
            class_code="G2-BONIFACIO", class_name="Grade 2 - BONIFACIO",
            grade_level="Grade 2", section="BONIFACIO", teacher=self.teacher,
            subject="Filipino", is_active=True,
        )
        self.other_section = test_section_create(
            class_code="G2-MABINI", class_name="Grade 2 - MABINI",
            grade_level="Grade 2", section="MABINI", teacher=self.other_teacher,
            subject="Filipino", is_active=True,
        )
        self.student = User.objects.create(
            custom_id="STU-BONIFACIO", role="student", first_name="Bonifacio", last_name="Student",
            sex="male", birth_month=1, birth_day=1, birth_year=2016,
            email="bonifacio.student@example.com", password_hash=make_password("password"),
        )
        self.outside_student = User.objects.create(
            custom_id="STU-MABINI", role="student", first_name="Mabini", last_name="Student",
            sex="female", birth_month=1, birth_day=1, birth_year=2016,
            email="mabini.student@example.com", password_hash=make_password("password"),
        )
        self.section.add_student(self.student)
        self.other_section.add_student(self.outside_student)
        self._login(self.teacher)

    def _login(self, user):
        session = self.client.session
        session.update({"user_id": user.id, "user_role": user.role, "custom_id": user.custom_id})
        session.save()

    def _create_material(self, source_type="personal", class_id=None):
        data = {
            "title": f"{source_type.title()} canonical reading",
            "content": "Aso\nPusa",
            "reading_type": "word",
            "status": "published",
            "source_type": source_type,
            "language": "Filipino",
        }
        if class_id is not None:
            data["class_id"] = f"section-{class_id}"
            data["class_code"] = self.section.class_code
        return self.client.post(
            reverse("add_reading_material"), json.dumps(data), content_type="application/json",
        )

    def test_template_creation_targets_exact_existing_section_without_creating_section_or_course(self):
        section_count = Section.objects.count()
        course_count = Course.objects.count()
        section_pk = self.section.pk
        response = self.client.post(
            reverse("add_reading_material"),
            json.dumps({
                "title": "Template canonical reading",
                "content": json.dumps({
                    "template_title": "Letter & Sound Matching", "template_lesson": "Lesson 1",
                    "template_source": "template", "items": ["A", "B"],
                }),
                "reading_type": "word", "status": "published", "source_type": "template",
                "template_title": "Letter & Sound Matching", "template_lesson": "Lesson 1",
                "class_id": f"section-{self.section.pk}", "class_code": self.section.class_code,
                "language": "Filipino",
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        material = Material.objects.get(pk=response.json()["material"]["raw_id"])
        self.assertEqual(material.section_id, section_pk)
        self.assertEqual(list(material.assigned_sections.values_list("pk", flat=True)), [section_pk])
        self.assertEqual(Section.objects.count(), section_count)
        self.assertEqual(Course.objects.count(), course_count)
        self.section.refresh_from_db()
        self.assertEqual(self.section.pk, section_pk)

    def test_personal_create_assign_reassign_and_edit_preserve_canonical_section(self):
        section_count = Section.objects.count()
        course_count = Course.objects.count()
        create_response = self._create_material()
        self.assertEqual(create_response.status_code, 200)
        material = Material.objects.get(pk=create_response.json()["material"]["raw_id"])
        self.assertEqual(material.teacher_id, self.teacher.id)
        self.assertIsNone(material.section_id)

        for _ in range(2):
            assign_response = self.client.post(
                reverse("add_material_to_course"),
                json.dumps({"course_id": f"section-{self.section.pk}", "material_id": material.pk}),
                content_type="application/json",
            )
            self.assertEqual(assign_response.status_code, 200)

        material.refresh_from_db()
        self.assertEqual(material.section_id, self.section.pk)
        self.assertEqual(material.assigned_sections.filter(pk=self.section.pk).count(), 1)
        assignment_ids = set(material.assigned_sections.values_list("pk", flat=True))

        edit_response = self.client.post(
            reverse("teacher_update_material"),
            json.dumps({
                "material_id": f"material-{material.pk}", "title": "Edited canonical reading",
                "content": "Aso\nPusa\nIbon", "reading_type": "word", "status": "published",
                "language": "Filipino",
            }),
            content_type="application/json",
        )
        self.assertEqual(edit_response.status_code, 200)
        material.refresh_from_db()
        self.assertEqual(material.title, "Edited canonical reading")
        self.assertEqual(material.language, "Filipino")
        self.assertEqual(material.content_json.get("language"), "Filipino")
        self.assertEqual(set(material.assigned_sections.values_list("pk", flat=True)), assignment_ids)
        self.assertEqual(Section.objects.count(), section_count)
        self.assertEqual(Course.objects.count(), course_count)

    def test_shared_material_assigns_to_existing_section_without_changing_owner(self):
        shared = Material.objects.create(
            teacher=self.other_teacher, title="Shared library reading", item_type="word",
            content_text="Araw", content_json={"items": ["Araw"]}, type="assessment",
            source_type="shared", status="published", is_active=True,
        )
        section_count = Section.objects.count()
        response = self.client.post(
            reverse("add_material_to_course"),
            json.dumps({"course_id": f"section-{self.section.pk}", "material_id": shared.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        shared.refresh_from_db()
        self.assertEqual(shared.teacher_id, self.other_teacher.id)
        self.assertTrue(shared.assigned_sections.filter(pk=self.section.pk).exists())
        self.assertEqual(Section.objects.count(), section_count)
        self.assertEqual(Course.objects.count(), 0)

    def test_other_teachers_private_unassigned_material_cannot_be_assigned(self):
        private = Material.objects.create(
            teacher=self.other_teacher, title="Private reading", item_type="word",
            content_text="Lihim", content_json={"items": ["Lihim"]}, type="assessment",
            source_type="personal", status="published", is_active=True,
        )
        response = self.client.post(
            reverse("add_material_to_course"),
            json.dumps({"course_id": f"section-{self.section.pk}", "material_id": private.pk}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 403)
        self.assertFalse(private.assigned_sections.exists())

    def test_only_students_enrolled_in_assigned_section_can_fetch_material(self):
        material = Material.objects.create(
            teacher=self.teacher, section=self.section, title="Bonifacio only", item_type="word",
            content_text="Bayani", content_json={"items": ["Bayani"]}, type="assessment",
            source_type="personal", status="published", is_active=True,
        )
        material.assigned_sections.add(self.section)

        self._login(self.student)
        response = self.client.get(reverse("get_class_materials"), {"class_code": self.section.class_code})
        self.assertEqual(response.status_code, 200)
        self.assertTrue(any(item["raw_id"] == material.pk for item in response.json()["all_materials"]))

        self._login(self.outside_student)
        response = self.client.get(reverse("get_class_materials"), {"class_code": self.section.class_code})
        self.assertEqual(response.status_code, 403)


class PracticeReaderMaterialTests(TestCase):
    def setUp(self):
        self.student = User.objects.create(
            custom_id="STD-PRACT",
            role="student",
            first_name="Practice",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2012,
            email="practice-student@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 1",
        )
        session = self.client.session
        session["user_id"] = self.student.id
        session["user_role"] = self.student.role
        session["first_name"] = self.student.first_name
        session["last_name"] = self.student.last_name
        session["email"] = self.student.email
        session["custom_id"] = self.student.custom_id
        session.save()

    def test_word_reader_receives_active_published_practice_items(self):
        Material.objects.create(
            title="Easy syllables",
            item_type="word",
            content_text="HA\nhe\nhi\nho\nhu",
            content_json={"source": "admin_practice", "difficulty": "easy", "items": ["HA", "he", "hi", "ho", "hu"]},
            status="published",
            difficulty_level="easy",
            is_active=True,
        )
        Material.objects.create(
            title="Draft syllables",
            item_type="word",
            content_text="draft",
            content_json={"items": ["draft"]},
            status="draft",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.get(reverse("practice_word_page"), {"difficulty": "easy"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="practiceMaterialsData"', html=False)
        self.assertContains(response, '"items": ["HA", "he", "hi", "ho", "hu"]', html=False)
        self.assertNotContains(response, "Draft syllables")

    def test_word_reader_parses_comma_separated_content_text_when_json_items_missing(self):
        Material.objects.create(
            title="Comma syllables",
            item_type="word",
            content_text="HA, he, hi, ho, hu",
            content_json={},
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.get(reverse("practice_word_page"), {"difficulty": "easy"})

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '"items": ["HA", "he", "hi", "ho", "hu"]', html=False)

    def test_practice_completion_records_student_done_status(self):
        material = Material.objects.create(
            title="Completion syllables",
            item_type="word",
            content_text="HA\nhe",
            content_json={"items": ["HA", "he"]},
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": f"practice-{material.id}",
                "activity_type": "practice",
                "stars_earned": 20,
                "items_completed": 2,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertEqual(payload["material_id"], f"practice-{material.id}")
        self.assertEqual(payload["status"], "Done")

        material.refresh_from_db()
        completion = material.content_json["student_completions"][str(self.student.id)]
        self.assertEqual(completion["status"], "completed")
        self.assertEqual(completion["stars_earned"], 20)
        self.assertEqual(completion["items_completed"], 2)
        self.assertEqual(material.status, "published")

    def test_practice_completion_saves_detailed_results_metrics(self):
        material = Material.objects.create(
            title="Scored practice",
            item_type="word",
            content_text="HA\nhe",
            content_json={"items": ["HA", "he"]},
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": f"practice-{material.id}",
                "activity_type": "practice",
                "stars_earned": 20,
                "items_completed": 2,
                "correct_responses": 2,
                "incorrect_responses": 0,
                "reading_time_seconds": 45,
                "attempt_number": 1,
                "total_practice_items": 2,
                "total_read_words": 2,
                "total_skipped_words": 0,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])

        material.refresh_from_db()
        completion = material.content_json["student_completions"][str(self.student.id)]
        self.assertEqual(completion["score"], 100)
        self.assertEqual(completion["accuracy"], 100)
        self.assertEqual(completion["correct_responses"], 2)
        self.assertEqual(completion["incorrect_responses"], 0)
        self.assertEqual(completion["reading_time_seconds"], 45)
        self.assertEqual(completion["attempt_number"], 1)
        self.assertEqual(completion["total_practice_items"], 2)
        self.assertEqual(completion["total_read_words"], 2)
        self.assertEqual(completion["total_skipped_words"], 0)

    def test_color_mode_replay_never_reduces_saved_stars(self):
        material = Material.objects.create(
            title="Color score protection",
            item_type="word",
            content_text="HA\nhe\nhi\nho\nhu",
            content_json={"mode": "color", "difficulty": "easy", "level": "level_1"},
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )
        endpoint = reverse("record_assessment_completion")
        base_payload = {
            "material_id": f"practice-{material.id}",
            "activity_type": "practice",
            "game_mode": "color",
            "items_completed": 5,
        }

        first = self.client.post(endpoint, data=json.dumps({**base_payload, "stars_earned": 50}), content_type="application/json")
        replay = self.client.post(endpoint, data=json.dumps({**base_payload, "stars_earned": 20}), content_type="application/json")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(replay.status_code, 200)
        material.refresh_from_db()
        completion = material.content_json["student_completions"][str(self.student.id)]
        self.assertEqual(completion["stars_earned"], 50)

    def _hunt_material(self, level="level_1"):
        return Material.objects.create(
            title=f"Hunt {level}", item_type="word", content_text="one\ntwo\nthree\nfour\nfive",
            content_json={"mode": "hunt", "difficulty": "easy", "level": level},
            type="practice", status="published", difficulty_level="easy", is_active=True,
        )

    def _award_hunt(self, material, points):
        stars = 3 if points >= 8 else 2 if points >= 5 else 1
        return self.client.post(reverse("award_hunt_mode_stars"), data=json.dumps({
            "student_id": self.student.id, "level_id": f"practice-{material.id}",
            "total_points": points, "percentage": points * 10, "earned_stars": stars,
        }), content_type="application/json")

    def test_hunt_star_first_completion(self):
        payload = self._award_hunt(self._hunt_material(), 5).json()
        self.assertEqual((payload["earned_stars"], payload["star_delta"], payload["best_stars"]), (2, 2, 2))
        self.assertEqual((payload["total_stars_earned"], payload["available_stars"]), (2, 2))

    def test_hunt_star_improved_replay(self):
        material = self._hunt_material()
        self._award_hunt(material, 5)
        payload = self._award_hunt(material, 8).json()
        self.assertEqual((payload["star_delta"], payload["best_stars"], payload["available_stars"]), (1, 3, 3))

    def test_hunt_star_equal_replay(self):
        material = self._hunt_material()
        self._award_hunt(material, 5)
        payload = self._award_hunt(material, 7).json()
        self.assertEqual((payload["star_delta"], payload["best_stars"], payload["available_stars"]), (0, 2, 2))

    def test_hunt_star_lower_replay(self):
        material = self._hunt_material()
        self._award_hunt(material, 8)
        payload = self._award_hunt(material, 2).json()
        self.assertEqual((payload["star_delta"], payload["best_stars"], payload["available_stars"]), (0, 3, 3))

    def test_hunt_stars_are_independent_per_level(self):
        first = self._hunt_material("level_1")
        second = self._hunt_material("level_2")
        self._award_hunt(first, 5)
        payload = self._award_hunt(second, 8).json()
        self.assertEqual((payload["star_delta"], payload["total_stars_earned"], payload["available_stars"]), (3, 5, 5))

    def test_hunt_star_endpoint_requires_authenticated_student(self):
        material = self._hunt_material()
        self.client.logout()
        response = self._award_hunt(material, 5)
        self.assertIn(response.status_code, {302, 401, 403})

    def test_practice_results_route_redirects_to_shared_practice_flow(self):
        material = Material.objects.create(
            title="Results practice",
            item_type="word",
            content_text="HA\nhe",
            content_json={
                "items": ["HA", "he"],
                "student_completions": {
                    str(self.student.id): {
                        "student_id": self.student.id,
                        "status": "completed",
                        "completed_at": timezone.now().isoformat(),
                        "score": 80,
                        "accuracy": 80,
                        "correct_responses": 2,
                        "incorrect_responses": 1,
                        "reading_time_seconds": 60,
                        "attempt_number": 1,
                    }
                },
            },
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.get(reverse("practice_results"), {"id": f"practice-{material.id}"})

        self.assertEqual(response.status_code, 302)
        self.assertEqual(response.url, reverse("practice"))

    def test_progression_page_renders_unlock_states_for_game_levels(self):
        Material.objects.create(
            title="Free Easy Level 1",
            item_type="word",
            content_text="sun",
            content_json={
                "mode": "free",
                "difficulty": "easy",
                "level": "level_1",
                "items": ["sun"],
                "student_completions": {
                    str(self.student.id): {
                        "student_id": self.student.id,
                        "status": "completed",
                        "completed_at": timezone.now().isoformat(),
                        "stars_earned": 3,
                    }
                },
            },
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )
        Material.objects.create(
            title="Free Easy Level 2",
            item_type="word",
            content_text="moon",
            content_json={
                "mode": "free",
                "difficulty": "easy",
                "level": "level_2",
                "items": ["moon"],
            },
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Free Mode Adventure")
        self.assertContains(response, "Level 1")
        self.assertContains(response, "Level 2")
        self.assertContains(response, "Free Mode • Easy")
        self.assertContains(response, "Complete Easy Level 1 to unlock this level.")

    def test_progression_page_uses_dynamic_unlock_copy_across_difficulties(self):
        Material.objects.create(
            title="Free Medium Level 1",
            item_type="word",
            content_text="sun",
            content_json={
                "mode": "free",
                "difficulty": "medium",
                "level": "level_1",
                "items": ["sun"],
            },
            type="practice",
            status="published",
            difficulty_level="medium",
            is_active=True,
        )
        Material.objects.create(
            title="Free Medium Level 2",
            item_type="word",
            content_text="moon",
            content_json={
                "mode": "free",
                "difficulty": "medium",
                "level": "level_2",
                "items": ["moon"],
            },
            type="practice",
            status="published",
            difficulty_level="medium",
            is_active=True,
        )
        Material.objects.create(
            title="Free Hard Level 1",
            item_type="word",
            content_text="star",
            content_json={
                "mode": "free",
                "difficulty": "hard",
                "level": "level_1",
                "items": ["star"],
            },
            type="practice",
            status="published",
            difficulty_level="hard",
            is_active=True,
        )

        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Free Mode • Medium")
        self.assertContains(response, "Free Mode • Hard")
        self.assertContains(response, "Complete Easy Level 1 to unlock this level.")
        self.assertContains(response, "Complete Medium Level 1 to unlock this level.")
        self.assertContains(response, "Complete all Easy levels to unlock this difficulty.")
        self.assertContains(response, "Complete all Medium levels to unlock this difficulty.")

    def test_progression_page_marks_levels_without_content_as_unavailable(self):
        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Content unavailable")

    def test_progression_page_shows_mode_tutorial_and_how_to_play_button(self):
        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "How to Play")
        self.assertContains(response, "Welcome to Free Mode! Read the word aloud.")
        self.assertContains(response, 'data-tutorial-mode="free"')

    def test_progression_page_prefers_active_session_for_current_challenge(self):
        for difficulty, level, expected in [
            ("easy", "level_4", "Easy • Level 4"),
            ("medium", "level_2", "Medium • Level 2"),
            ("hard", "level_3", "Hard • Level 3"),
        ]:
            session = self.client.session
            session["practice_active_session"] = {"game": "free", "difficulty": difficulty, "level": level}
            session.save()

            response = self.client.get(reverse("practice_game_progression", args=["free"]))

            self.assertEqual(response.status_code, 200)
            self.assertEqual(response.context["game_progression_summary"]["current_challenge_label"], expected)
            self.assertEqual(response.context["game_progression_summary"]["current_difficulty"], difficulty.title())
            self.assertEqual(response.context["game_progression_summary"]["current_level"], f"Level {level.split('_', 1)[1]}")

    def test_progression_page_falls_back_to_summary_when_active_session_is_missing(self):
        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["game_progression_summary"]["current_challenge_label"], "Easy • Level 1")

    def test_student_theme_shop_renders_ui_only_catalog(self):
        response = self.client.get(reverse("theme_shop"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Theme Shop")
        self.assertContains(response, "Sky Island")
        self.assertContains(response, "Magic Library")
        self.assertContains(response, "Light and Dark Mode stay separate.")

    def test_profile_uses_the_students_equipped_theme(self):
        self.student.equipped_theme = "forest"
        self.student.save(update_fields=["equipped_theme", "updated_at"])

        response = self.client.get(reverse("profile"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["student_theme_slug"], "forest")
        self.assertContains(response, 'data-student-theme="forest"')

    def test_forest_dark_mode_styles_shared_custom_select_menus(self):
        template = (Path(__file__).resolve().parent / "templates" / "pabasa_app" / "base_dashboard.html").read_text(encoding="utf-8")
        selector = 'body.dark-theme.dashboard-body[data-student-theme="forest"] .custom-device-menu'
        self.assertIn(selector, template)
        self.assertIn('.custom-device-option.is-selected', template)
        self.assertIn('.custom-device-option-title {color:#f3f2d3 !important}', template)
        self.assertIn('scrollbar-color:#708b55 #102e25', template)

    def test_every_student_theme_has_readable_dark_custom_selects(self):
        template = (Path(__file__).resolve().parent / "templates" / "pabasa_app" / "base_dashboard.html").read_text(encoding="utf-8")
        self.assertIn('body.dark-theme.dashboard-body[data-student-theme] .custom-device-menu', template)
        for slug in ("forest", "treasure", "ocean", "space", "zoo", "library"):
            self.assertIn(f'data-student-theme="{slug}"] {{--theme-dark-accent:', template)
        self.assertIn('.custom-device-option-title {color:#f8fafc !important}', template)
        self.assertIn('background:#111827 !important', template)

    def test_theme_unlock_is_charged_once_and_can_be_equipped(self):
        self.student.available_stars = 200
        self.student.unlocked_themes = ["sky"]
        self.student.equipped_theme = "sky"
        self.student.save(update_fields=["available_stars", "unlocked_themes", "equipped_theme", "updated_at"])
        endpoint = reverse("student_theme_action")

        first = self.client.post(endpoint, data=json.dumps({"theme": "forest", "action": "unlock"}), content_type="application/json")
        duplicate = self.client.post(endpoint, data=json.dumps({"theme": "forest", "action": "unlock"}), content_type="application/json")
        equipped = self.client.post(endpoint, data=json.dumps({"theme": "forest", "action": "equip"}), content_type="application/json")

        self.assertEqual(first.status_code, 200)
        self.assertEqual(duplicate.status_code, 200)
        self.assertEqual(equipped.status_code, 200)
        self.student.refresh_from_db()
        self.assertEqual(self.student.available_stars, 125)
        self.assertIn("forest", self.student.unlocked_themes)
        self.assertEqual(self.student.equipped_theme, "forest")

    def test_tutorial_header_shows_refined_storybook_prompt(self):
        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Quick Guide")
        self.assertContains(response, "Learn at your own pace.")

    def test_first_time_tutorial_overlay_shows_start_button(self):
        response = self.client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'class="tutorial-overlay is-open')
        self.assertContains(response, 'tutorial-start-btn')
        self.student.refresh_from_db()
        self.assertTrue(self.student.preference.get("free_mode_tutorial_seen"))

    def test_tutorial_auto_opens_only_on_first_mode_visit(self):
        first_response = self.client.get(reverse("practice_game_progression", args=["hunt"]))
        second_response = self.client.get(reverse("practice_game_progression", args=["hunt"]))

        self.assertContains(first_response, 'class="tutorial-overlay is-open')
        self.assertNotContains(second_response, 'class="tutorial-overlay is-open')

    def test_seen_tutorial_overlay_does_not_auto_open_after_new_session(self):
        self.student.preference = {"free_mode_tutorial_seen": True}
        self.student.save(update_fields=["preference", "updated_at"])
        new_client = Client()
        session = new_client.session
        session["user_id"] = self.student.id
        session["user_role"] = self.student.role
        session["first_name"] = self.student.first_name
        session["last_name"] = self.student.last_name
        session["email"] = self.student.email
        session["custom_id"] = self.student.custom_id
        session.save()

        response = new_client.get(reverse("practice_game_progression", args=["free"]))

        self.assertEqual(response.status_code, 200)
        self.assertNotContains(response, 'class="tutorial-overlay is-open')
        self.assertContains(response, 'tutorial-start-btn')

    def test_mark_tutorial_seen_sets_user_preference_flag(self):
        response = self.client.post(reverse("practice_mark_tutorial_seen", args=["color"]))

        self.assertEqual(response.status_code, 302)
        self.assertTrue(self.client.session.get("color_mode_tutorial_seen"))
        self.student.refresh_from_db()
        self.assertTrue(self.student.preference.get("color_mode_tutorial_seen"))

    def test_practice_reader_template_shows_results_breakdown(self):
        response = self.client.get(reverse("practice_word_page"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Score Breakdown")
        self.assertContains(response, "Total Practice Items")
        self.assertContains(response, "Total Read Words")
        self.assertContains(response, "Total Skipped Words")

    def test_practice_hub_marks_only_completed_student_material_done(self):
        material = Material.objects.create(
            title="Done for one student",
            item_type="word",
            content_text="HA\nhe",
            content_json={
                "items": ["HA", "he"],
                "student_completions": {
                    str(self.student.id): {
                        "student_id": self.student.id,
                        "status": "completed",
                        "completed_at": timezone.now().isoformat(),
                    }
                },
            },
            type="practice",
            status="published",
            difficulty_level="easy",
            is_active=True,
        )

        response = self.client.get(reverse("practice"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'"id": "practice-{material.id}"', html=False)
        self.assertContains(response, '"status": "Done"', html=False)
        self.assertContains(response, '"is_done": true', html=False)

        other_student = User.objects.create(
            custom_id="STD-OTHER-PRACT",
            role="student",
            first_name="Other",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=2012,
            email="other-practice-student@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 1",
        )
        session = self.client.session
        session["user_id"] = other_student.id
        session["user_role"] = other_student.role
        session["first_name"] = other_student.first_name
        session["last_name"] = other_student.last_name
        session["email"] = other_student.email
        session["custom_id"] = other_student.custom_id
        session.save()

        response = self.client.get(reverse("practice"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, f'"id": "practice-{material.id}"', html=False)
        self.assertContains(response, '"status": "published"', html=False)
        self.assertContains(response, '"is_done": false', html=False)


class AdminPracticeAssessmentListTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create(
            custom_id="ADM-PRACT-LIST",
            role="admin",
            first_name="Practice",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="practice-admin-list@example.com",
            password_hash=make_password("admin-password"),
        )

    def test_admin_practice_assessment_uses_mode_for_item_label(self):
        Material.objects.create(
            title="Free Hard Words",
            item_type='word',
            prompt_text='',
            content_text='sun\nmoon',
            content_json={'mode': 'free', 'difficulty': 'hard', 'level': 'level_1', 'items': ['sun', 'moon']},
            type='practice',
            status='published',
            difficulty_level='hard',
            language='English',
            is_active=True,
        )
        Material.objects.create(
            title="Color Easy Sentences",
            item_type='sentence',
            prompt_text='',
            content_text='One sentence.\nTwo sentence.',
            content_json={'mode': 'color', 'difficulty': 'easy', 'level': 'level_1', 'items': ['One sentence.', 'Two sentence.']},
            type='practice',
            status='published',
            difficulty_level='easy',
            language='English',
            is_active=True,
        )
        Material.objects.create(
            title="Hunt Medium Paragraphs",
            item_type='paragraph',
            prompt_text='',
            content_text='First paragraph.\n\nSecond paragraph.',
            content_json={'mode': 'hunt', 'difficulty': 'medium', 'level': 'level_1', 'items': ['First paragraph.', 'Second paragraph.']},
            type='practice',
            status='published',
            difficulty_level='medium',
            language='English',
            is_active=True,
        )

        session = self.client.session
        session['user_id'] = self.admin.id
        session['user_role'] = self.admin.role
        session['first_name'] = self.admin.first_name
        session['last_name'] = self.admin.last_name
        session['email'] = self.admin.email
        session['custom_id'] = self.admin.custom_id
        session.save()

        response = self.client.get(reverse('admin_practice_assessment'))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '2 words')
        self.assertContains(response, '2 sentences')
        self.assertContains(response, '2 paragraphs')


class AdminPracticeMaterialFormTests(TestCase):
    def test_free_mode_accepts_words_for_any_difficulty(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'free',
            'difficulty_level': 'hard',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'sun\nmoon',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.practice_items(), ['sun', 'moon'])

    def test_color_mode_items_are_limited_to_five_sentences_per_level(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'color',
            'difficulty_level': 'medium',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'One sentence.\nTwo sentence.\nThree sentence.\nFour sentence.\nFive sentence.\nSix sentence.',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('content_text', form.errors)

    def test_color_mode_requires_sentences_even_when_difficulty_is_easy(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'color',
            'difficulty_level': 'easy',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'sun',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('content_text', form.errors)

    def test_hunt_mode_accepts_paragraphs_for_any_difficulty(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'hunt',
            'difficulty_level': 'easy',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'First paragraph here.\nSecond paragraph here.',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.practice_items(), ['First paragraph here.', 'Second paragraph here.'])

    def test_duplicate_mode_difficulty_and_level_is_rejected(self):
        Material.objects.create(
            title='Existing Practice',
            item_type='word',
            prompt_text='',
            content_text='sun',
            content_json={'mode': 'free', 'difficulty': 'easy', 'level': 'level_1'},
            type='practice',
            status='published',
            difficulty_level='easy',
            is_active=True,
        )

        form = AdminPracticeMaterialForm(data={
            'mode': 'free',
            'difficulty_level': 'easy',
            'level': 'level_1',
            'status': 'draft',
            'content_text': 'sun',
        })

        self.assertFalse(form.is_valid())
        self.assertIn('__all__', form.errors)

    def test_color_mode_sentence_keeps_commas_inside_the_sentence(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'color',
            'difficulty_level': 'hard',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'The cat ran home, and it slept on the couch.',
        })

        self.assertTrue(form.is_valid(), form.errors)
        self.assertEqual(form.practice_items(), ['The cat ran home, and it slept on the couch.'])

    def test_save_helper_uses_mode_for_item_type(self):
        form = AdminPracticeMaterialForm(data={
            'mode': 'color',
            'difficulty_level': 'hard',
            'level': 'level_1',
            'status': 'draft',
            'language': 'English',
            'content_text': 'First sentence.\nSecond sentence.',
        })

        self.assertTrue(form.is_valid(), form.errors)
        material = _save_admin_practice_material(form)

        self.assertEqual(material.item_type, 'sentence')
        self.assertEqual(material.content_text, 'First sentence.\nSecond sentence.')

    def test_occupied_levels_are_detected_for_a_configuration(self):
        Material.objects.create(
            title='Existing Practice',
            item_type='word',
            prompt_text='',
            content_text='sun',
            content_json={'mode': 'free', 'difficulty': 'easy', 'level': 'level_1'},
            type='practice',
            status='published',
            difficulty_level='easy',
            is_active=True,
        )

        form = AdminPracticeMaterialForm()
        occupied_levels = form.get_occupied_levels('free', 'easy')

        self.assertEqual(occupied_levels, ['level_1'])


class AdminPracticeTemplateTests(TestCase):
    def test_practice_create_template_uses_mode_based_item_copy(self):
        template_path = Path(__file__).resolve().parent / 'templates' / 'pabasa_app' / 'admin_practice_create.html'
        content = template_path.read_text(encoding='utf-8')

        self.assertIn('Choose a game mode to see item guidance.', content)
        self.assertIn('Color Mode allows only up to 5 sentences for each level.', content)
        self.assertIn('Words Added', content)
        self.assertIn('Enter one paragraph per line...', content)

    def test_practice_edit_template_uses_mode_based_item_copy(self):
        template_path = Path(__file__).resolve().parent / 'templates' / 'pabasa_app' / 'admin_practice_edit.html'
        content = template_path.read_text(encoding='utf-8')

        self.assertIn('Choose a game mode to see item guidance.', content)
        self.assertIn('This section will change based on Game Mode', content)
        self.assertIn('Enter one paragraph per line', content)


class PracticeAccessControlTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id="TCH-PRACT",
            role="teacher",
            first_name="Practice",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="practice-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.admin = User.objects.create(
            custom_id="ADM-PRACT",
            role="admin",
            first_name="Practice",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="practice-admin@example.com",
            password_hash=make_password("admin-password"),
        )

    def test_teacher_is_redirected_from_practice_reader(self):
        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session["first_name"] = self.teacher.first_name
        session["last_name"] = self.teacher.last_name
        session["email"] = self.teacher.email
        session["custom_id"] = self.teacher.custom_id
        session.save()

        response = self.client.get(reverse("practice_word_page"))

        self.assertEqual(response.status_code, 302)
        self.assertIn(reverse("auth"), response.url)

    def test_admin_can_open_practice_assessment_management(self):
        session = self.client.session
        session["user_id"] = self.admin.id
        session["user_role"] = self.admin.role
        session["first_name"] = self.admin.first_name
        session["last_name"] = self.admin.last_name
        session["email"] = self.admin.email
        session["custom_id"] = self.admin.custom_id
        session.save()

        response = self.client.get(reverse("admin_practice_assessment"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Practice Content")


class SettingsViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(
            custom_id="STD-0001",
            role="student",
            first_name="Settings",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2012,
            email="settings@example.com",
            password_hash=make_password("old-password"),
            grade_level="Grade 1",
        )
        session = self.client.session
        session["user_id"] = self.user.id
        session["user_role"] = self.user.role
        session["first_name"] = self.user.first_name
        session["last_name"] = self.user.last_name
        session["email"] = self.user.email
        session["custom_id"] = self.user.custom_id
        session.save()

    def test_settings_password_change_updates_user_hash(self):
        response = self.client.post(
            reverse("settings"),
            {
                "settings_action": "change_password",
                "current_password": "old-password",
                "new_password": "new-password",
                "confirm_password": "new-password",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Password changed successfully.")

        self.user.refresh_from_db()
        self.assertTrue(check_password("new-password", self.user.password_hash))

    def test_settings_saves_push_notification_preferences(self):
        response = self.client.post(
            reverse("settings"),
            {
                "settings_action": "save_notifications",
                "push_enabled": "on",
                "new_materials": "on",
                "progress_updates": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Push notification preferences saved.")

        self.user.refresh_from_db()
        self.assertIn({
            "notification_settings": {
                "push_enabled": True,
                "email_notifications": False,
                "weekly_digest_enabled": False,
                "new_materials": True,
                "reading_reminders": False,
                "progress_updates": True,
            }
        }, self.user.tags)

    def test_settings_saves_weekly_digest_preference(self):
        response = self.client.post(
            reverse("settings"),
            {
                "settings_action": "save_notifications",
                "push_enabled": "on",
                "email_notifications": "on",
                "weekly_digest_enabled": "on",
                "new_materials": "on",
                "reading_reminders": "on",
                "progress_updates": "on",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.user.refresh_from_db()
        self.assertTrue(self.user.preference["notification_settings"]["weekly_digest_enabled"])

        fresh_response = self.client.get(reverse("settings"))
        self.assertEqual(fresh_response.status_code, 200)
        self.assertContains(fresh_response, 'id="weeklyDigestEnabled"')
        self.assertContains(fresh_response, 'name="weekly_digest_enabled" checked')


class AdminSettingsRenderTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(
            custom_id="ADM-9999",
            role="admin",
            first_name="Admin",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="admin-settings@example.com",
            password_hash=make_password("admin-password"),
        )
        session = self.client.session
        session["user_id"] = self.user.id
        session["user_role"] = self.user.role
        session["first_name"] = self.user.first_name
        session["last_name"] = self.user.last_name
        session["email"] = self.user.email
        session["custom_id"] = self.user.custom_id
        session.save()

    def test_admin_settings_renders_new_settings_ui(self):
        response = self.client.get(reverse("admin_settings"))

        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "pabasa_app/admin_settings.html")
        self.assertNotContains(response, "Assessment Window Management")
        self.assertNotContains(response, "Active Assessment Window")
        self.assertNotContains(response, "Save Window")
        self.assertContains(response, "Push Notifications")
        self.assertContains(response, "Enable Notifications")
        self.assertContains(response, "Email Notifications")
        self.assertContains(response, "Save Preferences")
        self.assertContains(response, "Change Password")
        self.assertContains(response, "Current Password")
        self.assertContains(response, "New Password")
        self.assertContains(response, "Confirm Password")
        self.assertContains(response, "Update Password")
        self.assertNotContains(response, "Settings placeholder. CRUD is not implemented yet.")


class PrincipalNotificationTests(TestCase):
    def test_notify_principals_creates_in_app_notifications(self):
        principal = User.objects.create(
            custom_id=f"PRN-{uuid.uuid4().hex[:8].upper()}",
            role="principal",
            first_name="Principal",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="principal-notify@example.com",
            password_hash=make_password("principal-password"),
        )

        result = _notify_principals("School update", "A principal alert should be stored.", "success")

        self.assertEqual(result, 1)
        self.assertTrue(Notification.objects.filter(recipient=principal, title="School update").exists())

    def test_teacher_cannot_create_a_canonical_section(self):
        principal = User.objects.create(
            custom_id=f"PRN-{uuid.uuid4().hex[:8].upper()}",
            role="principal",
            first_name="Principal",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="principal-class@example.com",
            password_hash=make_password("principal-password"),
        )
        teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}",
            role="teacher",
            first_name="Teacher",
            last_name="User",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="teacher-class@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session["first_name"] = teacher.first_name
        session["last_name"] = teacher.last_name
        session["email"] = teacher.email
        session["custom_id"] = teacher.custom_id
        session.save()

        response = self.client.post(
            reverse("create_reading_class"),
            json.dumps({
                "grade": "Grade 1",
                "section": "A",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 410)
        payload = response.json()
        self.assertFalse(payload['success'])
        self.assertIn('Admin', payload['error'])
        self.assertFalse(Section.objects.filter(grade_level='Grade 1', section='A').exists())
        self.assertFalse(Notification.objects.filter(recipient=principal, title__icontains="new class").exists())

    def test_legacy_create_class_endpoint_rejects_any_payload(self):
        teacher = User.objects.create(
            custom_id=f"TCH-{uuid.uuid4().hex[:8].upper()}", role="teacher",
            first_name="Teacher", last_name="Validation", middle_initial="", suffix="",
            sex="female", birth_month=1, birth_day=1, birth_year=1990,
            email="teacher-validation@example.com", password_hash=make_password("password"),
            teacher_role="Teacher",
        )
        session = self.client.session
        session.update({"user_id": teacher.id, "user_role": "teacher", "email": teacher.email})
        session.save()
        response = self.client.post(
            reverse("create_reading_class"), json.dumps({"grade": "", "section": ""}),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 410)
        self.assertIn('Admin', response.json()['error'])


class PreferenceDeliveryTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(
            custom_id="TCH-PREF",
            role="teacher",
            first_name="Pref",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="pref-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )

    def test_disabled_in_app_alerts_do_not_create_notification_rows(self):
        self.user.preference = {
            "notification_settings": {
                "push_enabled": False,
                "email_notifications": False,
                "weekly_digest_enabled": False,
            }
        }
        self.user.save(update_fields=["preference", "updated_at"])

        result = _create_notification(
            self.user,
            "Hidden alert",
            "This should not be stored.",
            send_email=False,
        )

        self.assertIsNone(result)
        self.assertFalse(Notification.objects.filter(recipient=self.user).exists())

    @patch("pabasa_app.views.send_mail")
    def test_email_alerts_respect_email_preference(self, mock_send_mail):
        self.user.preference = {
            "notification_settings": {
                "push_enabled": True,
                "email_notifications": False,
                "weekly_digest_enabled": False,
            }
        }
        self.user.save(update_fields=["preference", "updated_at"])

        _create_notification(self.user, "Stored only", "No email should be sent.")

        self.assertEqual(Notification.objects.filter(recipient=self.user).count(), 1)
        mock_send_mail.assert_not_called()

        self.user.preference["notification_settings"]["email_notifications"] = True
        self.user.save(update_fields=["preference", "updated_at"])
        _create_notification(self.user, "Stored and emailed", "Email should be sent.")

        self.assertEqual(Notification.objects.filter(recipient=self.user).count(), 2)
        mock_send_mail.assert_called_once()


class PrincipalSettingsViewTests(TestCase):
    def setUp(self):
        self.user = User.objects.create(
            custom_id="PRN-SES",
            role="principal",
            first_name="Jobelyn",
            last_name="Valdez",
            middle_initial="A",
            suffix="",
            sex="female",
            birth_month=6,
            birth_day=3,
            birth_year=1980,
            email="principal@example.com",
            password_hash=make_password("old-password"),
        )
        session = self.client.session
        session["user_id"] = self.user.id
        session["user_role"] = self.user.role
        session["first_name"] = self.user.first_name
        session["last_name"] = self.user.last_name
        session["email"] = self.user.email
        session["custom_id"] = self.user.custom_id
        session.save()

    def test_principal_settings_saves_school_information(self):
        response = self.client.post(
            reverse("principal_settings"),
            {
                "settings_action": "save_school_info",
                "school_name": "Example Elementary School",
                "school_code": "EX-001",
                "municipality": "Imus",
                "province": "Cavite",
                "district": "District 5",
                "region": "CALABARZON",
                "address": "Example Street",
                "contact": "0917-123-4567",
                "email": "principal@example.org",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "School information updated.")

        self.user.refresh_from_db()
        self.assertEqual(self.user.preference["principal_school_info"]["name"], "Example Elementary School")
        self.assertEqual(self.user.preference["principal_school_info"]["code"], "EX-001")

    def test_principal_settings_changes_password(self):
        response = self.client.post(
            reverse("principal_settings"),
            {
                "settings_action": "change_password",
                "current_password": "old-password",
                "new_password": "new-password",
                "confirm_password": "new-password",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Password changed successfully.")

        self.user.refresh_from_db()
        self.assertTrue(check_password("new-password", self.user.password_hash))

    def test_principal_settings_updates_personal_information(self):
        response = self.client.post(
            reverse("principal_settings"),
            {
                "settings_action": "save_personal_info",
                "first_name": "Maria",
                "last_name": "Cruz",
                "middle_initial": "L",
                "email": "maria.cruz@example.org",
                "contact_number": "09171234567",
                "position": "Principal II",
            },
        )

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Personal information updated.")

        self.user.refresh_from_db()
        self.assertEqual(self.user.first_name, "Maria")
        self.assertEqual(self.user.last_name, "Cruz")
        self.assertEqual(self.user.middle_initial, "L")
        self.assertEqual(self.user.email, "maria.cruz@example.org")
        self.assertEqual(self.user.contact_no, "09171234567")
        self.assertEqual(self.user.preference["principal_profile_info"]["position"], "Principal II")


class LiveAssessmentWaitingRoomTemplateTests(TestCase):
    def test_waiting_room_template_does_not_render_a_separate_countdown(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "live_assessment_waiting_room.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertNotIn('waitingRoomCountdown', content)
        self.assertNotIn('startCountdownClock', content)

    def test_reader_template_still_uses_the_live_countdown_overlay(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "reading_assessment_base.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn('id="liveCountdownOverlay"', content)

    def test_dashboard_template_prioritizes_waiting_room_over_modal_for_pending_sessions(self):
        template_path = Path(__file__).resolve().parent / "templates" / "pabasa_app" / "base_dashboard.html"
        content = template_path.read_text(encoding="utf-8")

        self.assertIn("['waiting', 'countdown'].includes(sessionStatus)", content)
        self.assertIn("window.location.assign(session.join_url)", content)


class AssessmentCompletionNotificationTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id="TCH-9001",
            role="teacher",
            first_name="Taylor",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="teacher9001@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.student = User.objects.create(
            custom_id="STU-9001",
            role="student",
            first_name="Jane",
            last_name="Doe",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=2,
            birth_day=2,
            birth_year=2012,
            email="student9001@example.com",
            password_hash=make_password("student-password"),
        )
        self.admin = User.objects.create(
            custom_id="ADM-9001",
            role="admin",
            first_name="Alex",
            last_name="Admin",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=3,
            birth_day=3,
            birth_year=1985,
            email="admin9001@example.com",
            password_hash=make_password("admin-password"),
        )
        self.principal = User.objects.create(
            custom_id="PRN-9001",
            role="principal",
            first_name="Paula",
            last_name="Principal",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=4,
            birth_day=4,
            birth_year=1980,
            email="principal9001@example.com",
            password_hash=make_password("principal-password"),
        )
        self.section = test_section_create(
            teacher=self.teacher,
            class_name="Class A",
            class_code="CLS-A9001",
            subject="Reading",
            is_active=True,
        )
        self.section.add_student(self.student)
        self.assessment = Assessment.objects.create(
            title="Reading Fluency Test",
            code="ASM-9001",
            assessment_type="word",
            content="cat\ndog",
            teacher=self.teacher,
            section=self.section,
            is_active=True,
        )

    def _login_student(self):
        session = self.client.session
        session["user_id"] = self.student.id
        session["user_role"] = self.student.role
        session.save()

    def _login_teacher(self):
        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session.save()

    def _login_admin(self):
        session = self.client.session
        session["user_id"] = self.admin.id
        session["user_role"] = self.admin.role
        session.save()

    def _login_principal(self):
        session = self.client.session
        session["user_id"] = self.principal.id
        session["user_role"] = self.principal.role
        session.save()

    def test_assessment_material_creation_leaves_assessments_table_empty_until_completion(self):
        material = Material.objects.create(
            title="New Assessment Material",
            item_type="word",
            content_text="cat\ndog",
            content_json={"items": ["cat", "dog"]},
            type="assessment",
            status="published",
            is_active=True,
        )

        self.assertIsNone(material.assessment)
        self.assertEqual(Assessment.objects.count(), 0)

        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": f"material-{material.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        material.refresh_from_db()
        self.assertIsNotNone(material.assessment)
        self.assertEqual(Assessment.objects.count(), 1)
        self.assertEqual(material.assessment.get_student_attempt_count(self.student), 1)

    def test_phrase_reading_requires_all_ten_phrases_and_persists_completion(self):
        material = Material.objects.create(
            title="Phrase Reading",
            item_type="phrase",
            content_json={
                "template_title": "Phrase Reading",
                "items": [f"Phrase {index}" for index in range(1, 11)],
            },
            type="assessment",
            status="published",
            is_active=True,
            section=self.section,
        )
        self._login_student()
        endpoint = reverse("record_assessment_completion")
        partial_payload = {
            "material_id": f"material-{material.id}",
            "activity_type": "phrase_reading",
            "items_completed": 9,
            "scores": {"completed_phrases": list(range(9))},
        }

        partial = self.client.post(endpoint, data=json.dumps(partial_payload), content_type="application/json")

        self.assertEqual(partial.status_code, 400)
        self.assertFalse(material.has_student_completed(self.student))

        complete_payload = {
            **partial_payload,
            "items_completed": 10,
            "scores": {"completed_phrases": list(range(10))},
        }
        complete = self.client.post(endpoint, data=json.dumps(complete_payload), content_type="application/json")

        self.assertEqual(complete.status_code, 200)
        self.assertTrue(complete.json()["success"])
        self.assertTrue(material.has_student_completed(self.student))
        result = material.assessment_results.get(student=self.student, attempt_status="completed")
        self.assertEqual(result.items_completed, 10)
        self.assertEqual(result.correct_items, 10)

        reopened = self.client.get(reverse("phrase_reading_page"), {"id": material.id})

        self.assertEqual(reopened.status_code, 200)
        restored = json.loads(reopened.context["phrase_reading_completion_json"])
        self.assertTrue(restored["completed"])
        self.assertEqual(restored["correct_items"], 10)
        self.assertEqual(restored["total_items"], 10)
        self.assertEqual(restored["total_score"], 100)
        self.assertEqual(restored["completed_phrases"], list(range(10)))
        self.assertEqual(material.assessment_results.filter(student=self.student).count(), 1)

    def test_repeated_assessment_completions_create_separate_assessment_rows(self):
        material = Material.objects.create(
            title="Retake Assessment",
            item_type="word",
            content_text="cat\ndog",
            content_json={"items": ["cat", "dog"]},
            type="assessment",
            status="published",
            is_active=True,
        )
        self._login_student()

        payload = json.dumps({
            "material_id": f"material-{material.id}",
            "activity_type": "assessment",
            "class_code": self.section.class_code,
        })

        first = self.client.post(reverse("record_assessment_completion"), data=payload, content_type="application/json")
        second = self.client.post(reverse("record_assessment_completion"), data=payload, content_type="application/json")

        self.assertTrue(first.json()["success"])
        self.assertTrue(second.json()["success"])

        material.refresh_from_db()
        self.assertIsNotNone(material.assessment)
        self.assertGreaterEqual(Assessment.objects.filter(code__startswith=material.assessment.code).count(), 2)
        attempts = material.assessment.get_attempts(self.student)
        self.assertEqual(len(attempts), 2)
        self.assertEqual(attempts[0]["attempt_number"], 1)
        self.assertEqual(attempts[1]["attempt_number"], 2)
        self.assertNotEqual(attempts[0]["attempt_id"], attempts[1]["attempt_id"])

    def test_assessment_completion_creates_teacher_notification(self):
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "assessment_id": f"assessment-{self.assessment.id}",
                "material_id": f"assessment-{self.assessment.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        notification = Notification.objects.filter(
            recipient=self.teacher,
            created_by=self.student,
            notification_type="assessment",
        ).first()
        self.assertIsNotNone(notification)
        self.assertIn("Jane Doe completed the assessment 'Reading Fluency Test' in Class A.", notification.message)
        self.assertFalse(notification.is_read)

    def test_assessment_completion_saves_scores_and_crla_classification(self):
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "assessment_id": f"assessment-{self.assessment.id}",
                "material_id": f"assessment-{self.assessment.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
                "scores": {
                    "fluency_score": 90,
                    "accuracy": 88,
                    "pronunciation_score": 86,
                    "time_score": 94,
                    "total_score": 89.5,
                    "wpm": 72,
                    "duration_seconds": 15,
                    "word_count": 18,
                    "transcript": "cat dog",
                    "speech_recognition_used": True,
                },
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        self.assessment.refresh_from_db()
        attempt = self.assessment.get_attempts()[-1]
        self.assertEqual(attempt["fluency_score"], 90)
        self.assertEqual(attempt["accuracy"], 88)
        self.assertEqual(attempt["pronunciation_score"], 86)
        self.assertEqual(attempt["time_score"], 94)
        self.assertEqual(attempt["total_score"], 89.5)
        self.assertEqual(attempt["crla_classification"], "Transitioning Readers")
        self.assertEqual(attempt["wpm"], 72)

        self.student.refresh_from_db()
        self.assertEqual(self.student.reading_level, "Transitioning Readers")
        profile = self.student.preference.get("student_profile", {})
        self.assertEqual(profile["accuracy"], "88")
        self.assertEqual(profile["wpm"], "72")
        self.assertEqual(profile["crla_classification"], "Transitioning Readers")

    def test_crla_completion_keeps_aral_eligible_students_on_completion_state(self):
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "assessment_id": f"assessment-{self.assessment.id}",
                "material_id": f"assessment-{self.assessment.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
                "scores": {
                    "fluency_score": 90,
                    "accuracy": 88,
                    "pronunciation_score": 86,
                    "time_score": 94,
                    "total_score": 89.5,
                    "wpm": 72,
                    "duration_seconds": 15,
                    "word_count": 18,
                },
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertNotIn("redirect_url", payload)
        self.student.refresh_from_db()
        self.assertEqual(
            self.student.preference.get("reading_assessment_state", {}).get("current_phase"),
            "materials",
        )

    def test_crla_completion_keeps_non_eligible_students_on_completion_state(self):
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "assessment_id": f"assessment-{self.assessment.id}",
                "material_id": f"assessment-{self.assessment.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
                "scores": {
                    "fluency_score": 18,
                    "accuracy": 20,
                    "pronunciation_score": 22,
                    "time_score": 24,
                    "total_score": 21,
                    "wpm": 12,
                    "duration_seconds": 15,
                    "word_count": 18,
                },
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload["success"])
        self.assertNotIn("redirect_url", payload)
        self.student.refresh_from_db()
        self.assertEqual(
            self.student.preference.get("reading_assessment_state", {}).get("current_phase"),
            "complete",
        )

    def test_numeric_material_id_records_assessment_attempt_by_assessment_id(self):
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": str(self.assessment.id),
                "activity_type": "assessment",
                "class_code": self.section.class_code,
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        self.assessment.refresh_from_db()
        self.assertEqual(len(self.assessment.get_attempts(self.student)), 1)
        self.assertEqual(self.assessment.get_student_attempt_count(self.student), 1)

    def test_duplicate_assessment_completion_records_multiple_attempts_with_unique_ids(self):
        self._login_student()
        payload = json.dumps({
            "assessment_id": f"assessment-{self.assessment.id}",
            "material_id": f"assessment-{self.assessment.id}",
            "activity_type": "assessment",
            "class_code": self.section.class_code,
        })

        first = self.client.post(
            reverse("record_assessment_completion"),
            data=payload,
            content_type="application/json",
        )
        second = self.client.post(
            reverse("record_assessment_completion"),
            data=payload,
            content_type="application/json",
        )

        self.assertTrue(first.json()["success"])
        self.assertTrue(second.json()["success"])

        self.assessment.refresh_from_db()
        attempts = self.assessment.get_attempts(self.student)
        self.assertEqual(len(attempts), 2)
        self.assertNotEqual(attempts[0]["attempt_id"], attempts[1]["attempt_id"])
        self.assertEqual(attempts[0]["attempt_number"], 1)
        self.assertEqual(attempts[1]["attempt_number"], 2)

    def test_teacher_update_material_does_not_create_assessment_record_when_none_exists(self):
        teacher = self.teacher
        material = Material.objects.create(
            title="Draft Assessment",
            item_type="word",
            content_text="cat\ndog",
            content_json={"items": ["cat", "dog"]},
            type="assessment",
            status="draft",
            is_active=False,
        )
        self.assertIsNone(material.assessment)

        session = self.client.session
        session["user_id"] = teacher.id
        session["user_role"] = teacher.role
        session.save()

        response = self.client.post(
            reverse("teacher_update_material"),
            data=json.dumps({
                "material_id": f"material-{material.id}",
                "title": "Draft Assessment Updated",
                "content": "cat dog",
                "status": "published",
                "usage_type": "assessment",
                "language": "English",
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        material.refresh_from_db()
        self.assertIsNone(material.assessment)
        self.assertEqual(Assessment.objects.filter(title="Draft Assessment Updated").count(), 0)

    def test_numeric_material_id_records_assessment_attempt_by_assessment_id(self):
        self._login_student()
        payload = json.dumps({
            "assessment_id": f"assessment-{self.assessment.id}",
            "material_id": f"assessment-{self.assessment.id}",
            "activity_type": "assessment",
            "class_code": self.section.class_code,
        })
        first = self.client.post(
            reverse("record_assessment_completion"),
            data=payload,
            content_type="application/json",
        )
        second = self.client.post(
            reverse("record_assessment_completion"),
            data=payload,
            content_type="application/json",
        )
        self.assertTrue(first.json()["success"])
        self.assertTrue(second.json()["success"])
        self.assertEqual(
            Notification.objects.filter(
                recipient=self.teacher,
                created_by=self.student,
                notification_type="assessment",
            ).count(),
            1,
        )

    def test_class_materials_only_marks_completed_after_completed_attempt(self):
        self._login_student()

        self.assessment.record_attempt(self.student, status="started")
        response = self.client.get(
            reverse("get_class_materials"),
            {"class_code": self.section.class_code},
        )
        self.assertEqual(response.status_code, 200)
        item = response.json()["materials"]["word"][0]
        self.assertEqual(item["attempt_count"], 1)
        self.assertEqual(item["completed_attempt_count"], 0)
        self.assertFalse(item["student_has_completed"])

        self.assessment.record_attempt(self.student, status="completed")
        response = self.client.get(
            reverse("get_class_materials"),
            {"class_code": self.section.class_code},
        )
        self.assertEqual(response.status_code, 200)
        item = response.json()["materials"]["word"][0]
        self.assertEqual(item["attempt_count"], 2)
        self.assertEqual(item["completed_attempt_count"], 1)
        self.assertTrue(item["student_has_completed"])

    def test_class_materials_include_latest_time_score_for_completed_attempt(self):
        self._login_student()

        self.assessment.record_attempt(
            self.student,
            status="completed",
            time_score=82,
            total_score=88,
        )
        response = self.client.get(
            reverse("get_class_materials"),
            {"class_code": self.section.class_code},
        )

        self.assertEqual(response.status_code, 200)
        item = response.json()["materials"]["word"][0]
        self.assertEqual(item["latest_time_score"], 82)
        self.assertEqual(item["latest_attempt_summary"]["time_score"], 82)

    def test_class_materials_do_not_duplicate_materials_with_assessment_rows(self):
        self._login_student()
        material = Material.objects.create(
            title="Duplicate Prevention Assessment",
            item_type="word",
            content_text="cat\ndog",
            content_json={"items": ["cat", "dog"]},
            type="assessment",
            status="published",
            section=self.section,
            is_active=True,
        )
        self.assertIsNone(material.assessment)

        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": f"material-{material.id}",
                "activity_type": "assessment",
                "class_code": self.section.class_code,
            }),
            content_type="application/json",
        )
        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        response = self.client.get(
            reverse("get_class_materials"),
            {"class_code": self.section.class_code},
        )
        self.assertEqual(response.status_code, 200)
        items = response.json()["materials"]["word"]
        self.assertEqual(len(items), 1)
        self.assertEqual(items[0]["id"], f"material-{material.id}")

    def test_teacher_can_fetch_unread_notification(self):
        Notification.objects.create(
            recipient=self.teacher,
            created_by=self.student,
            title="📝 Student Completed an Assessment",
            message="Jane Doe completed the assessment 'Reading Fluency Test' in Class A.",
            notification_type="assessment",
            action_url=f"/dashboard/teacher/students/detail/?student_id={self.student.custom_id}",
        )
        self._login_teacher()
        response = self.client.get(reverse("get_notifications"))
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(len(data["notifications"]), 1)
        self.assertFalse(data["notifications"][0]["is_read"])

    def test_mark_notification_read_endpoint_updates_state(self):
        notification = Notification.objects.create(
            recipient=self.teacher,
            created_by=self.student,
            title="Unread teacher update",
            message="A fresh notification for the shared panel.",
            notification_type="assessment",
        )
        self._login_teacher()

        response = self.client.post(
            reverse("mark_notification_read"),
            data=json.dumps({"notification_id": notification.id}),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        notification.refresh_from_db()
        self.assertTrue(notification.is_read)

    def test_notifications_page_uses_shared_mount_for_all_roles(self):
        for login in (
            self._login_admin,
            self._login_teacher,
            self._login_student,
            self._login_principal,
        ):
            login()
            response = self.client.get(reverse("notifications"))
            self.assertEqual(response.status_code, 200)
            self.assertContains(response, 'data-notifications-mount="page"', html=False)

    def test_practice_completion_notifies_admin_in_app_only(self):
        material = Material.objects.create(
            title="Practice Words",
            item_type="word",
            content_text="cat\ndog",
            content_json={"items": ["cat", "dog"]},
            status="published",
            is_active=True,
        )
        self._login_student()
        response = self.client.post(
            reverse("record_assessment_completion"),
            data=json.dumps({
                "material_id": f"material-{material.id}",
                "activity_type": "practice",
                "class_code": self.section.class_code,
            }),
            content_type="application/json",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])

        teacher_notification = Notification.objects.filter(
            recipient=self.teacher,
            created_by=self.student,
            notification_type="assessment",
        ).first()
        self.assertIsNone(teacher_notification)

        admin_notification = Notification.objects.filter(
            recipient=self.admin,
            created_by=self.student,
            notification_type="assessment",
        ).first()
        self.assertIsNotNone(admin_notification)
        self.assertIn("Jane Doe read \"Practice Words\"", admin_notification.message)
        self.assertFalse(admin_notification.is_read)


class WeeklyDigestTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id="TCH-DIGEST",
            role="teacher",
            first_name="Digest",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="digest-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
            preference={
                "notification_settings": {
                    "push_enabled": True,
                    "email_notifications": True,
                    "weekly_digest_enabled": True,
                }
            },
        )
        self.student = User.objects.create(
            custom_id="STD-DIGEST",
            role="student",
            first_name="Digest",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=2,
            birth_day=2,
            birth_year=2013,
            email="digest-student@example.com",
            password_hash=make_password("student-password"),
        )
        self.section = test_section_create(
            teacher=self.teacher,
            class_name="Digest Class",
            class_code="DIG-001",
            subject="Reading",
            is_active=True,
        )
        self.section.add_student(self.student)
        self.assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section,
            title="Digest Assessment",
            code="DIG-ASM-001",
            assessment_type="word",
            content="cat\ndog",
            is_active=True,
        )
        self.start = timezone.now() - timedelta(days=7)
        self.end = timezone.now() + timedelta(seconds=1)
        self.assessment.record_attempt(
            self.student,
            status="completed",
            completed_at=(timezone.now() - timedelta(days=1)).isoformat(),
            total_score=88,
            accuracy=87,
            fluency_score=86,
            pronunciation_score=85,
        )

    @patch("pabasa_app.weekly_digest.send_mail")
    def test_weekly_digest_skips_disabled_user(self, mock_send_mail):
        self.teacher.preference["notification_settings"]["weekly_digest_enabled"] = False
        self.teacher.save(update_fields=["preference", "updated_at"])

        result = send_weekly_digest(self.teacher, self.start, self.end)

        self.assertEqual(result["skipped"], "weekly_digest_disabled")
        mock_send_mail.assert_not_called()

    @patch("pabasa_app.weekly_digest.send_mail")
    def test_teacher_weekly_digest_sends_html_email_and_records_window(self, mock_send_mail):
        result = send_weekly_digest(self.teacher, self.start, self.end)

        self.assertTrue(result["sent"])
        mock_send_mail.assert_called_once()
        email_body = mock_send_mail.call_args[0][1]
        html_body = mock_send_mail.call_args.kwargs["html_message"]
        self.assertIn("Assessments completed by students: 1", email_body)
        self.assertIn("Average class reading performance: 88.0%", email_body)
        self.assertIn("<html", html_body)
        self.assertIn("Your Weekly PABASA Digest", html_body)
        self.assertIn("pabasalogo.png", html_body)
        self.assertIn("Assessments made", html_body)
        self.assertIn("Class average", html_body)

        self.teacher.refresh_from_db()
        digest_meta = self.teacher.preference["weekly_digest"]
        self.assertEqual(digest_meta["last_window_start"], self.start.isoformat())
        self.assertEqual(digest_meta["last_window_end"], self.end.isoformat())

        duplicate = send_weekly_digest(self.teacher, self.start, self.end)
        self.assertEqual(duplicate["skipped"], "duplicate_window")
        mock_send_mail.assert_called_once()

    @patch("pabasa_app.weekly_digest.send_mail")
    def test_student_weekly_digest_sends_html_email(self, mock_send_mail):
        self.student.preference = {
            "notification_settings": {
                "push_enabled": True,
                "email_notifications": True,
                "weekly_digest_enabled": True,
            }
        }
        self.student.save(update_fields=["preference", "updated_at"])

        result = send_weekly_digest(self.student, self.start, self.end)

        self.assertTrue(result["sent"])
        html_body = mock_send_mail.call_args.kwargs["html_message"]
        self.assertIn("Assessments done", html_body)
        self.assertIn("Practice sessions", html_body)
        self.assertIn("Best Assessment", html_body)
        self.assertIn("Pending Assessments", html_body)


class TeacherStudentsDirectoryTests(TestCase):
    def setUp(self):
        self.teacher = User.objects.create(
            custom_id="TCH-DIR1",
            role="teacher",
            first_name="Directory",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="directory-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        self.student = User.objects.create(
            custom_id="STD-DIR1",
            role="student",
            first_name="Single",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=2,
            birth_day=2,
            birth_year=2013,
            email="single-student@example.com",
            password_hash=make_password("student-password"),
        )
        self.section_a = test_section_create(
            teacher=self.teacher,
            class_name="Reading One",
            class_code="READ-ONE",
            subject="Reading",
            is_active=True,
        )
        self.section_b = test_section_create(
            teacher=self.teacher,
            class_name="Reading Two",
            class_code="READ-TWO",
            subject="Reading",
            is_active=True,
        )
        self.section_a.add_student(self.student)
        self.section_b.add_student(self.student)

        entries = self.section_b.get_enrolled_students()
        entries[0]["student_id"] = str(entries[0]["student_id"])
        self.section_b.students = entries
        self.section_b.save(update_fields=["students", "updated_at"])

        session = self.client.session
        session["user_id"] = self.teacher.id
        session["user_role"] = self.teacher.role
        session["first_name"] = self.teacher.first_name
        session["last_name"] = self.teacher.last_name
        session["email"] = self.teacher.email
        session["custom_id"] = self.teacher.custom_id
        session.save()

    def test_teacher_students_api_returns_one_row_with_joined_classes(self):
        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(len(data["students"]), 1)
        self.assertEqual(data["students"][0]["custom_id"], self.student.custom_id)
        self.assertCountEqual(data["students"][0]["classes"], ["Reading One", "Reading Two"])
        self.assertCountEqual(data["students"][0]["section_ids"], [self.section_a.id, self.section_b.id])

    def test_teacher_overview_counts_unique_students_across_classes(self):
        response = self.client.get(reverse("get_teacher_overview"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["classes_count"], 2)
        self.assertEqual(data["total_students"], 1)

    def test_teacher_students_api_uses_latest_assessment_classification(self):
        assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Oral Reading Check",
            code="ASM-DIR-001",
            assessment_type="paragraph",
            status="published",
            is_active=True,
        )
        assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            total_score=87,
        )

        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["students"][0]["level"], "Transitioning Readers")
        self.assertTrue(data["students"][0]["has_completed_assessment"])

    def test_teacher_students_api_prefers_persisted_crla_classification_over_score(self):
        assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Official CRLA",
            code="CRLA-DIR-AUTHORITY",
            assessment_type="paragraph",
            status="published",
            is_active=True,
        )
        Material.objects.create(
            assessment=assessment,
            teacher=self.teacher,
            section=self.section_a,
            title=assessment.title,
            code="CRLA-DIR-AUTHORITY-MAT",
            item_type="paragraph",
            type="assessment",
            assessment_kind="crla",
            status="published",
            is_active=True,
            is_official_reading=True,
        )
        assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            total_score=95,
        )
        attempt = Assessment.objects.get(source_assessment=assessment, student=self.student)
        attempt.crla_classification = "Low Emerging Readers"
        attempt.classification = "Low Emerging Readers"
        attempt.save(update_fields=["crla_classification", "classification", "updated_at"])

        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["students"][0]["level"], "Low Emerging Readers")
        self.assertEqual(data["level_counts"]["Low Emerging Readers"], 1)
        self.assertEqual(data["dashboard_metrics"]["grade_level_ready_count"], 0)

    def test_teacher_students_api_uses_adapted_level_from_attempt_history(self):
        word_assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Word Reading Check",
            code="ASM-DIR-002",
            assessment_type="word",
            status="published",
            is_active=True,
        )
        word_assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            total_score=56,
        )

        paragraph_assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Paragraph Reading Check",
            code="ASM-DIR-003",
            assessment_type="paragraph",
            status="published",
            is_active=True,
        )
        paragraph_assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-02T09:00:00+00:00",
            total_score=80,
        )

        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["students"][0]["adapted_reading_level"], "Developing")
        self.assertEqual(data["students"][0]["level"], "Transitioning Readers")

    def test_teacher_students_api_exposes_latest_completion_duration(self):
        assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Oral Reading Check",
            code="ASM-DIR-002",
            assessment_type="paragraph",
            status="published",
            is_active=True,
        )
        assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            total_score=87,
            duration_seconds=75,
        )

        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["students"][0]["duration_seconds"], 75)

    def test_teacher_students_api_returns_pending_without_assessment_data(self):
        response = self.client.get(reverse("get_teacher_students_api"))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data["students"][0]["level"], "Pending")
        self.assertFalse(data["students"][0]["has_completed_assessment"])
        self.assertIsNone(data["students"][0]["accuracy"])
        self.assertIsNone(data["students"][0]["wpm"])

    def test_crla_directory_uses_enrollment_and_source_material_for_all_term_phase_filters(self):
        system_owner = User.objects.create(
            custom_id="ADM-CRLA-DIR", role="admin", first_name="CRLA", last_name="System",
            middle_initial="", suffix="", sex="female", birth_month=1, birth_day=1,
            birth_year=1990, email="crla-directory-system@example.com",
            password_hash=make_password("system-password"),
        )
        root = Assessment.objects.create(
            teacher=system_owner, section=None, title="Term 2 Midline CRLA",
            code="CRLA-DIR-MID", assessment_type="paragraph", status="published",
            is_active=True, is_system_owned=True, system_assessment_period="midline",
            system_assessment_phase="midtest",
        )
        Material.objects.create(
            assessment=root, teacher=system_owner, title=root.title, code="CRLA-DIR-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
            source_type="personal", status="published", is_active=True,
            is_official_reading=True, is_system_owned=True, official_term=2,
            system_assessment_period="midline", system_assessment_phase="midtest",
        )
        root.record_attempt(
            self.student, status="completed", completed_at="2026-08-01T09:00:00+00:00",
            total_score=75, accuracy=75,
        )
        # Legacy completed rows can lack copied CRLA metadata. The persisted
        # source assessment/material remains authoritative.
        Assessment.objects.filter(source_assessment=root, student=self.student).update(
            is_system_owned=False, system_assessment_key='',
            system_assessment_period='', system_assessment_phase='', material=None,
        )

        for term in (1, 2, 3):
            for phase in ('pretest', 'midtest', 'posttest'):
                response = self.client.get(
                    reverse("get_teacher_students_api"),
                    {"term": term, "assessment": phase},
                )
                self.assertEqual(response.status_code, 200)
                students = response.json()["students"]
                if (term, phase) == (2, 'midtest'):
                    self.assertEqual([student["id"] for student in students], [self.student.id])
                    self.assertEqual(students[0]["assessment_id"], root.id)
                    self.assertTrue(students[0]["has_completed_assessment"])
                else:
                    self.assertEqual([student["id"] for student in students], [self.student.id])
                    self.assertEqual(students[0]["level"], "Pending")
                    self.assertFalse(students[0]["has_completed_assessment"])
                    self.assertIsNone(students[0]["assessment_id"])
                    self.assertIsNone(students[0]["accuracy"])
                    self.assertIsNone(students[0]["wpm"])

    def test_crla_directory_keeps_students_without_matching_attempts_pending(self):
        pending_student = User.objects.create(
            custom_id="STD-DIR2", role="student", first_name="Pending", last_name="Student",
            middle_initial="", suffix="", sex="male", birth_month=3, birth_day=3,
            birth_year=2013, email="pending-student@example.com",
            password_hash=make_password("student-password"),
        )
        self.section_a.add_student(pending_student)
        system_owner = User.objects.create(
            custom_id="ADM-CRLA-MIX", role="admin", first_name="CRLA", last_name="Owner",
            middle_initial="", suffix="", sex="female", birth_month=1, birth_day=1,
            birth_year=1990, email="crla-mixed-owner@example.com",
            password_hash=make_password("system-password"),
        )
        root = Assessment.objects.create(
            teacher=system_owner, title="Term 1 Pre-Test CRLA", code="CRLA-DIR-MIX",
            assessment_type="paragraph", status="published", is_active=True,
            is_system_owned=True, system_assessment_period="bosy",
            system_assessment_phase="pretest",
        )
        material = Material.objects.create(
            assessment=root, teacher=system_owner, title=root.title, code="CRLA-DIR-MIX-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
            source_type="personal", status="published", is_active=True,
            is_official_reading=True, is_system_owned=True, official_term=1,
            system_assessment_period="bosy", system_assessment_phase="pretest",
        )
        root.record_attempt(
            self.student, status="completed", completed_at="2026-06-01T09:00:00+00:00",
            total_score=82, accuracy=82,
        )
        Assessment.objects.filter(source_assessment=root, student=self.student).update(
            material=material, official_term=1,
        )

        response = self.client.get(
            reverse("get_teacher_students_api"),
            {"term": 1, "assessment": "pretest"},
        )

        self.assertEqual(response.status_code, 200)
        students = {student["id"]: student for student in response.json()["students"]}
        self.assertEqual(set(students), {self.student.id, pending_student.id})
        self.assertTrue(students[self.student.id]["has_completed_assessment"])
        self.assertNotEqual(students[self.student.id]["level"], "Pending")
        self.assertEqual(students[self.student.id]["accuracy"], 82)
        self.assertEqual(students[pending_student.id]["level"], "Pending")
        self.assertFalse(students[pending_student.id]["has_completed_assessment"])
        self.assertIsNone(students[pending_student.id]["accuracy"])
        self.assertIsNone(students[pending_student.id]["wpm"])

    def test_crla_directory_preserves_persisted_zero_accuracy_and_wpm(self):
        system_owner = User.objects.create(
            custom_id="ADM-CRLA-ZERO", role="admin", first_name="CRLA", last_name="Owner",
            middle_initial="", suffix="", sex="female", birth_month=1, birth_day=1,
            birth_year=1990, email="crla-zero-owner@example.com",
            password_hash=make_password("system-password"),
        )
        root = Assessment.objects.create(
            teacher=system_owner, title="Term 1 Zero CRLA", code="CRLA-DIR-ZERO",
            assessment_type="paragraph", status="published", is_active=True,
            is_system_owned=True, system_assessment_period="bosy",
            system_assessment_phase="pretest",
        )
        material = Material.objects.create(
            assessment=root, teacher=system_owner, title=root.title, code="CRLA-DIR-ZERO-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
            source_type="personal", status="published", is_active=True,
            is_official_reading=True, is_system_owned=True, official_term=1,
            system_assessment_period="bosy", system_assessment_phase="pretest",
        )
        root.record_attempt(
            self.student, status="completed", completed_at="2026-06-01T09:00:00+00:00",
            total_score=0, accuracy=0, wpm=0,
        )
        Assessment.objects.filter(source_assessment=root, student=self.student).update(
            material=material, official_term=1,
        )

        response = self.client.get(
            reverse("get_teacher_students_api"),
            {"term": 1, "assessment": "pretest"},
        )

        self.assertEqual(response.status_code, 200)
        student = response.json()["students"][0]
        self.assertTrue(student["has_completed_assessment"])
        self.assertEqual(student["accuracy"], 0)
        self.assertEqual(student["wpm"], 0)

    def test_official_crla_attempts_persist_calendar_term_independently_from_stage(self):
        calendar = SchoolCalendar.objects.create(
            school_year='2026-2027', current_term=1, is_active=True,
        )
        system_owner = User.objects.create(
            custom_id='ADM-CRLA-TERM', role='admin', first_name='CRLA', last_name='Owner',
            middle_initial='', suffix='', sex='female', birth_month=1, birth_day=1,
            birth_year=1990, email='crla-term-owner@example.com',
            password_hash=make_password('system-password'),
        )
        phase_config = (
            ('pretest', 'bosy', 'pre_assessment'),
            ('midtest', 'midline', 'midline_assessment'),
            ('posttest', 'eosy', 'post_assessment'),
        )

        expected = {}
        for term in (1, 2, 3):
            for offset, (phase, period, event_type) in enumerate(phase_config, start=1):
                taken_on = date(2026, term + 7, offset)
                CalendarEvent.objects.create(
                    school_calendar=calendar, term=term, title=f'Term {term} {phase}',
                    event_type=event_type, start_date=taken_on, end_date=taken_on,
                    is_published=True,
                )
                material = Material.objects.create(
                    teacher=system_owner, title=f'Term {term} {phase}',
                    code=f'CRLA-T{term}-{phase}', item_type='paragraph', type='assessment',
                    assessment_kind='crla', source_type='shared', status='published',
                    is_active=True, is_official_reading=True, is_system_owned=True,
                    system_assessment_period=period, system_assessment_phase=phase,
                )
                material.record_assessment_result(
                    self.student, status='completed',
                    completed_at=timezone.make_aware(datetime.combine(taken_on, datetime.min.time())),
                    total_score=80,
                )
                result = Assessment.objects.get(material=material, student=self.student)
                expected[(term, phase)] = result.official_term

        self.assertEqual(
            expected,
            {(term, phase): term for term in (1, 2, 3) for phase, _, _ in phase_config},
        )

    def test_teacher_course_assessments_api_includes_section_assigned_material_assessments(self):
        course = Course.objects.create(
            teacher=self.teacher,
            title="Course Assigned Material",
            code="CRS-MAT-001",
            description="Course with section-assigned assessment material",
        )
        course.sections.add(self.section_a)

        other_teacher = User.objects.create(
            custom_id="TCH-SHARED",
            role="teacher",
            first_name="Shared",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="shared-teacher@example.com",
            password_hash=make_password("shared-password"),
            teacher_role="Teacher",
        )
        material = Material.objects.create(
            title="Shared Assessment",
            teacher=other_teacher,
            item_type="paragraph",
            type="assessment",
            status="published",
            is_active=True,
        )
        material.assigned_sections.add(self.section_a)

        material.record_assessment_result(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            accuracy=82,
            wpm=65,
            fluency_score=78,
            pronunciation_score=80,
            time_score=85,
            total_score=83,
        )

        response = self.client.get(
            reverse("get_teacher_assessments_api"),
            {"course_id": course.id},
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )
        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(len(data["assessments"]), 1)
        self.assertEqual(data["assessments"][0]["title"], "Shared Assessment")
        self.assertEqual(data["assessments"][0]["attempt_count"], 1)
        self.assertIn("created_at", data["assessments"][0])
        self.assertIn("updated_at", data["assessments"][0])
        self.assertIsNotNone(data["assessments"][0]["updated_at"])

    def test_students_template_uses_static_renderer_only(self):
        response = self.client.get(reverse("students"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "pabasa_app/js/students.js", html=False)
        self.assertNotContains(response, "Students directory: prefer server")

    def test_students_export_card_reuses_existing_crla_export_endpoint(self):
        response = self.client.get(reverse("students"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn('/assessment/${encodeURIComponent(assessmentId)}/export-crla/?source=student-directory', content)
        self.assertNotIn('exportVisibleStudentsWorkbook', content)
        self.assertNotIn('XLSX.writeFile', content)

    def test_student_directory_can_download_system_crla_for_enrolled_completed_student(self):
        system_owner = User.objects.create(
            custom_id='ADM-CRLA-EXPORT', role='admin', first_name='CRLA', last_name='System',
            middle_initial='', suffix='', sex='female', birth_month=1, birth_day=1,
            birth_year=1990, email='crla-export-system@example.com',
            password_hash=make_password('system-password'),
        )
        root = Assessment.objects.create(
            teacher=system_owner, title='Official CRLA Post', code='CRLA-EXPORT-ROOT',
            assessment_type='paragraph', status='published', is_active=True,
            is_system_owned=True, system_assessment_key='eosy_crla_posttest',
            system_assessment_period='eosy', system_assessment_phase='posttest',
        )
        root.record_attempt(
            self.student, status='completed', completed_at=timezone.now(), total_score=80,
        )

        normal_response = self.client.get(reverse('export_crla_assessment', args=[root.id]))
        self.assertEqual(normal_response.status_code, 403)

        response = self.client.get(
            reverse('export_crla_assessment', args=[root.id]),
            {'source': 'student-directory'},
        )
        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        self.assertIn('attachment; filename="CRLA_', response['Content-Disposition'])
        self.assertTrue(response.content.startswith(b'PK'))

    def test_course_detail_refresh_script_reloads_students_after_assessment_change(self):
        response = self.client.get(reverse("courses"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        self.assertIn("async function refreshOpenCourseAfterAssessmentChange", content)
        self.assertIn("loadCourseStudents(openCourseId)", content)
        self.assertIn("Could not refresh course students after assessment change", content)

    def test_material_export_generates_student_activity_results_workbook(self):
        teacher = User.objects.create(
            custom_id="TCH-MAT-EXP",
            role="teacher",
            first_name="Material",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="material-export-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        student_one = User.objects.create(
            custom_id="PABASA-0001",
            role="student",
            first_name="Juan",
            last_name="Dela Cruz",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=8,
            birth_day=15,
            birth_year=2014,
            email="juan@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
            section="Rizal",
        )
        student_two = User.objects.create(
            custom_id="PABASA-0002",
            role="student",
            first_name="Maria",
            last_name="Santos",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=9,
            birth_day=20,
            birth_year=2014,
            email="maria@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
            section="Rizal",
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Rizal",
            class_code="RIZ-001",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student_one)
        section.add_student(student_two)

        material = Material.objects.create(
            teacher=teacher,
            section=section,
            title="Phrase Reading",
            item_type="sentence",
            type="assessment",
            status="published",
            is_active=True,
            source_type='personal',
            content_json={'template_title': 'Phrase Reading'},
        )
        material.record_assessment_result(
            student_one,
            status="completed",
            completed_at="2026-08-30T09:00:00+00:00",
            correct_items=8,
            items_completed=10,
        )

        session = self.client.session
        session['user_id'] = teacher.id
        session['user_role'] = teacher.role
        session.save()

        response = self.client.get(reverse('export_material_results'), {'material_id': f'material-{material.id}'})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response['Content-Type'],
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.sheetnames[0], 'Activity Results')
        sheet = workbook.active
        self.assertEqual(sheet['A1'].value, 'PABASA — Student Activity Results')
        self.assertEqual(sheet['A2'].value, 'Activity: Phrase Reading')
        header = [cell.value for cell in sheet[4]]
        self.assertEqual(header[:10], [
            'Student Name',
            'PABASA ID',
            'Grade',
            'Section',
            'Activity Title',
            'Activity Type',
            'Score',
            'Percentage',
            'Status',
            'Date Completed',
        ])
        self.assertTrue(all(cell.font.bold for cell in sheet[4]))
        self.assertFalse(any(cell.font.bold for cell in sheet[5]))
        self.assertFalse(any(cell.font.bold for cell in sheet[6]))

        first_row_values = [cell.value for cell in sheet[5]]
        self.assertEqual(first_row_values[0], 'Juan Dela Cruz')
        self.assertEqual(first_row_values[1], 'PABASA-0001')
        self.assertEqual(first_row_values[4], 'Phrase Reading')
        self.assertEqual(first_row_values[6], '8/10')
        self.assertEqual(first_row_values[7], '80%')
        self.assertEqual(first_row_values[8], 'Completed')

        second_row_values = [cell.value for cell in sheet[6]]
        self.assertEqual(second_row_values[0], 'Maria Santos')
        self.assertEqual(second_row_values[8], 'Not Attempted')
        self.assertEqual(second_row_values[6], '—')
        self.assertEqual(second_row_values[7], '—')

    def test_material_export_humanizes_activity_type_slug_in_excel(self):
        teacher = User.objects.create(
            custom_id="TCH-ACT-TYPE",
            role="teacher",
            first_name="Type",
            last_name="Teacher",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=1990,
            email="activity-type-teacher@example.com",
            password_hash=make_password("teacher-password"),
            teacher_role="Teacher",
        )
        student = User.objects.create(
            custom_id="PABASA-0009",
            role="student",
            first_name="Sample",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="female",
            birth_month=1,
            birth_day=1,
            birth_year=2014,
            email="activity-type-student@example.com",
            password_hash=make_password("student-password"),
            grade_level="Grade 2",
            section="Rizal",
        )
        section = test_section_create(
            teacher=teacher,
            class_name="Rizal",
            class_code="RIZ-ACT",
            subject="Reading",
            is_active=True,
        )
        section.add_student(student)
        material = Material.objects.create(
            teacher=teacher,
            section=section,
            title="Sound Play",
            item_type="word",
            type="assessment",
            status="published",
            is_active=True,
            source_type='personal',
            content_json={'activity_type': 'syllable_blending'},
        )
        material.record_assessment_result(
            student,
            status="completed",
            completed_at="2026-08-30T09:00:00+00:00",
            correct_items=4,
            items_completed=5,
        )

        session = self.client.session
        session['user_id'] = teacher.id
        session['user_role'] = teacher.role
        session.save()

        response = self.client.get(reverse('export_material_results'), {'material_id': f'material-{material.id}'})

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content), data_only=False)
        self.assertEqual(workbook.active['F5'].value, 'Syllable Blending')

    def test_course_report_recipients_do_not_use_local_storage_students(self):
        response = self.client.get(reverse("courses"))

        self.assertEqual(response.status_code, 200)
        content = response.content.decode()
        helper_start = content.index("async function fetchStudentsForCourse(course)")
        helper_end = content.index("// Course-scoped Reports loader", helper_start)
        helper_body = content[helper_start:helper_end]

        self.assertIn("/dashboard/teacher/students-api/", helper_body)
        self.assertNotIn("pabasa_added_students", helper_body)

    def test_course_update_composer_includes_report_preview_container(self):
        response = self.client.get(reverse("courses"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, 'id="courseSelectedStudentReportPreview"', html=False)

    @patch("pabasa_app.views.EmailMultiAlternatives")
    def test_send_course_update_emails_student_and_stores_note(self, mock_email_cls):
        course = Course.objects.create(
            teacher=self.teacher,
            title="Chapter 2",
            code="CRS-TEST-001",
            description="Course update test",
        )
        course.sections.add(self.section_a)
        assessment = Assessment.objects.create(
            teacher=self.teacher,
            section=self.section_a,
            title="Oral Reading Check",
            code="ASM-COURSE-001",
            assessment_type="paragraph",
            status="published",
            is_active=True,
        )
        assessment.record_attempt(
            self.student,
            status="completed",
            completed_at="2026-06-01T09:00:00+00:00",
            accuracy=88,
            wpm=72,
            fluency_score=84,
            pronunciation_score=86,
            time_score=90,
            total_score=87,
            crla_classification="Transitioning Readers",
        )

        response = self.client.post(
            reverse("send_course_update"),
            data=json.dumps({
                "course_id": course.id,
                "student_ids": [self.student.id],
                "update_type": "general",
                "message": "Hello {name}, keep practicing.",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["sent_count"], 1)
        self.assertTrue(data["report_included"])
        self.assertIn("report_summary", data["sent"][0])
        mock_email_cls.assert_called_once()
        self.assertEqual(mock_email_cls.call_args[0][0], "Student Reading Progress Report – PABASA")
        email_instance = mock_email_cls.return_value
        email_body = mock_email_cls.call_args[0][1]
        self.assertIn("Dear Parent/Guardian", email_body)
        self.assertIn("Attached is the latest Reading Progress Report", email_body)
        self.assertIn("PABASA Team", email_body)
        self.assertNotIn("Reading Performance Report", email_body)
        self.assertNotIn("Accuracy: 88%", email_body)
        self.assertNotIn("Words Per Minute: 72 WPM", email_body)
        email_instance.attach.assert_called_once()
        self.assertEqual(email_instance.attach.call_args[0][0], "Single_Student_reading_report.pdf")
        self.assertEqual(email_instance.attach.call_args[0][2], "application/pdf")
        email_instance.send.assert_called_once_with(fail_silently=False)

        note = Note.objects.get(teacher=self.teacher, student=self.student)
        self.assertEqual(note.note_type, "course_update:general")
        self.assertIn("Chapter 2", note.note_text)
        self.assertIn("Hello Single Student, keep practicing.", note.note_text)
        self.assertIn("Reading Performance Report", note.note_text)
        self.assertIn("Suggested Home Support", note.note_text)

    @patch("pabasa_app.views.EmailMultiAlternatives")
    def test_send_course_update_includes_baseline_message_when_metrics_missing(self, mock_email_cls):
        course = Course.objects.create(
            teacher=self.teacher,
            title="Chapter 3",
            code="CRS-TEST-002",
            description="Missing metrics test",
        )
        course.sections.add(self.section_a)

        response = self.client.post(
            reverse("send_course_update"),
            data=json.dumps({
                "course_id": course.id,
                "student_ids": [self.student.id],
                "update_type": "followup",
                "message": "Hello {name}, we will check your baseline soon.",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertEqual(mock_email_cls.call_args[0][0], "Student Reading Progress Report – PABASA")
        email_body = mock_email_cls.call_args[0][1]
        self.assertIn("Dear Parent/Guardian", email_body)
        self.assertIn("Attached is the latest Reading Progress Report", email_body)
        self.assertIn("PABASA Team", email_body)
        self.assertNotIn("No completed assessment yet", email_body)
        mock_email_cls.return_value.attach.assert_called_once()

        note = Note.objects.get(teacher=self.teacher, student=self.student)
        self.assertIn("No completed assessment yet", note.note_text)

    @patch("pabasa_app.views.EmailMultiAlternatives")
    def test_send_course_update_commendation_sends_certificate_attachment(self, mock_email_cls):
        course = Course.objects.create(
            teacher=self.teacher,
            title="Chapter 5",
            code="CRS-TEST-004",
            description="Commendation test",
        )
        course.sections.add(self.section_a)

        response = self.client.post(
            reverse("send_course_update"),
            data=json.dumps({
                "course_id": course.id,
                "student_ids": [self.student.id],
                "update_type": "commendation",
                "message": "Congratulations {name}!",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertFalse(response.json()["report_included"])
        self.assertEqual(mock_email_cls.call_args[0][0], "Performance Commendation – PABASA")
        email_body = mock_email_cls.call_args[0][1]
        self.assertIn("Congratulations", email_body)
        self.assertIn("certificate is attached", email_body.lower())
        self.assertIn("outstanding reading performance", email_body.lower())
        self.assertEqual(mock_email_cls.return_value.attach.call_count, 1)
        attachment_name, attachment_bytes, mime_type = mock_email_cls.return_value.attach.call_args.args
        self.assertIn("certificate", attachment_name.lower())
        self.assertEqual(mime_type, "application/pdf")
        self.assertTrue(attachment_bytes)

    @patch("pabasa_app.views.EmailMultiAlternatives")
    def test_send_course_update_assessment_notice_sends_details_without_attachment(self, mock_email_cls):
        course = Course.objects.create(
            teacher=self.teacher,
            title="Chapter 6",
            code="CRS-TEST-005",
            description="Assessment notice test",
        )
        course.sections.add(self.section_a)

        response = self.client.post(
            reverse("send_course_update"),
            data=json.dumps({
                "course_id": course.id,
                "student_ids": [self.student.id],
                "update_type": "assessment",
                "message": "Please prepare for your upcoming reading assessment.",
                "assessment_title": "Oral Reading Check",
                "scheduled_at": "2026-07-10 09:00",
                "reading_material": "The Little Red Hen",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        self.assertTrue(response.json()["success"])
        self.assertFalse(response.json()["report_included"])
        self.assertEqual(mock_email_cls.call_args[0][0], "Scheduled Assessment Notice – PABASA")
        email_body = mock_email_cls.call_args[0][1]
        self.assertIn("Oral Reading Check", email_body)
        self.assertIn("July 10, 2026 at 09:00 AM", email_body)
        self.assertIn("The Little Red Hen", email_body)
        self.assertIn("Please prepare", email_body)
        mock_email_cls.return_value.attach.assert_not_called()

    @patch("pabasa_app.views.EmailMultiAlternatives")
    def test_send_course_update_skips_unenrolled_and_missing_email_students(self, mock_email_cls):
        no_email_student = User.objects.create(
            custom_id="STD-NOEMAIL",
            role="student",
            first_name="No",
            last_name="Email",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=3,
            birth_day=3,
            birth_year=2013,
            email="",
            password_hash=make_password("student-password"),
        )
        outsider = User.objects.create(
            custom_id="STD-OUTSIDE",
            role="student",
            first_name="Outside",
            last_name="Student",
            middle_initial="",
            suffix="",
            sex="male",
            birth_month=4,
            birth_day=4,
            birth_year=2013,
            email="outside@example.com",
            password_hash=make_password("student-password"),
        )
        self.section_a.add_student(no_email_student)
        course = Course.objects.create(
            teacher=self.teacher,
            title="Chapter 4",
            code="CRS-TEST-003",
            description="Skipped recipient test",
        )
        course.sections.add(self.section_a)

        response = self.client.post(
            reverse("send_course_update"),
            data=json.dumps({
                "course_id": course.id,
                "student_ids": [self.student.id, no_email_student.id, outsider.id],
                "update_type": "general",
                "message": "Hello {name}, keep reading.",
            }),
            content_type="application/json",
            HTTP_X_REQUESTED_WITH="XMLHttpRequest",
        )

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertTrue(data["success"])
        self.assertEqual(data["sent_count"], 1)
        self.assertEqual(len(data["skipped"]), 2)
        self.assertCountEqual([item["reason"] for item in data["skipped"]], ["missing_email", "not_enrolled"])
        mock_email_cls.assert_called_once()


class AdminSingleSchoolTests(TestCase):
    def setUp(self):
        self.admin = User.objects.create(
            custom_id='ADM-SALAWAG', role='admin', first_name='Admin', last_name='User',
            sex='female', birth_month=1, birth_day=1, birth_year=1990,
            email='admin.salawag@example.com', password_hash=make_password('password'),
        )
        self.salawag = _primary_school()
        self.other_school = School.objects.create(name='Other Elementary School', code='OTHER-ES')
        session = self.client.session
        session.update({'user_id': self.admin.id, 'user_role': 'admin', 'custom_id': self.admin.custom_id})
        session.save()

    def test_school_list_route_redirects_to_salawag_detail(self):
        response = self.client.get(reverse('admin_school'))

        self.assertRedirects(
            response,
            reverse('admin_school_detail', args=[self.salawag.id]),
            fetch_redirect_response=False,
        )

    def test_school_creation_post_is_disabled_without_modifying_records(self):
        school_count = School.objects.count()

        response = self.client.post(reverse('admin_school'), {
            'school_name': 'Another School', 'school_code': 'ANOTHER-ES',
        })

        self.assertEqual(response.status_code, 405)
        self.assertEqual(School.objects.count(), school_count)
        self.assertTrue(School.objects.filter(pk=self.salawag.pk).exists())

    def test_only_salawag_detail_workspace_is_available(self):
        response = self.client.get(reverse('admin_school_detail', args=[self.salawag.id]))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, '<h1 class="h3 fw-bold mb-0">School</h1>', html=True)
        self.assertContains(response, 'School ID')
        self.assertContains(response, '107912')
        self.assertContains(response, '4114 Paliparan Road Dasmariñas Calabarzon')
        self.assertContains(response, 'Salawag Elementary School')
        self.assertContains(response, 'href="/dashboard/admin/school/"')
        self.assertContains(response, '>School<')
        self.salawag.refresh_from_db()
        self.assertEqual(self.salawag.name, 'Salawag Elementary School')
        self.assertEqual(self.salawag.code, '107912')
        self.assertEqual(self.salawag.address, '4114 Paliparan Road Dasmariñas Calabarzon')
        self.assertEqual(School.objects.filter(name='Salawag Elementary School').count(), 1)
        self.assertEqual(
            self.client.get(reverse('admin_school_detail', args=[self.other_school.id])).status_code,
            404,
        )

    def test_school_displays_predefined_grade_two_sections_without_manual_add_form(self):
        active_calendar = SchoolCalendar.objects.create(
            school_year='2026-2027', current_term=1, is_active=True,
        )

        response = self.client.get(reverse('admin_school_detail', args=[self.salawag.id]))

        expected = [
            'AGUINALDO', 'ALONZO', 'AQUINO', 'BALAGTAS', 'BALTAZAR', 'BONIFACIO',
            'DAGOHOY', 'DEL PILAR', 'ESCODA', 'JACINTO', 'LAPU-LAPU', 'LUNA',
            'MABINI', 'MAGSAYSAY', 'MALVAR', 'RICARTE', 'RIZAL', 'SAKAY',
        ]
        self.assertNotContains(response, 'Add Section')
        self.assertNotContains(response, 'name="section" class="form-control"')
        self.assertEqual(
            list(Section.objects.filter(
                school=self.salawag, school_calendar=active_calendar, grade_level='Grade 2',
            ).order_by('section').values_list('section', flat=True)),
            sorted(expected),
        )
        for section_name in expected:
            self.assertContains(response, section_name)

    def test_signup_uses_the_same_predefined_sections_for_students_and_teachers(self):
        SchoolCalendar.objects.create(school_year='2026-2027', current_term=1, is_active=True)
        expected = [
            'AGUINALDO', 'ALONZO', 'AQUINO', 'BALAGTAS', 'BALTAZAR', 'BONIFACIO',
            'DAGOHOY', 'DEL PILAR', 'ESCODA', 'JACINTO', 'LAPU-LAPU', 'LUNA',
            'MABINI', 'MAGSAYSAY', 'MALVAR', 'RICARTE', 'RIZAL', 'SAKAY',
        ]

        student_response = self.client.get(reverse('signup_sections'), {
            'role': 'student', 'grade_level': 'Grade 2',
        })
        teacher_response = self.client.get(reverse('signup_sections'), {
            'role': 'teacher', 'grade_level': 'Grade 2',
        })

        self.assertEqual(student_response.status_code, 200)
        self.assertEqual(teacher_response.status_code, 200)
        self.assertEqual([item['section'] for item in student_response.json()['sections']], expected)
        self.assertEqual([item['section'] for item in teacher_response.json()['sections']], expected)

    def test_signup_sections_only_returns_active_school_year_sections(self):
        old_calendar = SchoolCalendar.objects.create(
            school_year='2025-2026', current_term=3, is_active=False,
        )
        active_calendar = SchoolCalendar.objects.create(
            school_year='2026-2027', current_term=1, is_active=True,
        )
        old_section = Section.objects.create(
            school=self.salawag, school_calendar=old_calendar, class_code='OLDY-001',
            class_name='Grade 2 - Orchid', subject='Reading', grade_level='Grade 2', section='ORCHID',
        )

        response = self.client.get(reverse('signup_sections'), {'role': 'student', 'grade_level': 'Grade 2'})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload['school_year'], active_calendar.school_year)
        self.assertEqual(len(payload['sections']), 18)
        self.assertNotIn(old_section.id, [section['id'] for section in payload['sections']])

    def test_signup_sections_requires_an_active_school_year(self):
        SchoolCalendar.objects.create(school_year='2025-2026', current_term=3, is_active=False)

        response = self.client.get(reverse('signup_sections'), {'role': 'student'})

        self.assertEqual(response.status_code, 409)
        self.assertIn('No active School Year', response.json()['error'])

    @patch('pabasa_app.views._notify_admins')
    @patch('pabasa_app.views.send_student_confirmation_email')
    @patch('pabasa_app.views.send_student_signup_otp_email')
    def test_student_signup_assigns_grade_two_and_active_school_year(self, mock_otp, mock_confirmation, mock_notify):
        active_calendar = SchoolCalendar.objects.create(
            school_year='2026-2027', current_term=1, is_active=True,
        )
        section = Section.objects.create(
            school=self.salawag, school_calendar=active_calendar, class_code='STSY-001',
            class_name='Grade 2 - Sampaguita', subject='Reading', grade_level='Grade 2', section='SAMPAGUITA',
        )
        payload = {
            'first_name': 'Mia', 'last_name': 'Rivera', 'email': 'mia.calendar@example.com',
            'password': 'Student123', 'confirm_password': 'Student123', 'lrn': '123456789012',
            'sex': 'female', 'birth_month': '1', 'birth_day': '5', 'birth_year': '2014',
            'section': section.id,
        }

        response = self.client.post(reverse('register_student'), payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.client.session['pending_student_signup']['grade_level'], 'Grade 2')

        response = self.client.post(reverse('verify_student_otp'), {
            'otp': self.client.session['pending_student_signup_otp'],
        })
        self.assertEqual(response.status_code, 200)
        student = User.objects.get(email=payload['email'])
        self.assertEqual(student.grade_level, 'Grade 2')
        self.assertEqual(student.school_calendar, active_calendar)

    @patch('pabasa_app.views.send_teacher_confirmation_email')
    @patch('pabasa_app.views.send_teacher_signup_otp_email')
    def test_teacher_signup_assigns_active_school_year(self, mock_otp, mock_confirmation):
        active_calendar = SchoolCalendar.objects.create(
            school_year='2026-2027', current_term=1, is_active=True,
        )
        section = Section.objects.create(
            school=self.salawag, school_calendar=active_calendar, class_code='TCSY-001',
            class_name='Grade 2 - Rosal', subject='Reading', grade_level='Grade 2', section='ROSAL',
        )
        payload = {
            'first_name': 'Tina', 'last_name': 'Teacher', 'email': 'tina.calendar@example.com',
            'password': 'Teacher123', 'confirm_password': 'Teacher123', 'sex': 'female',
            'birth_month': '1', 'birth_day': '5', 'birth_year': '1990',
            'section': section.id, 'department': 'Mathematics',
        }

        response = self.client.post(reverse('register_teacher'), payload)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(self.client.session['pending_teacher_signup']['grade_level'], 'Grade 2')

        response = self.client.post(reverse('verify_teacher_otp'), {
            'otp': self.client.session['pending_teacher_signup_otp'],
        })
        self.assertEqual(response.status_code, 200)
        teacher = User.objects.get(email=payload['email'])
        self.assertEqual(teacher.school_calendar, active_calendar)
        section.refresh_from_db()
        self.assertEqual(section.teacher, teacher)
