from io import BytesIO

from django.contrib.auth.hashers import make_password
from django.test import TestCase
from django.utils import timezone
from openpyxl import load_workbook
import uuid

from .models import Assessment, Material, School, Section, StoryReadingProgress, User
from .scoring import crla_sentence_score
from .utils.crla_export import _part_1_reading_level, _row_formulas, _story_number, export_crla_excel


def test_section_create(**kwargs):
    school = kwargs.pop("school", None)
    if school is None:
        suffix = uuid.uuid4().hex.upper()
        school = School.objects.create(name=f"Fixture School {suffix}", code=f"FIXTURE-{suffix}")
    return Section.objects.create(school=school, **kwargs)


class CrlaExportResultTests(TestCase):
    def test_sentence_score_uses_official_four_sentence_table(self):
        self.assertEqual(
            [crla_sentence_score(count) for count in range(5)],
            [0, 3, 5, 7, 10],
        )

    def test_part_1_reading_level_uses_column_i_boundaries(self):
        expected = {
            0: "Full Refresher", 10: "Full Refresher",
            11: "Moderate Refresher", 16: "Moderate Refresher",
            17: "Light Refresher", 26: "Light Refresher",
            27: "Grade Ready", 30: "Grade Ready",
        }
        for total, level in expected.items():
            with self.subTest(total=total):
                self.assertEqual(_part_1_reading_level(total), level)
        self.assertIsNone(_part_1_reading_level(None))

    def test_generated_formula_logic_matches_official_crla_ranges(self):
        cases = (
            (5, 4, None, 9, "Full Refresher"),
            (6, 5, None, 11, "Moderate Refresher"),
            (10, None, 4, 14, "Moderate Refresher"),
            (10, None, 7, 17, "Light Refresher"),
            (10, None, 10, 20, "Light Refresher"),
            (10, None, 17, 27, "Grade Ready"),
        )
        formulas = _row_formulas(20)
        self.assertEqual(
            formulas["I"],
            '=IF(AND(F20="",G20="",H20=""),"",SUM(F20:H20))',
        )
        self.assertEqual(
            formulas["J"],
            '=IF(I20="","",IF(I20<=10,"Full Refresher",IF(I20<17,"Moderate Refresher",IF(I20<27,"Light Refresher","Grade Ready"))))',
        )
        self.assertNotIn("P", formulas)
        self.assertNotIn("Q", formulas)
        for task1, rhymes, sentences, total, level in cases:
            with self.subTest(task1=task1, rhymes=rhymes, sentences=sentences):
                self.assertEqual(task1 + (rhymes or 0) + (sentences or 0), total)
                self.assertEqual(
                    "Full Refresher" if total <= 10 else
                    "Moderate Refresher" if total < 17 else
                    "Light Refresher" if total < 27 else "Grade Ready",
                    level,
                )

    def test_story_number_uses_persisted_selected_story_title(self):
        self.assertEqual(_story_number({}, {"selected_story": "Si Pagong at Kuneho"}), 1)
        self.assertEqual(_story_number({}, {"selected_story": "Isang Kakaibang Araw"}), 2)
        self.assertIsNone(_story_number({}, {"selected_story": "Unknown Story"}))

    def make_user(self, custom_id, role, first_name, last_name, **extra):
        return User.objects.create(
            custom_id=custom_id,
            role=role,
            first_name=first_name,
            last_name=last_name,
            middle_initial="",
            suffix="",
            sex=extra.pop("sex", "female"),
            birth_month=1,
            birth_day=1,
            birth_year=2018 if role == "student" else 1990,
            email=f"{custom_id.lower()}@example.com",
            password_hash=make_password("password"),
            **extra,
        )

    def test_export_uses_assigned_teacher_and_persisted_story_two_results(self):
        admin = self.make_user("CRLA-ADMIN", "admin", "PABASA", "Admin")
        teacher = self.make_user("CRLA-TEACHER", "teacher", "Maria", "Santos", school="Mabini School")
        student = self.make_user("CRLA-STUDENT", "student", "Lina", "Reyes", lrn="123456789012")
        section = test_section_create(
            class_code="G2-RIZAL", class_name="Grade 2 Rizal", teacher=teacher,
            subject="Filipino", students=[{"student_id": student.id, "is_active": True}],
        )
        root = Assessment.objects.create(
            teacher=admin, title="Official CRLA", code="CRLA-EXPORT-VALUES",
            assessment_type="paragraph", status="published", is_system_owned=True,
            system_assessment_key="eosy_crla_posttest",
        )
        material = Material.objects.create(
            assessment=root, section=section, teacher=admin, title="Official CRLA",
            code="CRLA-EXPORT-MATERIAL", item_type="paragraph", type="assessment",
            assessment_kind="crla", is_system_owned=True, is_official_reading=True,
            content_json={"passages": [{"title": "Story One"}, {"title": "Story Two"}]},
        )
        student.preference = {
            "reading_assessment_state": {
                "student_end_assessment_state": {
                    "material_id": str(material.id), "stage": "completed", "branch": "sentences",
                    "task1_score": 1, "task2_rhymes_score": 9, "task2_sentences_score": 29,
                    "part1_total_score": 10, "selected_story": "Story One", "total_words_read": 1,
                    "duration_seconds": 1, "wpm": 1, "story_read_percent": 1,
                    "correct_answers": 1, "learner_experience_rating": 4,
                    "classification": "Transitioning Reader",
                }
            }
        }
        student.save(update_fields=["preference", "updated_at"])
        Assessment.objects.create(
            teacher=teacher, section=section, material=material, source_assessment=root,
            student=student, title="CRLA result", code="CRLA-EXPORT-RESULT",
            assessment_type="paragraph", status="published", attempt_status="completed",
            completed_at=timezone.now(), duration_seconds=125, word_count=70, wpm=33.6,
            accuracy=70, correct_items=3, crla_score_data={
                "task1_score": 10, "task2_type": "Task 2H / Sentences", "task2_score": 4,
                "part1_total_score": 14, "story_number": None, "story_total_words": 96,
                "words_read": 96, "miscues": 0, "duration_seconds": 519.99, "wpm": 42,
                "passage_accuracy_percent": 80, "comprehension_total": 6,
                "comprehension_correct": 4,
            },
        )

        workbook = load_workbook(BytesIO(export_crla_excel(root.id).getvalue()), data_only=False)
        sheet = workbook["G2 MT Reading Scoresheet"]

        self.assertEqual(sheet["C6"].value, "Maria Santos")
        self.assertEqual(sheet["F11"].value, 10)
        self.assertIsNone(sheet["G11"].value)
        self.assertEqual(sheet["H11"].value, 10)
        self.assertTrue(str(sheet["I11"].value).startswith("="))
        self.assertTrue(str(sheet["J11"].value).startswith("="))
        self.assertIsNone(sheet["K11"].value)
        self.assertEqual(sheet["L11"].value, 0)
        self.assertEqual(sheet["M11"].value, 96)
        self.assertEqual((sheet["N11"].value, sheet["O11"].value), (8, 40))
        self.assertEqual(sheet["P11"].value, 42)
        self.assertEqual(sheet["Q11"].value, 0.8)
        self.assertEqual(sheet["R11"].value, 4)
        self.assertIsNone(sheet["S11"].value)
        self.assertEqual(sheet["T11"].value, "Level 3")
        self.assertEqual(sheet["U11"].value, "Transitioning Reader")
        self.assertEqual(sheet["V11"].value, "Needs continued reading practice")

    def test_low_emerging_branch_leaves_part_two_cells_blank(self):
        teacher = self.make_user("CRLA-T2", "teacher", "Ana", "Cruz")
        student = self.make_user("CRLA-S2", "student", "Nilo", "Dela Cruz")
        section = test_section_create(
            class_code="G2-BONI", class_name="Grade 2 Bonifacio", teacher=teacher,
            subject="Filipino", students=[{"student_id": student.id, "is_active": True}],
        )
        root = Assessment.objects.create(
            teacher=teacher, section=section, title="CRLA Early Exit", code="CRLA-EARLY",
            assessment_type="word", status="published",
        )
        material = Material.objects.create(
            assessment=root, section=section, teacher=teacher, code="CRLA-EARLY-MAT",
            item_type="word", type="assessment", assessment_kind="crla",
        )
        student.preference = {"reading_assessment_state": {"student_end_assessment_state": {
            "material_id": str(material.id), "stage": "early_completed_words", "branch": "rhymes",
            "task1_score": 1, "task2_rhymes_score": 9, "part1_total_score": 10,
            "classification": "Low Emerging Reader",
        }}}
        student.save(update_fields=["preference", "updated_at"])
        Assessment.objects.create(
            teacher=teacher, section=section, material=material, source_assessment=root,
            student=student, title="early result", code="CRLA-EARLY-RESULT",
            assessment_type="word", status="published", attempt_status="completed",
            completed_at=timezone.now(), duration_seconds=20, word_count=6, wpm=18, accuracy=60,
            crla_score_data={"task1_score": 6, "task2_type": "Task 2L / Rhymes", "task2_score": 3},
        )

        sheet = load_workbook(BytesIO(export_crla_excel(root.id).getvalue()), data_only=False)["G2 MT Reading Scoresheet"]
        self.assertEqual((sheet["F11"].value, sheet["G11"].value), (6, 3))
        self.assertIsNone(sheet["H11"].value)
        self.assertTrue(str(sheet["I11"].value).startswith("="))
        self.assertTrue(str(sheet["J11"].value).startswith("="))
        for column in ("K", "N", "O", "R", "S", "T"):
            self.assertIsNone(sheet[f"{column}11"].value)
        self.assertEqual(sheet["P11"].value, 18)
        self.assertIsNone(sheet["Q11"].value)
        self.assertEqual(sheet["U11"].value, "Low Emerging Reader")
        self.assertEqual(sheet["V11"].value, "Needs intensive reading intervention")

    def test_export_falls_back_to_persisted_state_for_reading_metrics(self):
        teacher = self.make_user("CRLA-STATE", "teacher", "Mila", "Garcia")
        student = self.make_user("CRLA-STATE-STUDENT", "student", "Tomas", "Reyes")
        section = test_section_create(
            class_code="G2-STATE", class_name="Grade 2 State", teacher=teacher,
            subject="Filipino", students=[{"student_id": student.id, "is_active": True}],
        )
        root = Assessment.objects.create(
            teacher=teacher, section=section, title="CRLA State Metrics", code="CRLA-STATE-METRICS",
            assessment_type="paragraph", status="published",
        )
        material = Material.objects.create(
            assessment=root, section=section, teacher=teacher, code="CRLA-STATE-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
        )
        student.preference = {"reading_assessment_state": {"student_end_assessment_state": {
            "material_id": str(material.id), "stage": "completed", "branch": "story",
            "task1_score": 8, "task2_sentences_score": 10, "selected_story": "Story One",
            "words_read": 72, "miscues": 4, "duration_seconds": 120, "wpm": 42,
            "passage_accuracy_percent": 80, "comprehension_correct": 5,
            "classification": "Reading At Grade Level",
        }}}
        student.save(update_fields=["preference", "updated_at"])
        Assessment.objects.create(
            teacher=teacher, section=section, material=material, source_assessment=root,
            student=student, title="state-only result", code="CRLA-STATE-RESULT",
            assessment_type="paragraph", status="published", attempt_status="completed",
            completed_at=timezone.now(), crla_score_data={},
        )

        sheet = load_workbook(BytesIO(export_crla_excel(root.id).getvalue()), data_only=False)["G2 MT Reading Scoresheet"]
        self.assertEqual(sheet["L11"].value, 4)
        self.assertEqual(sheet["M11"].value, 72)
        self.assertEqual((sheet["N11"].value, sheet["O11"].value), (2, 0))
        self.assertEqual(sheet["P11"].value, 42)
        self.assertEqual(sheet["Q11"].value, 0.8)

    def test_export_preserves_zero_wpm_and_accuracy(self):
        teacher = self.make_user("CRLA-ZERO", "teacher", "Zero", "Values")
        student = self.make_user("CRLA-ZERO-STUDENT", "student", "Zero", "Reader")
        root = Assessment.objects.create(
            teacher=teacher, title="CRLA Zero Metrics", code="CRLA-ZERO-METRICS",
            assessment_type="paragraph", status="published",
        )
        material = Material.objects.create(
            assessment=root, teacher=teacher, code="CRLA-ZERO-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
        )
        student.preference = {"reading_assessment_state": {"student_end_assessment_state": {
            "material_id": str(material.id), "words_read": 0, "miscues": 0,
            "duration_seconds": 0, "wpm": 0, "correct_words_percentage": 0,
        }}}
        student.save(update_fields=["preference", "updated_at"])
        Assessment.objects.create(
            teacher=teacher, material=material, source_assessment=root,
            student=student, title="zero result", code="CRLA-ZERO-RESULT",
            assessment_type="paragraph", status="published", attempt_status="completed",
            completed_at=timezone.now(), crla_score_data={},
        )

        sheet = load_workbook(BytesIO(export_crla_excel(root.id).getvalue()), data_only=False)["G2 MT Reading Scoresheet"]
        self.assertEqual(sheet["L11"].value, 0)
        self.assertEqual(sheet["M11"].value, 0)
        self.assertEqual(sheet["P11"].value, 0)
        self.assertEqual(sheet["Q11"].value, 0)

    def test_export_keeps_official_formula_cells_and_persists_story_miscues(self):
        teacher = self.make_user("CRLA-T3", "teacher", "Luz", "Delgado")
        student = self.make_user("CRLA-S3", "student", "Rosa", "Santos")
        section = test_section_create(
            class_code="G2-SAMPAGUITA", class_name="Grade 2 Sampaguita", teacher=teacher,
            subject="Filipino", students=[{"student_id": student.id, "is_active": True}],
        )
        root = Assessment.objects.create(
            teacher=teacher, section=section, title="CRLA Story Export", code="CRLA-STORY-EXPORT",
            assessment_type="paragraph", status="published",
        )
        material = Material.objects.create(
            assessment=root, section=section, teacher=teacher, code="CRLA-STORY-MAT",
            item_type="paragraph", type="assessment", assessment_kind="crla",
            content_json={"passages": [{"title": "Story One"}, {"title": "Story Two"}]},
        )
        student.preference = {"reading_assessment_state": {"student_end_assessment_state": {
            "material_id": str(material.id), "stage": "completed", "branch": "story",
            "task1_score": 1, "task2_rhymes_score": 9, "task2_sentences_score": 29,
            "part1_total_score": 10, "selected_story": "Story One", "total_words_read": 1, "miscues": 99,
            "duration_seconds": 1, "wpm": 1, "story_read_percent": 1,
            "correct_answers": 1, "classification": "Transitioning Reader",
        }}}
        student.save(update_fields=["preference", "updated_at"])
        Assessment.objects.create(
            teacher=teacher, section=section, material=material, source_assessment=root,
            student=student, title="story result", code="CRLA-STORY-RESULT",
            assessment_type="paragraph", status="published", attempt_status="completed",
            completed_at=timezone.now(), duration_seconds=125, word_count=70, wpm=33.6,
            accuracy=70, correct_items=3, crla_score_data={
                "task1_score": 8, "task2_type": "Task 2H / Sentences", "task2_score": 7,
                "story_number": 2, "words_read": 70, "miscues": 3,
                "duration_seconds": 125, "wpm": 33.6, "passage_accuracy_percent": 70,
                "comprehension_correct": 3,
            },
        )
        StoryReadingProgress.objects.create(
            student=student, material=material, story_title="Story Two", total_words=100,
            words_read=70, correct_words=70, miscues=3, accuracy=70, wpm=33.6,
            correct_sentences=3, duration_seconds=125, completed=True,
            completed_at=timezone.now(),
        )

        sheet = load_workbook(BytesIO(export_crla_excel(root.id).getvalue()), data_only=False)["G2 MT Reading Scoresheet"]
        self.assertEqual(sheet["F11"].value, 8)
        self.assertEqual(sheet["H11"].value, 10)
        self.assertTrue(str(sheet["I11"].value).startswith("="))
        self.assertTrue(str(sheet["J11"].value).startswith("="))
        self.assertEqual(sheet["K11"].value, 2)
        self.assertEqual(sheet["L11"].value, 3)
        self.assertEqual(sheet["M11"].value, 70)
        self.assertEqual((sheet["N11"].value, sheet["O11"].value), (2, 5))
        self.assertTrue(str(sheet["P11"].value).startswith("="))
        self.assertTrue(str(sheet["Q11"].value).startswith("="))
        self.assertEqual(sheet["R11"].value, 3)
